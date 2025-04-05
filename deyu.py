#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
德育维度模块，包含德育维度相关的API路由和功能
"""

import os
import json
import sqlite3
import traceback
import datetime
from flask import Blueprint, request, jsonify, render_template, redirect, url_for, current_app as app, session, send_file, send_from_directory
import openpyxl
from werkzeug.utils import secure_filename
from flask_login import login_required, current_user
from utils.class_filter import class_filter, user_can_access
from utils.grades_manager import GradesManager
import logging
import pandas as pd
import numpy as np
import glob
import io
import time
from markupsafe import escape
import uuid
import shutil

# 配置日志
logger = logging.getLogger(__name__)

# 德育蓝图
deyu_bp = Blueprint('deyu', __name__)

# 配置
UPLOAD_FOLDER = 'uploads'
DATABASE = 'students.db'
TEMPLATE_FOLDER = 'templates'  # 模板文件夹

# 初始化德育模块
def init_deyu(app):
    """初始化德育维度模块"""
    logger.info("初始化德育维度模块")
    # 确保students表包含德育维度相关字段
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    # 检查是否所有必要的列都存在，如果不存在则添加
    cursor.execute("PRAGMA table_info(students)")
    existing_columns = [row[1] for row in cursor.fetchall()]
    logger.info(f"现有数据库列: {existing_columns}")
    
    # 定义应该存在的德育维度列及其类型
    deyu_columns = {
        'pinzhi': 'INTEGER',  # 品质 30分
        'xuexi': 'INTEGER',   # 学习 20分
        'jiankang': 'INTEGER', # 健康 20分
        'shenmei': 'INTEGER',  # 审美 10分
        'shijian': 'INTEGER',  # 实践 10分
        'shenghuo': 'INTEGER'  # 生活 10分
    }
    
    # 添加缺失的列
    for column, col_type in deyu_columns.items():
        if column not in existing_columns:
            logger.warning(f"添加缺失的德育维度列: {column} ({col_type})")
            try:
                cursor.execute(f"ALTER TABLE students ADD COLUMN {column} {col_type}")
                conn.commit()
                logger.info(f"成功添加德育维度列: {column}")
            except sqlite3.Error as e:
                logger.error(f"添加德育维度列 {column} 时出错: {e}")
    
    # 提交并关闭连接
    conn.commit()
    conn.close()
    
    logger.info("德育维度模块初始化完成")

# 创建成绩管理器实例
grades_manager = GradesManager()

# 创建成绩导入模板
try:
    grades_manager.create_empty_template()
    print("德育模板创建完成")
except Exception as e:
    print(f"创建德育模板时出错: {str(e)}")

# 提供模板下载
@deyu_bp.route('/download/template/<filename>', methods=['GET'])
@login_required
def serve_template(filename):
    return send_from_directory(TEMPLATE_FOLDER, filename)

# 获取数据库连接
def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')  # 启用外键约束
    return conn

# 获取所有学生成绩
@deyu_bp.route('/api/deyu', methods=['GET'])
@login_required
def get_all_deyu():
    try:
        class_id = request.args.get('class_id', '')
        semester = request.args.get('semester', '')
        
        # 详细的请求日志
        logger.info(f"收到获取德育维度数据请求: class_id={class_id}, semester={semester}")
        logger.info(f"当前用户: {current_user.username}, ID: {current_user.id}, 角色: {'管理员' if current_user.is_admin else '普通用户'}, 关联班级: {current_user.class_id}")
        logger.info(f"请求参数: {request.args}")
        
        # 班主任只能查看自己班级的学生德育成绩
        if not current_user.is_admin:
            if not current_user.class_id:
                logger.warning("非管理员用户未分配班级")
                return jsonify({
                    'status': 'ok',
                    'deyu': [],
                    'message': '您尚未被分配班级，无法查看学生德育维度'
                })
            class_id = current_user.class_id
            logger.info(f"非管理员用户，使用其班级ID: {class_id}")
        
        if not class_id:
            logger.warning("未提供班级ID")
            return jsonify({'status': 'error', 'message': '请提供班级ID'})
        
        # 记录使用的查询参数
        logger.info(f"使用的班级ID参数: {class_id}")
        
        # 直接从数据库获取学生信息
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 数据库连接检查
        if not conn or not cursor:
            logger.error("无法获取数据库连接")
            return jsonify({'status': 'error', 'message': '数据库连接失败'})
        
        # 检查数据库表和列是否存在
        try:
            cursor.execute("PRAGMA table_info(students)")
            columns = cursor.fetchall()
            column_names = [col[1] for col in columns]
            logger.info(f"students表列名: {column_names}")
            
            # 确认必要的列存在
            required_columns = ['id', 'name', 'class', 'class_id', 'pinzhi', 'xuexi', 'jiankang', 'shenmei', 'shijian', 'shenghuo']
            missing_columns = [col for col in required_columns if col not in column_names]
            
            if missing_columns:
                logger.error(f"students表缺少必要的列: {missing_columns}")
                return jsonify({'status': 'error', 'message': f'数据库表结构不完整，缺少列: {missing_columns}'})
            
            # 检查表中的记录总数
            cursor.execute("SELECT COUNT(*) FROM students")
            total_students = cursor.fetchone()[0]
            logger.info(f"students表总记录数: {total_students}")
            
        except sqlite3.Error as e:
            logger.error(f"检查数据库结构时出错: {e}")
            return jsonify({'status': 'error', 'message': f'数据库结构检查失败: {str(e)}'})
        
        # 查询班级所有学生信息，包括德育维度字段
        # 支持使用班级ID或班级名称查询
        students = []
        query_attempts = 0
        
        try:
            # 尝试将class_id转为整数，如果能转，就当作ID查询
            int_class_id = int(class_id)
            logger.info(f"使用班级ID查询: {int_class_id}")
            query_attempts += 1
            
            cursor.execute('''
                SELECT id, name, class, class_id, pinzhi, xuexi, jiankang, shenmei, shijian, shenghuo 
                FROM students 
                WHERE class_id = ?
                ORDER BY id
            ''', (int_class_id,))
            students = cursor.fetchall()
            logger.info(f"按班级ID查询结果: 找到{len(students)}名学生")
            
            # 记录查询到的首个学生数据样本（如果有）
            if students and len(students) > 0:
                logger.info(f"首个学生样本: {dict(students[0])}")
            
        except (ValueError, TypeError):
            # 如果不能转为整数，就当作班级名称查询
            logger.info(f"使用班级名称查询: {class_id}")
            query_attempts += 1
            
            cursor.execute('''
                SELECT id, name, class, class_id, pinzhi, xuexi, jiankang, shenmei, shijian, shenghuo 
                FROM students 
                WHERE class = ?
                ORDER BY id
            ''', (class_id,))
            students = cursor.fetchall()
            logger.info(f"按班级名称查询结果: 找到{len(students)}名学生")
        
        # 如果没有找到学生，尝试使用模糊匹配班级名称
        if not students:
            logger.info("未找到学生，尝试使用模糊匹配")
            query_attempts += 1
            
            cursor.execute('''
                SELECT id, name, class, class_id, pinzhi, xuexi, jiankang, shenmei, shijian, shenghuo 
                FROM students 
                WHERE class LIKE ?
                ORDER BY id
            ''', (f'%{class_id}%',))
            students = cursor.fetchall()
            logger.info(f"模糊匹配结果: 找到{len(students)}名学生")
            
            # 如果还是没找到学生，尝试直接获取所有学生（管理员用户）
            if not students and current_user.is_admin:
                logger.info("所有查询都未找到学生，管理员用户尝试获取所有学生")
                query_attempts += 1
                
                cursor.execute('''
                    SELECT id, name, class, class_id, pinzhi, xuexi, jiankang, shenmei, shijian, shenghuo 
                    FROM students 
                    ORDER BY class, id
                ''')
                students = cursor.fetchall()
                logger.info(f"获取所有学生结果: 找到{len(students)}名学生")
            
            # 如果仍然没找到学生，检查数据库中是否有任何学生
            if not students:
                logger.warning("所有查询方式都未找到学生，检查数据库中是否有任何学生")
                query_attempts += 1
                
                cursor.execute('SELECT COUNT(*) FROM students')
                total_students = cursor.fetchone()[0]
                logger.info(f"数据库中总学生数: {total_students}")
                
                # 检查班级表中是否有匹配的班级
                cursor.execute('SELECT id, class_name FROM classes WHERE id = ? OR class_name LIKE ?', 
                               (class_id, f'%{class_id}%'))
                matching_classes = cursor.fetchall()
                if matching_classes:
                    class_info = [dict(c) for c in matching_classes]
                    logger.info(f"找到匹配的班级: {class_info}")
                else:
                    logger.warning(f"未找到匹配班级ID/名称: {class_id}")
        
        logger.info(f"进行了{query_attempts}次查询尝试")
        conn.close()
        
        # 转换结果，更详细记录处理过程
        result = []
        logger.info(f"开始处理查询结果，转换 {len(students)} 名学生的德育维度数据")
        
        for student in students:
            student_data = dict(student)
            
            # 记录处理每个维度的详细信息
            deyu_fields = {}
            for field in ['pinzhi', 'xuexi', 'jiankang', 'shenmei', 'shijian', 'shenghuo']:
                field_value = student_data.get(field)
                deyu_fields[field] = field_value or 0  # 将None转为0
            
            result.append({
                'student_id': student_data['id'],
                'name': student_data['name'],
                'class': student_data['class'],
                'class_id': student_data['class_id'],
                'pinzhi': deyu_fields['pinzhi'],      # 品质
                'xuexi': deyu_fields['xuexi'],        # 学习
                'jiankang': deyu_fields['jiankang'],  # 健康
                'shenmei': deyu_fields['shenmei'],    # 审美
                'shijian': deyu_fields['shijian'],    # 实践
                'shenghuo': deyu_fields['shenghuo']   # 生活
            })
        
        # 添加结果样本日志
        if result:
            sample_size = min(3, len(result))
            logger.info(f"结果样本 ({sample_size} 条):")
            for i in range(sample_size):
                logger.info(f"  {result[i]}")
        
        logger.info(f"成功获取班级 {class_id} 的所有学生德育维度数据，共 {len(result)} 条记录")
        return jsonify({'status': 'ok', 'deyu': result})
    except Exception as e:
        logger.error(f'获取学生德育维度时出错: {str(e)}')
        logger.error(traceback.format_exc())
        return jsonify({'status': 'error', 'message': f'获取学生德育维度失败: {str(e)}'})

# 获取单个学生成绩
@deyu_bp.route('/api/deyu/<student_id>', methods=['GET'])
@login_required
def get_student_deyu(student_id):
    try:
        semester = request.args.get('semester', None)
        class_id = request.args.get('class_id', None)
        
        # 班主任自动使用自己的班级ID
        if not current_user.is_admin:
            if not current_user.class_id:
                return jsonify({'status': 'error', 'message': '您尚未被分配班级，无法查看学生德育维度'}), 403
            class_id = current_user.class_id
        elif not class_id:
            return jsonify({'status': 'error', 'message': '管理员需提供班级ID参数'}), 400
        
        # 直接从数据库获取学生信息
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 查询学生信息，包括德育维度字段
        cursor.execute('''
            SELECT id, name, class, pinzhi, xuexi, jiankang, shenmei, shijian, shenghuo 
            FROM students 
            WHERE id = ? AND class_id = ?
        ''', (student_id, class_id))
        
        student = cursor.fetchone()
        conn.close()
        
        if not student:
            logger.error(f"未找到学生，ID: {student_id}, 班级ID: {class_id}")
            return jsonify({'status': 'error', 'message': '未找到学生'}), 404
        
        # 转换为字典
        student_data = dict(student)
        
        # 包装为API响应格式
        result = {
            'student_id': student_data['id'],
            'name': student_data['name'],
            'class': student_data['class'],
            'pinzhi': student_data['pinzhi'] or 0,      # 品质
            'xuexi': student_data['xuexi'] or 0,        # 学习
            'jiankang': student_data['jiankang'] or 0,  # 健康
            'shenmei': student_data['shenmei'] or 0,    # 审美
            'shijian': student_data['shijian'] or 0,    # 实践
            'shenghuo': student_data['shenghuo'] or 0   # 生活
        }
        
        logger.info(f"成功获取学生 {student_id} 的德育维度数据")
        return jsonify({'status': 'ok', 'deyu': result})
    except Exception as e:
        logger.error(f'获取学生德育维度时出错: {str(e)}')
        return jsonify({'status': 'error', 'message': f'获取学生德育维度失败: {str(e)}'})

# 保存学生成绩
@deyu_bp.route('/api/deyu/<student_id>', methods=['POST'])
@login_required
def save_student_deyu(student_id):
    try:
        data = request.get_json()
        logger.info(f"收到保存德育维度请求，学生ID: {student_id}")
        logger.info(f"请求数据: {data}")
        logger.info(f"当前用户: {current_user.username}, ID: {current_user.id}, 角色: {'管理员' if current_user.is_admin else '普通用户'}, 关联班级: {current_user.class_id}")
        
        semester = data.get('semester', '上学期')
        class_id = data.get('class_id')
        
        # 班主任自动使用自己的班级ID
        if not current_user.is_admin:
            if not current_user.class_id:
                logger.error("非管理员用户未分配班级，无法保存")
                return jsonify({'status': 'error', 'message': '您尚未被分配班级，无法保存学生德育维度'}), 403
            class_id = current_user.class_id
            logger.info(f"非管理员用户，使用其班级ID: {class_id}")
        elif not class_id:
            class_id = None  # 设为None，后续查询不使用班级ID条件
            logger.warning("管理员未提供班级ID，将直接按学生ID查询")
        
        # 记录最终使用的参数
        logger.info(f"更新操作最终参数 - 学生ID: {student_id}, 班级ID: {class_id}, 学期: {semester}")
        
        # 提取德育维度数据
        deyu_dimensions = {
            'pinzhi': data.get('pinzhi', 0),    # 品质 30分
            'xuexi': data.get('xuexi', 0),      # 学习 20分
            'jiankang': data.get('jiankang', 0), # 健康 20分
            'shenmei': data.get('shenmei', 0),   # 审美 10分
            'shijian': data.get('shijian', 0),   # 实践 10分
            'shenghuo': data.get('shenghuo', 0)  # 生活 10分
        }
        
        # 确保所有值都是整数
        for key, value in deyu_dimensions.items():
            try:
                deyu_dimensions[key] = int(value)
            except (ValueError, TypeError):
                logger.warning(f"维度 {key} 的值 '{value}' 不是有效整数，设为0")
                deyu_dimensions[key] = 0
        
        logger.info(f"提取的德育维度数据: {deyu_dimensions}")
        
        # 数据校验
        total_score = sum(deyu_dimensions.values())
        if total_score > 100:
            logger.warning(f"德育维度总分超过100: {total_score}")
        
        # 更新学生记录
        conn = get_db_connection()
        if not conn:
            logger.error("无法获取数据库连接")
            return jsonify({'status': 'error', 'message': '数据库连接失败'}), 500
            
        cursor = conn.cursor()
        
        # 获取当前时间
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 查询条件：如果有班级ID，则同时匹配学生ID和班级ID；否则只匹配学生ID
        query = "SELECT id, class_id, name, class FROM students WHERE id = ?"
        params = [student_id]
        
        if class_id:
            try:
                # 尝试将class_id转为整数
                int_class_id = int(class_id)
                query += " AND class_id = ?"
                params.append(int_class_id)
                logger.info(f"使用学生ID和班级ID查询: {student_id}, {int_class_id}")
            except (ValueError, TypeError):
                # 如果不能转为整数，就当作班级名称查询
                query += " AND class = ?"
                params.append(class_id)
                logger.info(f"使用学生ID和班级名称查询: {student_id}, {class_id}")
        
        # 检查学生是否存在
        logger.info(f"执行查询: {query}, 参数: {params}")
        cursor.execute(query, params)
        student = cursor.fetchone()
        
        if not student:
            logger.error(f"学生不存在或不在指定班级，ID: {student_id}, 班级ID: {class_id}")
            # 尝试不带班级条件查询学生，仅用于调试
            cursor.execute("SELECT id, class_id, name, class FROM students WHERE id = ?", (student_id,))
            debug_student = cursor.fetchone()
            if debug_student:
                debug_info = dict(debug_student)
                logger.info(f"找到学生记录，但班级不匹配: {debug_info}")
            conn.close()
            return jsonify({'status': 'error', 'message': '学生不存在或不在指定班级'}), 404
        
        # 从查询结果中获取实际的class_id
        student_info = dict(student)
        actual_class_id = student_info['class_id']
        actual_name = student_info['name']
        actual_class = student_info['class']
        
        logger.info(f"找到学生: {actual_name} (ID: {student_id}), 班级: {actual_class} (ID: {actual_class_id})")
        
        # 构建SQL更新语句
        update_fields = []
        update_params = []
        
        for field, value in deyu_dimensions.items():
            update_fields.append(f"{field} = ?")
            update_params.append(value)
        
        # 添加更新时间
        update_fields.append("updated_at = ?")
        update_params.append(now)
        
        # 添加WHERE条件参数 - 使用id和class_id作为复合主键
        update_params.append(student_id)
        update_params.append(actual_class_id)
        
        # 执行更新 - 修改WHERE条件，同时使用id和class_id
        sql = f"UPDATE students SET {', '.join(update_fields)} WHERE id = ? AND class_id = ?"
        logger.info(f"执行更新SQL: {sql}")
        logger.info(f"更新参数: {update_params}")
        cursor.execute(sql, update_params)
        
        affected_rows = cursor.rowcount
        logger.info(f"影响的行数: {affected_rows}")
        
        if affected_rows > 0:
            logger.info(f"成功更新学生 {student_id} (班级ID: {actual_class_id}) 的德育维度数据")
        else:
            logger.warning(f"更新学生 {student_id} (班级ID: {actual_class_id}) 的德育维度数据时未影响任何行")
            # 进一步检查
            cursor.execute(f"SELECT COUNT(*) FROM students WHERE id = ? AND class_id = ?", (student_id, actual_class_id))
            count = cursor.fetchone()[0]
            logger.info(f"再次检查: 找到 {count} 条匹配记录")
            
            if count == 0:
                logger.error(f"无法找到要更新的学生记录，ID: {student_id}, 班级ID: {actual_class_id}")
                conn.close()
                return jsonify({'status': 'error', 'message': '找不到要更新的学生记录'}), 404
        
        conn.commit()
        conn.close()
        
        # 返回更新后的总分
        total_score = sum(deyu_dimensions.values())
        
        logger.info(f"成功保存学生 {student_id} 的德育维度，总分: {total_score}")
        return jsonify({
            'status': 'ok', 
            'message': '成功保存学生德育维度',
            'total_score': total_score,
            'student_info': {
                'id': student_id,
                'name': actual_name,
                'class': actual_class,
                'class_id': actual_class_id
            }
        })
    except sqlite3.Error as db_err:
        logger.error(f'数据库操作错误: {str(db_err)}')
        logger.error(traceback.format_exc())
        return jsonify({'status': 'error', 'message': f'数据库操作失败: {str(db_err)}'}), 500
    except Exception as e:
        logger.error(f'保存学生德育维度时出错: {str(e)}')
        logger.error(traceback.format_exc())
        return jsonify({'status': 'error', 'message': f'保存学生德育维度失败: {str(e)}'}), 500

# 删除学生成绩
@deyu_bp.route('/api/deyu/<student_id>', methods=['DELETE'])
@login_required
def delete_student_deyu(student_id):
    try:
        class_id = request.args.get('class_id')
        
        # 班主任自动使用自己的班级ID
        if not current_user.is_admin:
            if not current_user.class_id:
                return jsonify({'status': 'error', 'message': '您尚未被分配班级，无法删除学生德育维度'}), 403
            class_id = current_user.class_id
        elif not class_id:
            return jsonify({'status': 'error', 'message': '管理员需提供班级ID参数'}), 400
            
        logger.info(f"收到删除德育维度请求，学生ID: {student_id}, 班级ID: {class_id}")
        
        # 直接从数据库更新学生信息
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取当前时间
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 检查学生是否存在
        cursor.execute("SELECT id FROM students WHERE id = ? AND class_id = ?", (student_id, class_id))
        if not cursor.fetchone():
            conn.close()
            logger.error(f"学生不存在，ID: {student_id}, 班级ID: {class_id}")
            return jsonify({'status': 'error', 'message': '学生不存在'}), 404
        
        # 执行更新，将所有德育维度字段设置为NULL
        cursor.execute('''
            UPDATE students 
            SET pinzhi = NULL, xuexi = NULL, jiankang = NULL, shenmei = NULL, shijian = NULL, shenghuo = NULL, 
                updated_at = ? 
            WHERE id = ? AND class_id = ?
        ''', (now, student_id, class_id))
        
        conn.commit()
        conn.close()
        
        logger.info(f"成功删除学生 {student_id} 的德育维度")
        return jsonify({'status': 'ok', 'message': '成功删除学生德育维度'})
    except Exception as e:
        logger.error(f'删除学生德育维度时出错: {str(e)}')
        return jsonify({'status': 'error', 'message': f'删除学生德育维度失败: {str(e)}'})

# 预览成绩导入
@deyu_bp.route('/api/deyu/preview-import', methods=['POST'])
@login_required
def preview_deyu_import():
    try:
        semester = request.form.get('semester', '上学期')
        class_id = request.form.get('class_id')
        
        # 班主任自动使用自己的班级ID
        if not current_user.is_admin:
            if not current_user.class_id:
                return jsonify({'status': 'error', 'message': '您尚未被分配班级，无法导入德育维度'}), 403
            class_id = current_user.class_id
        elif not class_id:
            return jsonify({'status': 'error', 'message': '管理员需提供班级ID参数'}), 400
            
        logger.info(f"收到预览德育维度导入请求，学期: {semester}, 班级: {class_id}")
        
        if 'file' not in request.files:
            logger.error("未提供文件")
            return jsonify({'status': 'error', 'message': '没有上传文件'}), 400
        
        file = request.files['file']
        logger.info(f"上传的文件名: {file.filename}")
        
        if file.filename == '':
            logger.error("未选择文件")
            return jsonify({'status': 'error', 'message': '未选择文件'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            logger.error("文件格式不正确")
            return jsonify({'status': 'error', 'message': '只能上传Excel (.xlsx/.xls) 文件'}), 400
        
        # 创建上传目录
        if not os.path.exists(UPLOAD_FOLDER):
            logger.info(f"创建上传目录: {UPLOAD_FOLDER}")
            os.makedirs(UPLOAD_FOLDER)
        
        # 保存文件
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        saved_filename = f"{timestamp}_{secure_filename(file.filename)}"
        file_path = os.path.join(UPLOAD_FOLDER, saved_filename)
        file.save(file_path)
        logger.info(f"保存文件到: {file_path}")
        
        # 确认文件是否成功保存
        if not os.path.exists(file_path):
            logger.error(f"文件保存失败: {file_path}")
            return jsonify({'status': 'error', 'message': '文件保存失败'}), 500
        
        logger.info(f"文件大小: {os.path.getsize(file_path)} 字节")
        
        # 调用预览功能
        logger.info("开始预览德育维度导入")
        result = grades_manager.preview_grades_from_excel(file_path, semester, class_id)
        
        if result['status'] == 'ok':
            logger.info(f"成功预览德育维度: {result['message']}")
            return jsonify(result)
        else:
            logger.error(f"预览德育维度失败: {result['message']}")
            return jsonify(result), 400
            
    except Exception as e:
        logger.error(f'预览德育维度导入时出错: {str(e)}')
        logger.error(traceback.format_exc())
        return jsonify({
            'status': 'error', 
            'message': f'预览德育维度导入失败: {str(e)}'
        }), 500

# 确认导入成绩
@deyu_bp.route('/api/deyu/confirm-import', methods=['POST'])
@login_required
def confirm_deyu_import():
    try:
        data = request.json
        logger.info(f"收到确认导入德育维度请求")
        
        if not data or 'file_path' not in data:
            logger.error("请求缺少文件路径")
            return jsonify({'status': 'error', 'message': '缺少文件路径参数'}), 400
        
        file_path = data.get('file_path')
        semester = data.get('semester', '上学期')
        class_id = data.get('class_id')
        
        # 班主任自动使用自己的班级ID
        if not current_user.is_admin:
            if not current_user.class_id:
                return jsonify({'status': 'error', 'message': '您尚未被分配班级，无法导入德育维度'}), 403
            class_id = current_user.class_id
        elif not class_id:
            return jsonify({'status': 'error', 'message': '管理员需提供班级ID参数'}), 400
        
        if not file_path:
            return jsonify({'status': 'error', 'message': '文件路径不能为空'}), 400
            
        if not os.path.exists(file_path):
            return jsonify({'status': 'error', 'message': '文件不存在，可能已被删除'}), 400
        
        # 导入成绩
        logger.info(f"开始导入德育维度，文件路径: {file_path}, 学期: {semester}, 班级: {class_id}")
        success, message = grades_manager.import_grades_from_excel(file_path, semester, class_id)
        
        if success:
            logger.info(f"成功导入德育维度: {message}")
            return jsonify({'status': 'ok', 'message': message})
        else:
            logger.error(f"导入德育维度失败: {message}")
            return jsonify({'status': 'error', 'message': message}), 400
    
    except Exception as e:
        logger.error(f'确认导入德育维度时出错: {str(e)}')
        logger.error(traceback.format_exc())
        return jsonify({
            'status': 'error', 
            'message': f'确认导入德育维度失败: {str(e)}'
        }), 500

# 导入学生成绩
@deyu_bp.route('/api/deyu/import', methods=['POST'])
@login_required
def import_deyu():
    try:
        semester = request.form.get('semester', '上学期')
        class_id = request.form.get('class_id')
        
        # 班主任自动使用自己的班级ID
        if not current_user.is_admin:
            if not current_user.class_id:
                return jsonify({'status': 'error', 'message': '您尚未被分配班级，无法导入德育维度'}), 403
            class_id = current_user.class_id
        elif not class_id:
            return jsonify({'status': 'error', 'message': '管理员需提供班级ID参数'}), 400
            
        logger.info(f"收到导入德育维度请求，学期: {semester}, 班级: {class_id}")
        
        if 'file' not in request.files:
            logger.error("未提供文件")
            return jsonify({'status': 'error', 'message': '没有上传文件'})
        
        file = request.files['file']
        logger.info(f"上传的文件名: {file.filename}")
        
        if file.filename == '':
            logger.error("未选择文件")
            return jsonify({'status': 'error', 'message': '未选择文件'})
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            logger.error("文件格式不正确")
            return jsonify({'status': 'error', 'message': '只能上传Excel (.xlsx/.xls) 文件'})
        
        # 创建上传目录
        if not os.path.exists(UPLOAD_FOLDER):
            logger.info(f"创建上传目录: {UPLOAD_FOLDER}")
            os.makedirs(UPLOAD_FOLDER)
        
        # 保存文件
        saved_filename = secure_filename(file.filename)
        file_path = os.path.join(UPLOAD_FOLDER, saved_filename)
        file.save(file_path)
        logger.info(f"保存文件到: {file_path}")
        
        # 确认文件是否成功保存
        if not os.path.exists(file_path):
            logger.error(f"文件保存失败: {file_path}")
            return jsonify({'status': 'error', 'message': '文件保存失败'})
        
        logger.info(f"文件大小: {os.path.getsize(file_path)} 字节")
        
        # 调用预览功能而不是直接导入
        logger.info("改为使用预览功能")
        result = grades_manager.preview_grades_from_excel(file_path, semester, class_id)
        
        if result['status'] == 'ok':
            logger.info(f"成功预览德育维度: {result['message']}")
            return jsonify(result)
        else:
            logger.error(f"预览德育维度失败: {result['message']}")
            return jsonify(result), 400
            
    except Exception as e:
        logger.error(f'导入德育维度时出错: {str(e)}')
        logger.error(traceback.format_exc())
        
        # 提供更明确的错误信息
        error_message = str(e)
        if "No such file or directory" in error_message:
            error_message = f"找不到文件或目录: {error_message}"
        elif "Permission denied" in error_message:
            error_message = f"权限被拒绝: {error_message}"
        
        return jsonify({'status': 'error', 'message': f'导入德育维度失败: {error_message}'})

# 下载德育维度导入模板
@deyu_bp.route('/api/deyu/template', methods=['GET'])
@login_required
def download_deyu_template():
    try:
        logger.info("收到下载德育维度导入模板请求")
        class_id = request.args.get('class_id')
        
        # 班主任自动使用自己的班级ID
        if not current_user.is_admin:
            if not current_user.class_id:
                return jsonify({'status': 'error', 'message': '您尚未被分配班级，无法下载德育维度模板'}), 403
            class_id = current_user.class_id
        elif not class_id and current_user.is_admin:
            logger.warning("管理员未指定班级ID")
            
        logger.info(f"准备下载德育维度模板，班级ID: {class_id}")
        
        template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates', 'grades_import_template.xlsx')
        logger.info(f"查找模板文件: {template_path}")
        
        # 生成文件名 - 如果指定了班级ID则包含在文件名中
        filename = '德育维度导入模板'
        if class_id:
            filename += f'_{class_id}'
        filename += '.xlsx'
        
        # 根据班级生成对应的模板
        template_path = grades_manager.create_empty_template(output_path=None, class_id=class_id)
        logger.info(f"创建的模板路径: {template_path}")
        
        if not template_path or not os.path.exists(template_path):
            logger.error("创建模板失败或模板路径无效")
            return jsonify({'status': 'error', 'message': '创建德育维度导入模板失败'})
        
        logger.info(f"准备发送模板文件: {template_path}")
        return send_file(template_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f'下载德育维度导入模板时出错: {str(e)}')
        logger.error(traceback.format_exc())
        return jsonify({'status': 'error', 'message': f'下载德育维度导入模板失败: {str(e)}'})

# API路由：下载模板文件
@deyu_bp.route('/api/deyu/download/<filename>', methods=['GET'])
def download_file(filename):
    """下载指定文件"""
    try:
        # 确保文件名安全
        safe_filename = secure_filename(filename)
        
        # 需要检查文件是否存在于模板文件夹中
        if not os.path.exists(os.path.join(TEMPLATE_FOLDER, safe_filename)):
            return jsonify({"status": "error", "message": "文件不存在"}), 404
        
        # 发送文件
        return send_from_directory(TEMPLATE_FOLDER, safe_filename, as_attachment=True)
        
    except Exception as e:
        logger.error(f"下载文件时发生错误: {str(e)}")
        return jsonify({"status": "error", "message": f"下载文件时发生错误: {str(e)}"}), 500

# API路由：批量更新德育维度
@deyu_bp.route('/api/deyu/batch-update', methods=['POST'])
def batch_update_grades():
    """批量更新德育维度"""
    try:
        data = request.json
        
        # 验证数据
        if not data:
            return jsonify({"status": "error", "message": "没有提供数据"}), 400
        
        # 所需的参数
        semester = data.get('semester', '')
        student_ids = data.get('student_ids', [])
        subjects = data.get('subjects', [])
        grade = data.get('grade', '')
        
        if not semester or not student_ids or not subjects:
            return jsonify({"status": "error", "message": "缺少必要的参数"}), 400
        
        # 调用管理器方法批量更新成绩
        success, message = grades_manager.batch_update_grades(semester, student_ids, subjects, grade)
        
        if success:
            return jsonify({"status": "ok", "message": message})
        else:
            return jsonify({"status": "error", "message": message}), 400
        
    except Exception as e:
        logger.error(f"批量更新成绩时发生错误: {str(e)}")
        return jsonify({"status": "error", "message": f"更新过程中发生错误: {str(e)}"}), 500

# API路由：清空所有德育维度
@deyu_bp.route('/api/deyu/clear-all', methods=['POST'])
def clear_all_grades():
    """清空所有德育维度"""
    try:
        data = request.json
        
        # 验证数据
        if not data or 'semester' not in data:
            return jsonify({"status": "error", "message": "缺少学期参数"}), 400
        
        semester = data['semester']
        
        # 调用管理器方法清空成绩
        success, message = grades_manager.clear_all_grades(semester)
        
        if success:
            return jsonify({"status": "ok", "message": message})
        else:
            return jsonify({"status": "error", "message": message}), 400
        
    except Exception as e:
        logger.error(f"清空成绩时发生错误: {str(e)}")
        return jsonify({"status": "error", "message": f"清空过程中发生错误: {str(e)}"}), 500

# API路由：导出德育维度
@deyu_bp.route('/api/deyu/export', methods=['GET'])
def export_grades():
    """导出德育维度"""
    try:
        semester = request.args.get('semester', '')
        
        if not semester:
            return jsonify({"status": "error", "message": "学期参数不能为空"}), 400
        
        # 获取所有成绩数据
        grades = grades_manager.get_all_grades(semester)
        
        if not grades:
            return jsonify({"status": "error", "message": "没有找到可导出的数据"}), 404
        
        # 创建一个新的Excel文件
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # 添加表头
        headers = ["学号", "姓名", "班级", "科目", "成绩"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # 科目映射
        subjects = {
            'daof': '道法',
            'yuwen': '语文',
            'shuxue': '数学',
            'yingyu': '英语',
            'laodong': '劳动',
            'tiyu': '体育',
            'yinyue': '音乐',
            'meishu': '美术',
            'kexue': '科学',
            'zonghe': '综合',
            'xinxi': '信息',
            'shufa': '书法'
        }
        
        # 添加数据
        row_idx = 2
        for student in grades:
            student_id = student.get('student_id', '')
            student_name = student.get('student_name', '')
            class_name = student.get('class', '')
            
            for subject_code, subject_name in subjects.items():
                if subject_code in student and student[subject_code]:
                    sheet.cell(row=row_idx, column=1, value=student_id)
                    sheet.cell(row=row_idx, column=2, value=student_name)
                    sheet.cell(row=row_idx, column=3, value=class_name)
                    sheet.cell(row=row_idx, column=4, value=subject_name)
                    sheet.cell(row=row_idx, column=5, value=student[subject_code])
                    row_idx += 1
        
        # 保存文件
        filename = f"德育维度_{semester}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        workbook.save(file_path)
        
        # 发送文件
        return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)
        
    except Exception as e:
        logger.error(f"导出成绩时发生错误: {str(e)}")
        return jsonify({"status": "error", "message": f"导出过程中发生错误: {str(e)}"}), 500 