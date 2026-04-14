#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
学生管理模块，包含学生相关的API路由和功能
"""

import os
import datetime
import traceback
import sqlite3
from flask import Blueprint, request, jsonify, send_from_directory, url_for
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from flask_login import login_required, current_user
from utils.class_filter import class_filter, user_can_access
import logging
import openpyxl

# 配置日志
logger = logging.getLogger(__name__)

# 学生蓝图
students_bp = Blueprint('students', __name__)

# 配置
UPLOAD_FOLDER = 'uploads'
TEMPLATE_FOLDER = 'templates'
DATABASE = 'students.db'

# 创建数据库连接
def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')  # 启用外键约束
    return conn

# 创建学生导入模板
def create_student_template():
    template_dir = 'templates'
    if not os.path.exists(template_dir):
        os.makedirs(template_dir)
    
    template_path = os.path.join(template_dir, 'student_template.xlsx')
    
    # 如果文件存在且被占用，则跳过创建
    if os.path.exists(template_path):
        try:
            # 尝试打开文件，如果可以打开就先删除
            with open(template_path, 'a'):
                pass
            os.remove(template_path)
        except:
            print("学生模板文件被占用，跳过创建")
            return template_path
    
    wb = Workbook()
    ws = wb.active
    ws.title = "学生信息"
    
    # 设置标题行
    headers = ['学号', '姓名', '性别', '班级', '身高(cm)', '体重(kg)', '胸围(cm)', '肺活量(ml)', '龋齿(个)', '视力左', '视力右']
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(i)].width = 15
    
    # 添加示例数据
    example_data = ['1', '张三', '男', '三年级一班', '135', '32', '65', '1500', '0', '5.0', '5.0']
    for i, value in enumerate(example_data, 1):
        ws.cell(row=2, column=i, value=value)
    
    # 添加说明文字
    ws.cell(row=4, column=1, value="说明事项：")
    ws.cell(row=5, column=1, value="1. 请按照示例格式填写学生信息")
    ws.cell(row=6, column=1, value='2. 性别请填写"男"或"女"')
    ws.cell(row=7, column=1, value="3. 班级格式: 三年级一班")
    ws.cell(row=8, column=1, value="4. 视力格式: 5.0 或 4.8 等")
    
    # 合并说明文字的单元格
    for i in range(4, 9):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
    
    wb.save(template_path)
    print("学生Excel模板创建完成")
    return template_path

# 获取所有学生API
@students_bp.route('/api/students', methods=['GET'], strict_slashes=False)
@login_required
def get_all_students():
    class_id = None
    
    # 班主任只能查看自己班级的学生
    if not current_user.is_admin:
        # 如果班主任没有被分配班级，则返回空列表
        if not current_user.class_id:
            logger.warning(f"班主任 {current_user.username} 未分配班级，不能查看任何学生")
            return jsonify({
                'status': 'ok',
                'count': 0,
                'students': [],
                'message': '您尚未被分配班级，无法查看学生信息'
            })
        class_id = current_user.class_id
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        if class_id:
            logger.info(f"仅获取班级 {class_id} 的学生")
            cursor.execute('SELECT * FROM students WHERE class_id = ? ORDER BY CAST(id AS INTEGER)', (class_id,))
        else:
            logger.info(f"获取所有学生")
            cursor.execute('SELECT * FROM students ORDER BY CAST(id AS INTEGER)')
            
        students = [dict(row) for row in cursor.fetchall()]
        
        return jsonify({
            'status': 'ok',
            'count': len(students),
            'students': students
        })
    except Exception as e:
        logger.error(f"获取学生列表时出错: {str(e)}")
        return jsonify({'status': 'error', 'message': f'获取学生列表失败: {str(e)}'}), 500
    finally:
        conn.close()

# 获取单个学生API
@students_bp.route('/api/student/<student_id>', methods=['GET'])
@login_required
def get_student(student_id):
    """获取学生详情"""
    try:
        # 获取班级ID
        class_id = request.args.get('class_id')
        if not class_id:
            return jsonify({'error': '缺少班级ID'}), 400
            
        # 获取数据库连接
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 使用复合主键查询
        cursor.execute('''
            SELECT s.*, c.class_name 
            FROM students s
            LEFT JOIN classes c ON s.class_id = c.id
            WHERE s.id = ? AND s.class_id = ?
        ''', (student_id, class_id))
        
        student = cursor.fetchone()
        if not student:
            return jsonify({'error': '未找到该学生'}), 404
            
        # 转换为字典
        student_dict = dict(student)
        
        # 处理数值字段
        numeric_fields = ['height', 'weight', 'chest_circumference', 'vital_capacity', 
                         'vision_left', 'vision_right']
        for field in numeric_fields:
            if field in student_dict and student_dict[field] is not None:
                student_dict[field] = float(student_dict[field])
        
        return jsonify(student_dict)
        
    except Exception as e:
        logger.error(f"获取学生详情时出错: {str(e)}")
        return jsonify({'error': f'获取学生详情失败: {str(e)}'}), 500
    finally:
        conn.close()

# 添加新学生API
@students_bp.route('/api/students', methods=['POST'], strict_slashes=False)
@login_required
def add_student():
    data = request.json
    
    if not data:
        return jsonify({'error': '无效的请求数据'}), 400
    
    required_fields = ['id', 'name', 'gender']
    for field in required_fields:
        if field not in data:
            return jsonify({'error': f'缺少必要的字段: {field}'}), 400
    
    # 如果是班主任且没有分配班级，不允许添加学生
    if not current_user.is_admin and not current_user.class_id:
        return jsonify({'error': '您尚未被分配班级，无法添加学生'}), 403
    
    # 如果是班主任，自动设置班级ID
    if not current_user.is_admin and current_user.class_id:
        # 检查是否已提供了class_id，如果是，确保与班主任的class_id一致
        if 'class_id' in data and data['class_id'] != current_user.class_id:
            return jsonify({'error': '班主任只能添加本班学生'}), 403
        
        # 设置班级ID
        data['class_id'] = current_user.class_id
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 检查学生ID是否已存在
    cursor.execute('SELECT id FROM students WHERE id = ?', (data['id'],))
    if cursor.fetchone():
        conn.close()
        return jsonify({'error': f'学号 {data["id"]} 已存在'}), 400
    
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # 确保数值类型字段的正确处理
    height = data.get('height')
    weight = data.get('weight')
    chest_circumference = data.get('chest_circumference')
    vital_capacity = data.get('vital_capacity')
    vision_left = data.get('vision_left')
    vision_right = data.get('vision_right')
    
    try:
        cursor.execute('''
        INSERT INTO students (
            id, name, gender, class, class_id, height, weight,
            chest_circumference, vital_capacity, dental_caries,
            vision_left, vision_right, physical_test_status, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['id'], data['name'], data['gender'], data.get('class', ''),
            data.get('class_id'), height, weight, 
            chest_circumference, vital_capacity, data.get('dental_caries', ''),
            vision_left, vision_right, data.get('physical_test_status', ''),
            now, now
        ))
        
        conn.commit()
        
        return jsonify({
            'status': 'ok',
            'message': f'成功添加学生: {data["name"]}',
            'student_id': data['id']
        })
    except Exception as e:
        conn.rollback()
        logger.error(f"添加学生时出错: {str(e)}")
        return jsonify({'error': f'添加学生时出错: {str(e)}'}), 500
    finally:
        conn.close()

# 更新学生API
@students_bp.route('/api/student/<student_id>', methods=['PUT'])
@login_required
def update_student(student_id):
    """更新学生信息"""
    try:
        data = request.json
        if not data:
            return jsonify({'error': '没有提供更新数据'}), 400
        
        # 如果是班主任且没有分配班级，不允许更新学生
        if not current_user.is_admin and not current_user.class_id:
            return jsonify({'error': '您尚未被分配班级，无法更新学生信息'}), 403
            
        # 获取班级ID
        class_id = data.get('class_id')
        if not class_id:
            return jsonify({'error': '缺少班级ID'}), 400
            
        # 班主任只能更新自己班级的学生
        if not current_user.is_admin and str(current_user.class_id) != str(class_id):
            return jsonify({'error': '您只能更新自己班级的学生信息'}), 403
        
        # 获取数据库连接
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 检查学生是否存在
        cursor.execute('SELECT id FROM students WHERE id = ? AND class_id = ?', 
                      (student_id, class_id))
        if not cursor.fetchone():
            return jsonify({'error': '未找到该学生'}), 404
        
        # 准备更新数据
        update_fields = []
        update_values = []
        
        # 处理数值字段
        numeric_fields = ['height', 'weight', 'chest_circumference', 'vital_capacity', 
                         'vision_left', 'vision_right']
        for field in numeric_fields:
            if field in data:
                try:
                    value = float(data[field]) if data[field] is not None else None
                    update_fields.append(f"{field} = ?")
                    update_values.append(value)
                except (ValueError, TypeError):
                    return jsonify({'error': f'字段 {field} 的值无效'}), 400
        
        # 处理其他字段
        for field, value in data.items():
            if field not in numeric_fields and field != 'class_id':
                update_fields.append(f"{field} = ?")
                update_values.append(value)
        
        # 添加更新时间
        update_fields.append("updated_at = ?")
        update_values.append(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # 添加WHERE条件
        update_values.extend([student_id, class_id])
        
        # 构建并执行更新SQL
        update_sql = f"""
            UPDATE students 
            SET {', '.join(update_fields)}
            WHERE id = ? AND class_id = ?
        """
        
        cursor.execute(update_sql, update_values)
        conn.commit()
        
        return jsonify({'message': '学生信息更新成功'})
        
    except Exception as e:
        logger.error(f"更新学生信息时出错: {str(e)}")
        return jsonify({'error': f'更新学生信息失败: {str(e)}'}), 500
    finally:
        conn.close()

@students_bp.route('/api/student/<student_id>/comments', methods=['PUT'])
@login_required
def update_student_comments(student_id):
    """更新学生评语"""
    try:
        data = request.json
        if not data or 'comments' not in data:
            return jsonify({'error': '没有提供评语内容'}), 400
            
        # 获取班级ID
        class_id = data.get('class_id')
        if not class_id:
            return jsonify({'error': '缺少班级ID'}), 400
            
        # 获取数据库连接
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 检查学生是否存在
        cursor.execute('SELECT id FROM students WHERE id = ? AND class_id = ?', 
                      (student_id, class_id))
        if not cursor.fetchone():
            return jsonify({'error': '未找到该学生'}), 404
        
        # 更新评语
        cursor.execute('''
            UPDATE students 
            SET comments = ?, updated_at = ?
            WHERE id = ? AND class_id = ?
        ''', (data['comments'], 
              datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
              student_id, class_id))
        
        conn.commit()
        return jsonify({'message': '评语更新成功'})
        
    except Exception as e:
        logger.error(f"更新学生评语时出错: {str(e)}")
        return jsonify({'error': f'更新评语失败: {str(e)}'}), 500
    finally:
        conn.close()

@students_bp.route('/api/student/<student_id>', methods=['DELETE'])
@login_required
def delete_student(student_id):
    """删除学生"""
    try:
        # 获取班级ID
        class_id = request.args.get('class_id')
        if not class_id:
            return jsonify({'error': '缺少班级ID'}), 400
            
        # 如果是班主任且没有分配班级，不允许删除学生
        if not current_user.is_admin and not current_user.class_id:
            logger.warning(f"班主任 {current_user.username} 未分配班级，不能删除学生")
            return jsonify({'error': '您尚未被分配班级，无法删除学生'}), 403
            
        # 班主任只能删除自己班级的学生
        if not current_user.is_admin and str(current_user.class_id) != str(class_id):
            return jsonify({'error': '您只能删除自己班级的学生'}), 403
            
        # 获取数据库连接
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 检查学生是否存在
        cursor.execute('SELECT id FROM students WHERE id = ? AND class_id = ?', 
                      (student_id, class_id))
        if not cursor.fetchone():
            return jsonify({'error': '未找到该学生'}), 404
        
        # 删除学生
        cursor.execute('DELETE FROM students WHERE id = ? AND class_id = ?', 
                      (student_id, class_id))
        
        conn.commit()
        return jsonify({'message': '学生删除成功'})
        
    except Exception as e:
        logger.error(f"删除学生时出错: {str(e)}")
        return jsonify({'error': f'删除学生失败: {str(e)}'}), 500
    finally:
        conn.close()

# 导入学生预览API
@students_bp.route('/api/import-students', methods=['POST'])
@login_required
def import_students_preview():
    """预览导入学生数据"""
    try:
        # 如果是班主任且没有分配班级，不允许导入学生
        if not current_user.is_admin and not current_user.class_id:
            logger.warning(f"班主任 {current_user.username} 未分配班级，不能导入学生")
            return jsonify({'error': '您尚未被分配班级，无法导入学生'}), 403
            
        # 获取上传的文件
        if 'file' not in request.files:
            return jsonify({'error': '没有上传文件'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400
            
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': '请上传Excel文件'}), 400
            
        # 保存文件
        filename = secure_filename(file.filename)
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        
        # 确保上传目录存在
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        
        # 尝试保存文件
        try:
            file.save(file_path)
            logger.info(f"文件已保存到: {file_path}")
        except Exception as e:
            logger.error(f"保存文件时出错: {str(e)}")
            return jsonify({'error': f'保存文件失败: {str(e)}'}), 500
        
        # 读取Excel文件
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
        except Exception as e:
            logger.error(f"读取Excel文件时出错: {str(e)}")
            return jsonify({'error': f'读取Excel文件失败: {str(e)}'}), 500
        
        # 获取表头
        headers = [cell.value for cell in ws[1]]
        
        # 获取数据库连接
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 准备数据
        students_data = []
        error_records = []
        skipped_count = 0
        updated_count = 0
        added_count = 0
        
        # 字段映射表，Excel列名 -> 数据库字段名
        field_mapping = {
            '学号': 'id',
            '姓名': 'name',
            '性别': 'gender',
            '班级': 'class',
            '身高(cm)': 'height',
            '身高': 'height',
            '体重(kg)': 'weight',
            '体重': 'weight',
            '胸围(cm)': 'chest_circumference',
            '胸围': 'chest_circumference',
            '肺活量(ml)': 'vital_capacity',
            '肺活量': 'vital_capacity',
            '龋齿(个)': 'dental_caries',
            '龋齿': 'dental_caries',
            '视力左': 'vision_left',
            '视力右': 'vision_right',
            '体测情况': 'physical_test_status',
        }
        
        # 处理每一行数据
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            student = {}
            
            # 处理每个单元格
            for header, cell in zip(headers, row):
                cell_value = cell.value
                field_name = field_mapping.get(header)
                
                if field_name:
                    # 数值字段处理
                    if field_name in ['height', 'weight', 'chest_circumference', 'vital_capacity', 
                                    'vision_left', 'vision_right']:
                        try:
                            if cell_value is not None and cell_value != '' and cell_value != '-':
                                student[field_name] = float(cell_value)
                            else:
                                student[field_name] = None
                        except (ValueError, TypeError):
                            student[field_name] = None
                            logger.warning(f"行 {row_idx}: '{header}' 值 '{cell_value}' 无法转换为数值")
                    else:
                        # 文本字段处理
                        if cell_value is not None and cell_value != '' and cell_value != '-':
                            student[field_name] = str(cell_value)
                        else:
                            student[field_name] = None
            
            # 检查必填字段
            if not student.get('id') or not student.get('name'):
                error_records.append({
                    'row': row_idx,
                    'reason': '学号或姓名为空'
                })
                skipped_count += 1
                continue
            
            # 如果是班主任，设置班级ID
            if not current_user.is_admin and current_user.class_id:
                student['class_id'] = current_user.class_id
            else:
                # 管理员模式：根据班级名称查找班级ID
                if student.get('class'):
                    cursor.execute('SELECT id FROM classes WHERE class_name = ?', (student['class'],))
                    class_result = cursor.fetchone()
                    if class_result:
                        student['class_id'] = class_result['id']
                    else:
                        error_records.append({
                            'row': row_idx,
                            'reason': f'班级 "{student["class"]}" 不存在'
                        })
                        skipped_count += 1
                        continue
                else:
                    error_records.append({
                        'row': row_idx,
                        'reason': '班级信息缺失'
                    })
                    skipped_count += 1
                    continue
            
            # 检查学生是否已存在（使用复合主键检查）
            cursor.execute('SELECT id FROM students WHERE id = ? AND class_id = ?', 
                          (student['id'], student['class_id']))
            if cursor.fetchone():
                updated_count += 1
            else:
                added_count += 1
                
            students_data.append(student)
        
        # 成功读取数据后，不删除文件，确保导入过程中文件可用
        # 注意: 稍后需要在确认导入后或一段时间后清理这些文件
        
        logger.info(f"成功解析Excel文件，发现 {len(students_data)} 条有效学生记录")
        
        return jsonify({
            'status': 'ok',
            'message': f'成功读取 {len(students_data)} 条学生记录',
            'preview': {
                'total': len(students_data),
                'added': added_count,
                'updated': updated_count,
                'skipped': skipped_count,
                'errors': error_records
            },
            'students': students_data,
            'file_path': file_path  # 返回文件路径以供后续使用
        })
        
    except Exception as e:
        logger.error(f"导入学生预览时出错: {str(e)}")
        return jsonify({'error': f'导入预览失败: {str(e)}'}), 500

@students_bp.route('/api/confirm-import', methods=['POST'])
@login_required
def confirm_import():
    """确认导入学生数据"""
    try:
        # 如果是班主任且没有分配班级，不允许导入学生
        if not current_user.is_admin and not current_user.class_id:
            logger.warning(f"班主任 {current_user.username} 未分配班级，不能确认导入学生")
            return jsonify({'error': '您尚未被分配班级，无法导入学生'}), 403
            
        # 获取JSON数据
        data = request.get_json()
        if not data or 'students' not in data:
            return jsonify({'error': '无效的数据格式'}), 400
            
        students = data['students']
        if not students:
            return jsonify({'error': '没有要导入的学生数据'}), 400
            
        # 获取数据库连接
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取当前时间
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 获取数据库列名
        cursor.execute("PRAGMA table_info(students)")
        db_columns = [row[1] for row in cursor.fetchall()]
        
        # 准备统计
        inserted_count = 0
        updated_count = 0
        error_count = 0
        error_details = []
        
        # 遍历学生数据，执行插入或更新
        for student in students:
            try:
                # 提取班级ID和学生ID
                class_id = student.get('class_id')
                student_id = student.get('id')
                
                if not class_id or not student_id:
                    error_count += 1
                    error_details.append(f"学生 {student.get('name', '未知')} (ID: {student_id or '未知'}) 缺少班级ID或学生ID")
                    continue
                
                # 检查该学生是否已存在
                cursor.execute('SELECT id FROM students WHERE id = ? AND class_id = ?', 
                            (student_id, class_id))
                existing_student = cursor.fetchone()
                
                # 准备数据，只包含数据库存在的列
                db_fields = []
                db_values = []
                update_pairs = []
                
                for key, value in student.items():
                    if key in db_columns:
                        db_fields.append(key)
                        db_values.append(value)
                        update_pairs.append(f"{key} = ?")
                
                # 添加创建和更新时间
                if 'created_at' in db_columns and not existing_student:
                    db_fields.append('created_at')
                    db_values.append(now)
                
                if 'updated_at' in db_columns:
                    db_fields.append('updated_at')
                    db_values.append(now)
                
                # 执行SQL
                if existing_student:
                    # 更新现有学生
                    query = f"UPDATE students SET {', '.join(update_pairs)} WHERE id = ? AND class_id = ?"
                    cursor.execute(query, db_values + [student_id, class_id])
                    updated_count += 1
                else:
                    # 插入新学生
                    placeholders = ', '.join(['?'] * len(db_fields))
                    query = f"INSERT INTO students ({', '.join(db_fields)}) VALUES ({placeholders})"
                    cursor.execute(query, db_values)
                    inserted_count += 1
                
            except Exception as e:
                logger.error(f"导入学生 {student.get('id', '未知')} 时出错: {str(e)}")
                error_count += 1
                error_details.append(f"学生 {student.get('name', '未知')} (ID: {student.get('id', '未知')}) 导入失败: {str(e)}")
        
        # 提交事务
        conn.commit()
        conn.close()
        
        # 清理临时文件（可选）
        # 注意：这里不再尝试删除可能不存在的文件
        
        # 生成响应
        success_count = inserted_count + updated_count
        status = 'ok' if error_count == 0 else 'partial' if success_count > 0 else 'error'
        
        logger.info(f"导入完成: 成功 {success_count} 条 (新增 {inserted_count}, 更新 {updated_count}), 失败 {error_count} 条")
        
        return jsonify({
            'status': status,
            'message': f'完成导入 {success_count} 条记录',
            'success_count': success_count,
            'inserted_count': inserted_count,
            'updated_count': updated_count,
            'error_count': error_count,
            'error_details': error_details
        })
        
    except Exception as e:
        logger.error(f"确认导入学生时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': f'确认导入失败: {str(e)}'}), 500

# 下载模板API
@students_bp.route('/api/template', methods=['GET'])
def download_template():
    template_path = os.path.join(TEMPLATE_FOLDER, 'student_template.xlsx')
    if not os.path.exists(template_path):
        create_student_template()
    
    return jsonify({
        'status': 'ok',
        'template_url': f'/download/template/student_template.xlsx'
    })

# 提供模板下载
@students_bp.route('/download/template/<filename>', methods=['GET'])
def serve_template(filename):
    return send_from_directory(TEMPLATE_FOLDER, filename)

@students_bp.route('/api/students/<student_id>/update-subject', methods=['POST'])
@login_required
def update_student_subject(student_id):
    """更新学生的单个或多个学科成绩"""
    try:
        # 获取JSON数据
        data = request.get_json()
        if not data:
            return jsonify({'status': 'error', 'message': '无效的数据格式'})
        
        logger.info(f"正在更新学生ID={student_id}的学科成绩: {data}")
        
        # 获取当前用户班级ID
        user_class_id = current_user.class_id if not current_user.is_admin else None
        
        # 如果是班主任且没有分配班级，不允许更新学生成绩
        if not current_user.is_admin and not user_class_id:
            logger.warning(f"班主任 {current_user.username} 未分配班级，不能更新学生成绩")
            return jsonify({'status': 'error', 'message': '您尚未被分配班级，无法更新学生成绩'})
        
        # 检查学生是否存在，并验证班级ID
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 查询学生时需要同时检查学生ID和班级ID
        if user_class_id:
            # 班主任只能修改自己班级的学生
            cursor.execute('SELECT * FROM students WHERE id = ? AND class_id = ?', (student_id, user_class_id))
        else:
            # 管理员可以修改任意学生，但仍需要查询该学生
            cursor.execute('SELECT * FROM students WHERE id = ?', (student_id,))
            
        student = cursor.fetchone()
        
        if not student:
            conn.close()
            if user_class_id:
                # 如果是班主任，提示可能是权限问题
                return jsonify({
                    'status': 'error', 
                    'message': f'找不到ID为{student_id}的学生，或该学生不在您负责的班级中'
                })
            else:
                return jsonify({'status': 'error', 'message': f'找不到ID为{student_id}的学生'})
        
        # 支持的学科列表
        subjects = ['daof', 'yuwen', 'shuxue', 'yingyu', 'laodong', 'tiyu', 
                   'yinyue', 'meishu', 'kexue', 'zonghe', 'xinxi', 'shufa']
        
        # 学科到显示名称的映射，用于日志
        subject_display_names = {
            'daof': '道法', 'yuwen': '语文', 'shuxue': '数学', 'yingyu': '英语',
            'laodong': '劳动', 'tiyu': '体育', 'yinyue': '音乐', 'meishu': '美术',
            'kexue': '科学', 'zonghe': '综合', 'xinxi': '信息技术', 'shufa': '书法'
        }
        
        # 验证有效的成绩等级
        valid_grades = ['优', '良', '及格', '待及格', '']
        
        # 记录更新了多少个学科
        updated_count = 0
        
        # 遍历提交的数据，更新各个学科的成绩
        for subject, value in data.items():
            if subject in subjects:
                # 验证成绩值是否有效
                if value not in valid_grades:
                    logger.warning(f"尝试设置无效的成绩值: subject={subject}, value={value}")
                    continue
                
                # 更新数据库中的成绩，同时确保只更新特定班级的学生
                try:
                    if user_class_id:
                        # 班主任：使用ID和班级ID的组合查询条件
                        cursor.execute(f"UPDATE students SET {subject} = ? WHERE id = ? AND class_id = ?", 
                                      (value, student_id, user_class_id))
                    else:
                        # 管理员：仅使用ID查询条件
                        cursor.execute(f"UPDATE students SET {subject} = ? WHERE id = ?", 
                                      (value, student_id))
                                      
                    updated_count += 1
                    logger.info(f"已更新学生{student_id}的{subject_display_names.get(subject, subject)}成绩为: {value}")
                except Exception as e:
                    logger.error(f"更新科目成绩出错: {str(e)}")
        
        # 保存更改
        if updated_count > 0:
            # 更新修改时间，同样需要确保只更新特定班级的学生
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if user_class_id:
                cursor.execute('UPDATE students SET updated_at = ? WHERE id = ? AND class_id = ?', 
                              (now, student_id, user_class_id))
            else:
                cursor.execute('UPDATE students SET updated_at = ? WHERE id = ?', 
                              (now, student_id))
                              
            conn.commit()
            logger.info(f"成功保存学生{student_id}的{updated_count}个学科成绩")
            return jsonify({'status': 'ok', 'message': f'成功更新{updated_count}个学科成绩'})
        else:
            conn.close()
            return jsonify({'status': 'warning', 'message': '没有发现要更新的学科成绩'})
    
    except Exception as e:
        if 'conn' in locals() and conn:
            conn.rollback()
        error_msg = f"更新学生学科成绩时出错: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return jsonify({'status': 'error', 'message': error_msg})

@students_bp.route('/api/students', methods=['GET'])
@login_required
def get_students():
    """获取学生列表，支持按班级筛选"""
    try:
        # 获取请求参数
        class_id = request.args.get('class_id')
        with_grades = request.args.get('with_grades', 'false').lower() == 'true'
        
        # 获取当前用户信息
        current_user_id = current_user.id
        current_user_class_id = current_user.class_id
        is_admin = current_user.is_admin
        
        # 记录日志
        logger.info(f"获取学生列表请求: user={current_user.username}, class_id参数={class_id}, 用户班级={current_user_class_id}")
        
        # 非管理员且未分配班级的班主任不能看到任何学生
        if not is_admin and not current_user_class_id:
            logger.warning(f"班主任 {current_user.username} 未分配班级，不能查看任何学生")
            return jsonify({
                'status': 'ok',
                'students': [],
                'total': 0,
                'message': '您尚未被分配班级，无法查看学生信息'
            })
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 非管理员只能查看自己负责班级的学生
        if not is_admin and current_user_class_id:
            logger.info(f"班主任查询自己班级 {current_user_class_id} 的学生")
            cursor.execute('SELECT * FROM students WHERE class_id = ? ORDER BY CAST(id AS INTEGER)', (current_user_class_id,))
        # 如果指定了班级ID，则按班级筛选
        elif class_id:
            # 尝试处理整数或字符串的班级ID
            try:
                # 尝试转为整数
                int_class_id = int(class_id)
                logger.info(f"以数字形式查询班级ID: {int_class_id}")
                cursor.execute('SELECT * FROM students WHERE class_id = ? ORDER BY CAST(id AS INTEGER)', (int_class_id,))
            except (ValueError, TypeError):
                # 如果无法转为整数，尝试按班级名称查询
                logger.info(f"尝试以班级名称查询: {class_id}")
                cursor.execute('SELECT * FROM students WHERE class = ? ORDER BY CAST(id AS INTEGER)', (class_id,))
            
            # 检查结果数量
            students = cursor.fetchall()
            if not students:
                # 尝试模糊匹配
                logger.info(f"未找到精确匹配班级，尝试模糊匹配: {class_id}")
                cursor.execute('SELECT * FROM students WHERE class LIKE ? ORDER BY CAST(id AS INTEGER)', (f'%{class_id}%',))
                students = cursor.fetchall()
            else:
                # 重置游标位置
                cursor.execute('SELECT * FROM students WHERE class_id = ? ORDER BY CAST(id AS INTEGER)', (int_class_id,))
        else:
            logger.info("管理员查询所有学生")
            cursor.execute('SELECT * FROM students ORDER BY CAST(id AS INTEGER)')
        
        students = [dict(row) for row in cursor.fetchall()]
        logger.info(f"查询结果: 找到 {len(students)} 名学生")
        
        conn.close()
        
        return jsonify({
            'status': 'ok',
            'students': students,
            'total': len(students)
        })
        
    except Exception as e:
        logger.error(f"获取学生列表时出错: {str(e)}", exc_info=True)
        return jsonify({'status': 'error', 'message': f'获取学生列表失败: {str(e)}'}) 