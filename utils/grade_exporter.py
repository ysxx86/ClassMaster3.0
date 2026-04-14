# -*- coding: utf-8 -*-
"""
成绩导出工具
支持导出成绩为PDF，包含分数段统计和成绩分析
"""

import sqlite3
import os
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import logging

logger = logging.getLogger(__name__)

class GradeExporter:
    """成绩导出器"""
    
    def __init__(self, db_path="students.db"):
        self.db_path = db_path
        
        # 学科映射
        self.subject_mapping = {
            'daof': '道法',
            'yuwen': '语文',
            'shuxue': '数学',
            'yingyu': '英语',
            'laodong': '劳动',
            'tiyu': '体育',
            'yinyue': '音乐',
            'meishu': '美术',
            'kexue': '科学',
            'zonghe': '综合',
            'xinxi': '信息',
            'shufa': '书法',
            'xinli': '心理'
        }
        
        # 成绩等级
        self.grade_levels = ['优', '良', '及格', '待及格', '/']
        
        # 不同年级的达优标准（根据成绩分析页面的设置）
        self.excellence_standards = {
            '1-2年级': {'优': 90, '良': 80, '及格': 60},
            '3-4年级': {'优': 85, '良': 75, '及格': 60},
            '5-6年级': {'优': 85, '良': 75, '及格': 60}
        }
        
        # 注册中文字体
        self._register_fonts()
    
    def _register_fonts(self):
        """注册中文字体"""
        try:
            # 尝试使用系统字体
            font_paths = [
                'C:/Windows/Fonts/simhei.ttf',  # 黑体
                'C:/Windows/Fonts/simsun.ttc',  # 宋体
                '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',  # Linux
                '/System/Library/Fonts/PingFang.ttc'  # macOS
            ]
            
            for font_path in font_paths:
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('Chinese', font_path))
                    logger.info(f"成功注册字体: {font_path}")
                    return
            
            logger.warning("未找到中文字体，PDF可能无法正确显示中文")
        except Exception as e:
            logger.error(f"注册字体失败: {str(e)}")
    
    def _get_grade_level(self, class_name):
        """根据班级名称判断年级段"""
        if '一年级' in class_name or '二年级' in class_name:
            return '1-2年级'
        elif '三年级' in class_name or '四年级' in class_name:
            return '3-4年级'
        elif '五年级' in class_name or '六年级' in class_name:
            return '5-6年级'
        else:
            return '3-4年级'  # 默认
    
    def _calculate_statistics(self, grades, grade_level):
        """
        计算成绩统计
        
        Args:
            grades: 成绩列表
            grade_level: 年级段 ('1-2年级', '3-4年级', '5-6年级')
        
        Returns:
            dict: 统计结果
        """
        if not grades:
            return {
                'total': 0,
                'excellent': 0,
                'good': 0,
                'pass': 0,
                'fail': 0,
                'absent': 0,
                'excellent_rate': 0,
                'good_rate': 0,
                'pass_rate': 0
            }
        
        # 统计各等级人数
        stats = {
            '优': 0,
            '良': 0,
            '及格': 0,
            '待及格': 0,
            '/': 0
        }
        
        for grade in grades:
            if grade in stats:
                stats[grade] += 1
        
        total = len(grades)
        valid_total = total - stats['/']  # 有效人数（排除缺考）
        
        # 计算比率
        excellent_rate = (stats['优'] / valid_total * 100) if valid_total > 0 else 0
        good_rate = ((stats['优'] + stats['良']) / valid_total * 100) if valid_total > 0 else 0
        pass_rate = ((stats['优'] + stats['良'] + stats['及格']) / valid_total * 100) if valid_total > 0 else 0
        
        return {
            'total': total,
            'excellent': stats['优'],
            'good': stats['良'],
            'pass': stats['及格'],
            'fail': stats['待及格'],
            'absent': stats['/'],
            'excellent_rate': round(excellent_rate, 1),
            'good_rate': round(good_rate, 1),
            'pass_rate': round(pass_rate, 1)
        }
    
    def export_class_grades_excel(self, class_id, semester, output_path=None):
        """
        导出班级成绩为Excel（与导入模板格式一致）
        
        Args:
            class_id: 班级ID
            semester: 学期
            output_path: 输出路径（可选）
        
        Returns:
            str: Excel文件路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
            
            # 获取班级信息和学生成绩
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # 获取班级名称
            cursor.execute('SELECT class_name FROM classes WHERE id = ?', (class_id,))
            class_row = cursor.fetchone()
            if not class_row:
                raise ValueError(f"班级ID {class_id} 不存在")
            
            class_name = class_row[0]
            
            # 获取学生成绩
            cursor.execute('''
                SELECT id, name, daof, yuwen, shuxue, yingyu, laodong, tiyu, 
                       yinyue, meishu, kexue, zonghe, xinxi, shufa, xinli
                FROM students
                WHERE class_id = ? AND semester = ?
                ORDER BY CAST(id AS INTEGER)
            ''', (class_id, semester))
            
            students = cursor.fetchall()
            conn.close()
            
            if not students:
                raise ValueError(f"班级 {class_name} 在学期 {semester} 没有成绩数据")
            
            # 生成输出路径
            if not output_path:
                export_dir = 'exports'
                os.makedirs(export_dir, exist_ok=True)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = os.path.join(export_dir, f'成绩表_{class_name}_{timestamp}.xlsx')
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "成绩表"
            
            # 设置列标题（与导入模板一致）
            columns = [
                {'header': '学号', 'width': 10},
                {'header': '姓名', 'width': 12},
                {'header': '班级', 'width': 15},
                {'header': '道法', 'width': 8},
                {'header': '语文', 'width': 8},
                {'header': '数学', 'width': 8},
                {'header': '英语', 'width': 8},
                {'header': '劳动', 'width': 8},
                {'header': '体育', 'width': 8},
                {'header': '音乐', 'width': 8},
                {'header': '美术', 'width': 8},
                {'header': '科学', 'width': 8},
                {'header': '综合', 'width': 8},
                {'header': '信息', 'width': 8},
                {'header': '书法', 'width': 8},
                {'header': '心理', 'width': 8}
            ]
            
            # 设置标题行样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for col_idx, column in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=column['header'])
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.column_dimensions[get_column_letter(col_idx)].width = column['width']
            
            # 填充学生数据
            for row_idx, student in enumerate(students, start=2):
                ws.cell(row=row_idx, column=1, value=student[0])  # 学号
                ws.cell(row=row_idx, column=2, value=student[1])  # 姓名
                ws.cell(row=row_idx, column=3, value=class_name)  # 班级
                
                # 各科成绩
                for col_idx in range(4, 16):
                    grade = student[col_idx - 2] if student[col_idx - 2] else ''
                    cell = ws.cell(row=row_idx, column=col_idx, value=grade)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 保存文件
            wb.save(output_path)
            logger.info(f"成功导出成绩Excel: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"导出成绩Excel失败: {str(e)}")
            raise
    
    def export_class_grades(self, class_id, semester, output_path=None):
        """
        导出班级成绩为PDF
        
        Args:
            class_id: 班级ID
            semester: 学期
            output_path: 输出路径（可选）
        
        Returns:
            str: PDF文件路径
        """
        try:
            # 获取班级信息和学生成绩
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # 获取班级名称
            cursor.execute('SELECT class_name FROM classes WHERE id = ?', (class_id,))
            class_row = cursor.fetchone()
            if not class_row:
                raise ValueError(f"班级ID {class_id} 不存在")
            
            class_name = class_row[0]
            grade_level = self._get_grade_level(class_name)
            
            # 获取学生成绩
            cursor.execute('''
                SELECT id, name, daof, yuwen, shuxue, yingyu, laodong, tiyu, 
                       yinyue, meishu, kexue, zonghe, xinxi, shufa, xinli
                FROM students
                WHERE class_id = ? AND semester = ?
                ORDER BY CAST(id AS INTEGER)
            ''', (class_id, semester))
            
            students = cursor.fetchall()
            conn.close()
            
            if not students:
                raise ValueError(f"班级 {class_name} 在学期 {semester} 没有成绩数据")
            
            # 生成输出路径
            if not output_path:
                export_dir = 'exports'
                os.makedirs(export_dir, exist_ok=True)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = os.path.join(export_dir, f'成绩报表_{class_name}_{timestamp}.pdf')
            
            # 创建PDF
            doc = SimpleDocTemplate(
                output_path,
                pagesize=landscape(A4),
                rightMargin=1*cm,
                leftMargin=1*cm,
                topMargin=1.5*cm,
                bottomMargin=1*cm
            )
            
            # 构建内容
            story = []
            styles = getSampleStyleSheet()
            
            # 标题样式
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName='Chinese',
                fontSize=18,
                alignment=TA_CENTER,
                spaceAfter=12
            )
            
            # 正文样式
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontName='Chinese',
                fontSize=10,
                alignment=TA_LEFT
            )
            
            # 添加标题
            title = Paragraph(f'{class_name} 成绩报表', title_style)
            story.append(title)
            
            subtitle = Paragraph(f'学期：{semester}　　　导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}', normal_style)
            story.append(subtitle)
            story.append(Spacer(1, 0.5*cm))
            
            # 准备成绩表格数据
            table_data = [['学号', '姓名', '道法', '语文', '数学', '英语', '劳动', '体育', 
                          '音乐', '美术', '科学', '综合', '信息', '书法', '心理']]
            
            # 收集各科成绩用于统计
            subject_grades = {field: [] for field in self.subject_mapping.keys()}
            
            for student in students:
                row = [student[0], student[1]]  # 学号、姓名
                for i, field in enumerate(self.subject_mapping.keys(), start=2):
                    grade = student[i] if student[i] else '/'
                    row.append(grade)
                    if grade and grade != '/':
                        subject_grades[field].append(grade)
                table_data.append(row)
            
            # 创建成绩表格
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Chinese'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            story.append(table)
            story.append(PageBreak())
            
            # 添加统计分析页
            story.append(Paragraph(f'{class_name} 成绩统计分析', title_style))
            story.append(Paragraph(f'年级段：{grade_level}', normal_style))
            story.append(Spacer(1, 0.5*cm))
            
            # 统计表格
            stats_data = [['学科', '总人数', '优秀', '良好', '及格', '待及格', '缺考', '优秀率', '优良率', '及格率']]
            
            for field, subject_name in self.subject_mapping.items():
                grades = subject_grades[field]
                stats = self._calculate_statistics(grades, grade_level)
                
                stats_data.append([
                    subject_name,
                    stats['total'],
                    stats['excellent'],
                    stats['good'],
                    stats['pass'],
                    stats['fail'],
                    stats['absent'],
                    f"{stats['excellent_rate']}%",
                    f"{stats['good_rate']}%",
                    f"{stats['pass_rate']}%"
                ])
            
            stats_table = Table(stats_data, repeatRows=1)
            stats_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Chinese'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            story.append(stats_table)
            story.append(Spacer(1, 1*cm))
            
            # 添加说明
            note_text = f'''
            <b>说明：</b><br/>
            1. 本报表基于 {grade_level} 的评价标准<br/>
            2. 优秀率 = 优秀人数 / 有效人数 × 100%<br/>
            3. 优良率 = (优秀人数 + 良好人数) / 有效人数 × 100%<br/>
            4. 及格率 = (优秀人数 + 良好人数 + 及格人数) / 有效人数 × 100%<br/>
            5. 有效人数 = 总人数 - 缺考人数
            '''
            story.append(Paragraph(note_text, normal_style))
            
            # 生成PDF
            doc.build(story)
            logger.info(f"成功导出成绩PDF: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"导出成绩PDF失败: {str(e)}")
            raise
