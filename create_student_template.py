#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
学生数据导入模板生成工具
根据班级信息生成定制化的Excel模板
"""

import os
import sqlite3
import argparse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 配置
TEMPLATE_FOLDER = 'templates'
DATABASE = 'students.db'

def get_db_connection():
    """创建数据库连接"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def get_class_info(class_id):
    """根据班级ID获取班级信息"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('SELECT id, class_name FROM classes WHERE id = ?', (class_id,))
    class_info = cursor.fetchone()
    conn.close()
    
    if not class_info:
        raise ValueError(f"未找到ID为 {class_id} 的班级")
    
    return class_info

def create_custom_template(class_id):
    """为特定班级创建学生导入模板"""
    # 获取班级信息
    try:
        class_info = get_class_info(class_id)
        class_name = class_info['class_name']
    except Exception as e:
        print(f"获取班级信息失败: {str(e)}")
        return None
    
    # 创建模板目录
    template_dir = TEMPLATE_FOLDER
    if not os.path.exists(template_dir):
        os.makedirs(template_dir)
    
    # 定义文件名（使用班级名称）
    template_name = f"student_template_{class_name}.xlsx"
    template_path = os.path.join(template_dir, template_name)
    
    # 如果文件存在且被占用，则跳过创建
    if os.path.exists(template_path):
        try:
            # 尝试打开文件，如果可以打开就先删除
            with open(template_path, 'a'):
                pass
            os.remove(template_path)
        except:
            print(f"模板文件 {template_path} 被占用，跳过创建")
            return template_path
    
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "学生信息"
    
    # 设置样式
    # 标题样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # 班级信息样式
    class_font = Font(bold=True, size=14)
    class_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    class_alignment = Alignment(horizontal='left', vertical='center')
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # 设置标题行（与通用模板保持一致，从第1行开始）
    headers = ['学号', '姓名', '性别', '班级', '身高', '体重', 
              '胸围', '肺活量', '龋齿', '视力左', '视力右', '体测情况',
              '语文', '数学', '英语', '劳动', '体育', '音乐', '美术', 
              '科学', '综合', '信息', '书法', '心理',
              '品质', '学习', '健康', '审美', '实践', '生活',
              '评语']
    
    # 创建红色字体样式（用于必填字段）
    red_font = Font(bold=True, color="FF0000")
    
    # 写入表头（第1行）
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        # 姓名和班级标红
        if header in ['姓名', '班级']:
            cell.font = red_font
        else:
            cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        # 为评语列设置更宽的列宽
        if header == '评语':
            ws.column_dimensions[get_column_letter(i)].width = 40
        else:
            ws.column_dimensions[get_column_letter(i)].width = 15
    
    # 添加示例数据行（从第2行开始）
    example_rows = [
        ['1', '张三', '男', class_name, '135', '32', '65', '1500', '0', '5.0', '5.0', '健康',
         '优', '良', '优', '良', '优', '良', '优', '良', '优', '良', '优', '良',
         '25', '18', '18', '8', '8', '8', '该学生表现优秀'],
        ['2', '李四', '女', class_name, '130', '28', '62', '1400', '0', '5.0', '4.8', '健康',
         '良', '优', '良', '优', '良', '优', '良', '优', '良', '优', '良', '优',
         '23', '17', '17', '7', '7', '7', '该学生学习认真'],
    ]
    
    for row_idx, row_data in enumerate(example_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # 班级列设为浅灰色背景
            if col_idx == 4:
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # 添加说明信息（从第5行开始）
    ws.cell(row=5, column=1, value="说明事项：")
    ws.cell(row=6, column=1, value="1. 带红色标注的字段为必填项（姓名、班级）")
    ws.cell(row=7, column=1, value='2. 学号从1开始递增，如：1, 2, 3...（不是01, 001）')
    ws.cell(row=8, column=1, value='3. 性别填写"男"或"女"')
    ws.cell(row=9, column=1, value=f'4. 班级格式：二年级4班、204、2024级4班等')
    ws.cell(row=10, column=1, value='5. 身高、体重等数值字段只填数字，不要带单位')
    ws.cell(row=11, column=1, value='6. 学科成绩填写：优、良、及格、待及格')
    ws.cell(row=12, column=1, value='7. 德育分数：品质(0~30)、学习(0~20)、健康(0~20)、审美(0~10)、实践(0~10)、生活(0~10)')
    
    # 班级预填提示（红色字体）
    class_note = ws.cell(row=13, column=1, value=f'★ 班级字段已预设为 "{class_name}"，导入时班级必须与当前所管理班级一致')
    class_note.font = Font(color="FF0000", bold=True)
    
    # 保存工作簿
    wb.save(template_path)
    print(f"已为班级 '{class_name}' 创建自定义模板: {template_path}")
    
    return template_path

def main():
    parser = argparse.ArgumentParser(description='为特定班级创建学生导入模板')
    parser.add_argument('--class_id', type=int, required=True, help='班级ID')
    
    args = parser.parse_args()
    template_path = create_custom_template(args.class_id)
    
    if template_path:
        print(f"模板已创建: {template_path}")
    else:
        print("模板创建失败")

if __name__ == "__main__":
    main() 