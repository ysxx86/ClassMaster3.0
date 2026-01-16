#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
成绩分析模块，包含成绩分析相关的API路由和功能
"""

import os
import json
import sqlite3
import traceback
import datetime
from flask import Blueprint, request, jsonify, render_template, redirect, url_for, current_app as app, session, send_file, send_from_directory
import openpyxl
from werkzeug.utils import secure_filename
from flask_login import login_required, current_user
from utils.class_filter import class_filter, user_can_access
import logging
import pandas as pd
import numpy as np
import io
import time
from markupsafe import escape

# 配置日志
logger = logging.getLogger(__name__)

# 成绩分析蓝图
grade_analysis_bp = Blueprint('grade_analysis', __name__)

# 配置
UPLOAD_FOLDER = 'uploads/exams'
DATABASE = 'students.db'

# 初始化成绩分析模块
def init_grade_analysis(app):
    """初始化成绩分析模块"""
    logger.info("初始化成绩分析模块")
    
    # 创建上传目录
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    
    # 连接数据库
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    # 创建考试信息表
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS exams (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        exam_name TEXT NOT NULL,
        exam_date TEXT NOT NULL,
        class_id INTEGER NOT NULL,
        subjects TEXT NOT NULL,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        has_paper INTEGER DEFAULT 0,
        paper_path TEXT
    )
    ''')
    
    # 创建考试成绩表
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS exam_scores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        exam_id INTEGER NOT NULL,
        student_id TEXT NOT NULL,
        student_name TEXT NOT NULL,
        class_id INTEGER NOT NULL,
        subject TEXT NOT NULL,
        score REAL,
        leave_status INTEGER DEFAULT 0,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        FOREIGN KEY (exam_id) REFERENCES exams(id) ON DELETE CASCADE,
        UNIQUE (exam_id, student_id, subject)
    )
    ''')
    
    conn.commit()
    conn.close()
    logger.info("成绩分析模块初始化完成")

# 获取数据库连接
def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')  # 启用外键约束
    return conn

# 获取所有考试列表
@grade_analysis_bp.route('/api/exams', methods=['GET'])
@login_required
def get_exams():
    try:
        class_id = request.args.get('class_id')
        
        # 班主任自动使用自己的班级ID
        if not current_user.is_admin:
            if not current_user.class_id:
                return jsonify({'status': 'error', 'message': '您尚未被分配班级，无法查看考试数据'}), 403
            class_id = current_user.class_id
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 查询条件
        params = []
        query = "SELECT * FROM exams"
        
        if class_id:
            query += " WHERE class_id = ?"
            params.append(class_id)
        
        query += " ORDER BY exam_date DESC"
        
        cursor.execute(query, params)
        exams = [dict(row) for row in cursor.fetchall()]
        
        # 解析subjects字段为列表
        for exam in exams:
            exam['subjects'] = json.loads(exam['subjects'])
        
        return jsonify({
            'status': 'ok',
            'exams': exams
        })
    except Exception as e:
        logger.error(f"获取考试列表时出错: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'获取考试列表失败: {str(e)}'
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()

# 创建考试
@grade_analysis_bp.route('/api/exams', methods=['POST'])
@login_required
def create_exam():
    try:
        data = request.json
        
        # 验证必填字段
        required_fields = ['exam_name', 'exam_date', 'subjects']
        for field in required_fields:
            if field not in data:
                return jsonify({'status': 'error', 'message': f'缺少字段: {field}'}), 400
        
        # 班主任使用自己的班级ID
        if not current_user.is_admin:
            if not current_user.class_id:
                return jsonify({'status': 'error', 'message': '您尚未被分配班级，无法创建考试'}), 403
            class_id = current_user.class_id
        else:
            class_id = data.get('class_id')
            if not class_id:
                return jsonify({'status': 'error', 'message': '管理员需提供班级ID参数'}), 400
        
        # 准备保存数据
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 序列化subjects字段
        subjects_json = json.dumps(data['subjects'], ensure_ascii=False)
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 插入考试数据
        cursor.execute('''
        INSERT INTO exams (exam_name, exam_date, class_id, subjects, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?)
        ''', (data['exam_name'], data['exam_date'], class_id, subjects_json, now, now))
        
        # 获取新创建的考试ID
        exam_id = cursor.lastrowid
        
        conn.commit()
        
        return jsonify({
            'status': 'ok',
            'message': '考试创建成功',
            'exam_id': exam_id
        })
    except Exception as e:
        logger.error(f"创建考试时出错: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'创建考试失败: {str(e)}'
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()

# 获取考试详情
@grade_analysis_bp.route('/api/exams/<int:exam_id>', methods=['GET'])
@login_required
def get_exam_detail(exam_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取考试信息
        cursor.execute("SELECT * FROM exams WHERE id = ?", (exam_id,))
        exam = cursor.fetchone()
        
        if not exam:
            return jsonify({'status': 'error', 'message': '考试不存在'}), 404
        
        exam_dict = dict(exam)
        exam_dict['subjects'] = json.loads(exam_dict['subjects'])
        
        # 检查权限
        if not current_user.is_admin and current_user.class_id != exam_dict['class_id']:
            return jsonify({'status': 'error', 'message': '您无权查看此考试'}), 403
            
        # 获取班级信息
        class_id = exam_dict['class_id']
        cursor.execute("SELECT * FROM classes WHERE id = ?", (class_id,))
        class_info = cursor.fetchone()
        class_name = class_info['class_name'] if class_info else f"班级ID: {class_id}"
        
        # 获取班级所有学生数量
        cursor.execute("SELECT COUNT(*) as count FROM students WHERE class_id = ?", (class_id,))
        class_student_count = cursor.fetchone()['count']
        logger.info(f"班级 {class_name}(ID:{class_id}) 的实际学生总数: {class_student_count}")
        
        # 获取参加此次考试的所有记录
        cursor.execute('''
        SELECT s.*, students.name as student_name, students.class as class_name
        FROM exam_scores s
        LEFT JOIN students ON s.student_id = students.id
        WHERE s.exam_id = ? AND s.class_id = ?
        ORDER BY s.student_id, s.subject
        ''', (exam_id, class_id))
        
        all_scores = [dict(row) for row in cursor.fetchall()]
        total_score_records = len(all_scores)
        logger.info(f"考试ID: {exam_id}, 班级ID: {class_id} 的总成绩记录数: {total_score_records}")
        
        # 获取参加考试的学生ID集合
        student_ids = set(score['student_id'] for score in all_scores)
        actual_student_count = len(student_ids)
        logger.info(f"实际参加考试的学生数: {actual_student_count}, 学生ID列表: {student_ids}")
        
        # 按学科组织成绩数据
        scores_by_subject = {}
        for score in all_scores:
            subject = score['subject']
            if subject not in scores_by_subject:
                scores_by_subject[subject] = []
            scores_by_subject[subject].append(score)
        
        # 记录每个学科的记录数
        logger.info(f"共有 {len(scores_by_subject)} 个学科的成绩")
        for subject, subject_scores in scores_by_subject.items():
            subject_student_ids = [s['student_id'] for s in subject_scores]
            unique_student_ids = set(subject_student_ids)
            logger.info(f"学科: {subject}, 记录数: {len(subject_scores)}, 去重后学生数: {len(unique_student_ids)}")
            
            if len(subject_student_ids) != len(unique_student_ids):
                # 发现重复记录，需要处理
                logger.warning(f"⚠️ {subject}科目中存在重复学生记录！总记录数: {len(subject_scores)}, 去重后: {len(unique_student_ids)}")
                # 按学生ID分组，保留最新的成绩记录
                unique_scores = {}
                for score in subject_scores:
                    student_id = score['student_id']
                    if student_id not in unique_scores or score['updated_at'] > unique_scores[student_id]['updated_at']:
                        unique_scores[student_id] = score
                
                # 更新为去重后的成绩列表
                scores_by_subject[subject] = list(unique_scores.values())
                logger.info(f"去重后 {subject} 科目的记录数: {len(scores_by_subject[subject])}")
        
        # 计算每个学科的统计信息
        stats = {}
        for subject, subject_scores in scores_by_subject.items():
            # 提取分数列表 - 确保每个学生只计算一次
            subject_student_ids = set()
            valid_scores = []
            leave_students = 0  # 请假学生数
            
            for score in subject_scores:
                student_id = score['student_id']
                if student_id in subject_student_ids:
                    continue  # 跳过已经计算过的学生
                
                subject_student_ids.add(student_id)
                score_value = score['score']
                
                # 检查是否为请假学生（成绩为0）
                if score_value == 0:
                    leave_students += 1
                # 检查其他有效分数
                elif 0 < score_value <= 100:
                    valid_scores.append(score_value)
                else:
                    logger.warning(f"⚠️ 发现无效分数: 科目 {subject}, 学生 {score['student_name']}(ID:{student_id}), 分数: {score_value}")
            
            # 计算各类人数
            on_exam_students = len(subject_student_ids)  # 应考人数
            actual_exam_students = on_exam_students - leave_students  # 实考人数
            
            logger.info(f"科目 {subject} 的应考人数: {on_exam_students}, 实考人数: {actual_exam_students}, 请假人数: {leave_students}")
            
            # 检查有效学生人数
            if actual_exam_students == 0:
                logger.warning(f"⚠️ 科目 {subject} 没有有效成绩！")
                # 设置默认统计数据
                stats[subject] = {
                    'average': 0,
                    'max': 0,
                    'min': 0,
                    'excellent_rate': 0,
                    'pass_rate': 0,
                    'excellent_threshold': 90,
                    'score_distribution': {'0-59': 0, '60-69': 0, '70-79': 0, '80-89': 0, '90-100': 0},
                    'total_students': 0,
                    'on_exam_students': on_exam_students,
                    'actual_exam_students': actual_exam_students,
                    'leave_students': leave_students,
                    'class_student_count': class_student_count
                }
                continue
            
            # 判断年级
            class_name = class_name if class_name else ""
            
            # 根据年级确定优秀标准
            excellent_threshold = 90  # 默认优秀标准
            if "一年级" in class_name or "二年级" in class_name:
                excellent_threshold = 90
            elif "三年级" in class_name or "四年级" in class_name:
                excellent_threshold = 85
            elif "五年级" in class_name or "六年级" in class_name:
                excellent_threshold = 80
            
            # 计算各分数段人数 - 基于实考人数(不包括请假学生)
            below_60 = sum(1 for s in valid_scores if s < 60)
            from_60_to_69 = sum(1 for s in valid_scores if 60 <= s < 70)
            from_70_to_79 = sum(1 for s in valid_scores if 70 <= s < 80)
            from_80_to_89 = sum(1 for s in valid_scores if 80 <= s < 90)
            from_90_to_100 = sum(1 for s in valid_scores if s >= 90)
            
            # 记录详细的分段信息用于调试
            logger.info(f"科目: {subject}, 实考人数: {actual_exam_students}, 请假人数: {leave_students}")
            logger.info(f"  0-59分: {below_60}人")
            logger.info(f"  60-69分: {from_60_to_69}人")
            logger.info(f"  70-79分: {from_70_to_79}人")
            logger.info(f"  80-89分: {from_80_to_89}人")
            logger.info(f"  90-100分: {from_90_to_100}人")
            
            # 验证总和是否等于实考学生总数
            total_count = below_60 + from_60_to_69 + from_70_to_79 + from_80_to_89 + from_90_to_100
            if total_count != actual_exam_students:
                logger.error(f"❌ 分数段统计总人数({total_count})与实考学生总数({actual_exam_students})不匹配！科目: {subject}")
            
            # 基于实考人数计算及格率和优秀率
            pass_count = from_60_to_69 + from_70_to_79 + from_80_to_89 + from_90_to_100
            pass_rate = round(pass_count / actual_exam_students * 100, 2) if actual_exam_students > 0 else 0
            
            # 根据不同年级的优秀标准计算优秀人数
            excellent_count = 0
            if excellent_threshold == 90:
                excellent_count = from_90_to_100  # 90分及以上
            elif excellent_threshold == 85:
                # 85分及以上，即部分80-89分段和全部90-100分段
                excellent_count = from_90_to_100 + sum(1 for s in valid_scores if 85 <= s < 90)
            elif excellent_threshold == 80:
                # 80分及以上，即全部80-89分段和全部90-100分段
                excellent_count = from_80_to_89 + from_90_to_100
            
            excellent_rate = round(excellent_count / actual_exam_students * 100, 2) if actual_exam_students > 0 else 0
            
            # 计算统计数据
            stats[subject] = {
                'average': round(sum(valid_scores) / actual_exam_students, 2) if actual_exam_students > 0 else 0,
                'max': max(valid_scores) if valid_scores else 0,
                'min': min(valid_scores) if valid_scores else 0,
                'excellent_rate': excellent_rate,
                'pass_rate': pass_rate,
                'excellent_threshold': excellent_threshold,
                'score_distribution': {
                    '0-59': below_60,
                    '60-69': from_60_to_69,
                    '70-79': from_70_to_79,
                    '80-89': from_80_to_89,
                    '90-100': from_90_to_100
                },
                'total_students': actual_exam_students,  # 为了兼容性，保留此字段但实际是实考人数
                'on_exam_students': on_exam_students,  # 应考人数
                'actual_exam_students': actual_exam_students,  # 实考人数
                'leave_students': leave_students,  # 请假人数
                'class_student_count': class_student_count  # 班级总人数
            }
        
        return jsonify({
            'status': 'ok',
            'exam': exam_dict,
            'scores': all_scores,
            'stats': stats
        })
    except Exception as e:
        logger.error(f"获取考试详情时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'status': 'error',
            'message': f'获取考试详情失败: {str(e)}'
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()

# 上传考试成绩
@grade_analysis_bp.route('/api/exams/<int:exam_id>/scores', methods=['POST'])
@login_required
def upload_scores(exam_id):
    try:
        # 检查文件是否存在
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': '没有上传文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': '未选择文件'}), 400
        
        # 检查文件格式
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'status': 'error', 'message': '只支持Excel文件(.xlsx, .xls)'}), 400
        
        # 获取考试信息
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM exams WHERE id = ?", (exam_id,))
        exam = cursor.fetchone()
        
        if not exam:
            return jsonify({'status': 'error', 'message': '考试不存在'}), 404
        
        exam_dict = dict(exam)
        
        # 检查权限
        if not current_user.is_admin and current_user.class_id != exam_dict['class_id']:
            return jsonify({'status': 'error', 'message': '您无权上传此考试成绩'}), 403
        
        # 保存文件
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"{timestamp}_{secure_filename(file.filename)}"
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        
        # 确保上传目录存在
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        file.save(file_path)
        
        # 解析Excel文件
        df = pd.read_excel(file_path)
        
        # 检查必要的列是否存在
        required_columns = ['学号', '姓名']
        subjects_list = json.loads(exam_dict['subjects'])
        
        for col in required_columns:
            if col not in df.columns:
                return jsonify({'status': 'error', 'message': f'缺少必要的列: {col}'}), 400
        
        # 检查学科列是否存在
        missing_subjects = [subject for subject in subjects_list if subject not in df.columns]
        if missing_subjects:
            return jsonify({'status': 'error', 'message': f'缺少学科列: {", ".join(missing_subjects)}'}), 400
        
        # 准备批量插入数据
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rows_to_insert = []
        
        for _, row in df.iterrows():
            student_id = str(row['学号'])
            student_name = row['姓名']
            
            # 检查学生是否属于该班级
            cursor.execute("SELECT id FROM students WHERE id = ? AND class_id = ?", 
                          (student_id, exam_dict['class_id']))
            student = cursor.fetchone()
            
            if not student:
                continue  # 跳过不属于该班级的学生
            
            # 为每个学科添加一条记录
            for subject in subjects_list:
                if pd.notna(row[subject]):  # 检查成绩是否有效
                    leave_status = 0  # 默认正常状态
                    score_value = row[subject]
                    
                    # 处理成绩值
                    try:
                        # 尝试转换为浮点数
                        score = float(score_value)
                        # 如果成绩为0，视为请假
                        if score == 0:
                            leave_status = 1
                    except (ValueError, TypeError):
                        # 如果无法转换为浮点数，检查是否为"请假"文本
                        if isinstance(score_value, str) and "请假" in score_value:
                            score = 0  # 请假学生分数记为0
                            leave_status = 1
                        else:
                            # 跳过无法识别的成绩
                            logger.warning(f"跳过无法识别的成绩: {score_value}, 学生: {student_name}({student_id}), 学科: {subject}")
                            continue
                    
                    rows_to_insert.append((
                        exam_id, student_id, student_name, exam_dict['class_id'],
                        subject, score, leave_status, now, now
                    ))
        
        # 批量插入成绩
        cursor.executemany('''
        INSERT OR REPLACE INTO exam_scores 
        (exam_id, student_id, student_name, class_id, subject, score, leave_status, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', rows_to_insert)
        
        conn.commit()
        
        return jsonify({
            'status': 'ok',
            'message': '成绩上传成功',
            'records_count': len(rows_to_insert)
        })
    except Exception as e:
        logger.error(f"上传考试成绩时出错: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'上传考试成绩失败: {str(e)}'
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()

# 预览考试成绩Excel
@grade_analysis_bp.route('/api/exams/preview', methods=['POST'])
@login_required
def preview_scores():
    try:
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': '没有上传文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': '未选择文件'}), 400
        
        # 检查文件格式
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'status': 'error', 'message': '只支持Excel文件(.xlsx, .xls)'}), 400
        
        # 保存文件
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"{timestamp}_{secure_filename(file.filename)}"
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        
        # 确保上传目录存在
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        file.save(file_path)
        
        # 解析Excel文件
        df = pd.read_excel(file_path)
        
        # 检查必要的列是否存在
        required_columns = ['学号', '姓名']
        for col in required_columns:
            if col not in df.columns:
                return jsonify({'status': 'error', 'message': f'缺少必要的列: {col}'}), 400
        
        # 提取学科列(除了学号和姓名)
        subject_columns = [col for col in df.columns if col not in required_columns]
        
        # 构建预览数据
        preview_data = []
        for _, row in df.iterrows():
            student_data = {
                'student_id': str(row['学号']),
                'student_name': row['姓名'],
                'scores': {}
            }
            
            for subject in subject_columns:
                if pd.notna(row[subject]):
                    score_value = row[subject]
                    leave_status = False  # 默认非请假状态
                    
                    # 处理成绩值
                    try:
                        # 尝试转换为浮点数
                        score = float(score_value)
                        # 如果成绩为0，视为请假
                        if score == 0:
                            leave_status = True
                    except (ValueError, TypeError):
                        # 如果无法转换为浮点数，检查是否为"请假"文本
                        if isinstance(score_value, str) and "请假" in score_value:
                            score = 0  # 请假学生分数记为0
                            leave_status = True
                        else:
                            # 无法识别的成绩，设为None
                            score = None
                    
                    student_data['scores'][subject] = {
                        'score': score,
                        'leave_status': leave_status
                    }
                else:
                    student_data['scores'][subject] = {
                        'score': None,
                        'leave_status': False
                    }
            
            preview_data.append(student_data)
        
        return jsonify({
            'status': 'ok',
            'preview': preview_data,
            'subjects': subject_columns,
            'file_path': file_path
        })
    except Exception as e:
        logger.error(f"预览考试成绩时出错: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'预览考试成绩失败: {str(e)}'
        }), 500
    finally:
        pass  # 不关闭连接，因为没有使用数据库

# 下载成绩模板
@grade_analysis_bp.route('/api/exams/template', methods=['GET'])
@login_required
def download_template():
    try:
        # 获取班级ID
        class_id = request.args.get('class_id')
        subjects = request.args.get('subjects', '').split(',')
        
        if not class_id:
            if not current_user.is_admin:
                if not current_user.class_id:
                    return jsonify({'status': 'error', 'message': '您尚未被分配班级'}), 403
                class_id = current_user.class_id
            else:
                return jsonify({'status': 'error', 'message': '管理员需提供班级ID参数'}), 400
        
        # 获取班级学生信息
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT id, name FROM students WHERE class_id = ?", (class_id,))
        students = cursor.fetchall()
        
        # 创建Excel模板
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "成绩导入模板"
        
        # 添加标题行
        headers = ['学号', '姓名'] + subjects
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header)
        
        # 添加学生数据
        for i, student in enumerate(students, 2):
            ws.cell(row=i, column=1, value=student['id'])
            ws.cell(row=i, column=2, value=student['name'])
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 返回文件
        return send_file(
            output,
            as_attachment=True,
            download_name=f"成绩导入模板_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"下载成绩模板时出错: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'下载成绩模板失败: {str(e)}'
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()

# 获取班级学生列表
@grade_analysis_bp.route('/api/exams/students', methods=['GET'])
@login_required
def get_class_students():
    try:
        class_id = request.args.get('class_id')
        
        if not class_id:
            if not current_user.is_admin:
                if not current_user.class_id:
                    return jsonify({'status': 'error', 'message': '您尚未被分配班级'}), 403
                class_id = current_user.class_id
            else:
                return jsonify({'status': 'error', 'message': '管理员需提供班级ID参数'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT id, name FROM students WHERE class_id = ? ORDER BY id", (class_id,))
        students = [dict(row) for row in cursor.fetchall()]
        
        return jsonify({
            'status': 'ok',
            'students': students
        })
    except Exception as e:
        logger.error(f"获取班级学生列表时出错: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'获取班级学生列表失败: {str(e)}'
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()

# 删除考试
@grade_analysis_bp.route('/api/exams/<int:exam_id>', methods=['DELETE'])
@login_required
def delete_exam(exam_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取考试信息
        cursor.execute("SELECT * FROM exams WHERE id = ?", (exam_id,))
        exam = cursor.fetchone()
        
        if not exam:
            return jsonify({'status': 'error', 'message': '考试不存在'}), 404
        
        # 检查权限
        if not current_user.is_admin and current_user.class_id != exam['class_id']:
            return jsonify({'status': 'error', 'message': '您无权删除此考试'}), 403
        
        # 删除考试及相关成绩
        cursor.execute("DELETE FROM exams WHERE id = ?", (exam_id,))
        
        conn.commit()
        
        return jsonify({
            'status': 'ok',
            'message': '考试删除成功'
        })
    except Exception as e:
        logger.error(f"删除考试时出错: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'删除考试失败: {str(e)}'
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()

# 更新考试
@grade_analysis_bp.route('/api/exams/<int:exam_id>', methods=['PUT', 'POST'])
@login_required
def update_exam(exam_id):
    try:
        data = request.json
        
        # 检查是否是POST请求但意图是PUT
        if request.method == 'POST':
            # 检查请求头或请求体中是否有方法覆盖的标识
            method_override = request.headers.get('X-HTTP-Method-Override')
            method_param = data.get('_method')
            
            # 如果不是PUT意图，则拒绝请求
            if method_override != 'PUT' and method_param != 'PUT':
                return jsonify({'status': 'error', 'message': '不支持POST方法更新考试'}), 405
            
            # 如果是覆盖方法，从请求数据中移除_method参数
            if '_method' in data:
                del data['_method']
        
        # 验证必填字段
        required_fields = ['exam_name', 'exam_date', 'subjects']
        for field in required_fields:
            if field not in data:
                return jsonify({'status': 'error', 'message': f'缺少字段: {field}'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取考试信息
        cursor.execute("SELECT * FROM exams WHERE id = ?", (exam_id,))
        exam = cursor.fetchone()
        
        if not exam:
            return jsonify({'status': 'error', 'message': '考试不存在'}), 404
        
        # 检查权限
        if not current_user.is_admin and current_user.class_id != exam['class_id']:
            return jsonify({'status': 'error', 'message': '您无权修改此考试'}), 403
        
        # 准备更新数据
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 序列化subjects字段
        subjects_json = json.dumps(data['subjects'], ensure_ascii=False)
        
        # 更新考试数据
        cursor.execute('''
        UPDATE exams 
        SET exam_name = ?, exam_date = ?, subjects = ?, updated_at = ?
        WHERE id = ?
        ''', (data['exam_name'], data['exam_date'], subjects_json, now, exam_id))
        
        conn.commit()
        
        return jsonify({
            'status': 'ok',
            'message': '考试更新成功'
        })
    except Exception as e:
        logger.error(f"更新考试时出错: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'更新考试失败: {str(e)}'
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()

# 更新学生成绩
@grade_analysis_bp.route('/api/exams/<int:exam_id>/scores/update', methods=['POST'])
@login_required
def update_student_score(exam_id):
    try:
        data = request.json
        
        # 验证必填字段
        required_fields = ['student_id', 'subject', 'score']
        for field in required_fields:
            if field not in data:
                return jsonify({'status': 'error', 'message': f'缺少字段: {field}'}), 400
        
        # 获取数据
        student_id = data['student_id']
        subject = data['subject']
        score = data['score']
        leave_status = data.get('leave_status', 0)  # 默认为0表示正常
        
        # 验证分数是否有效 (请假状态的成绩为0)
        if leave_status == 0 and (score < 0 or score > 100):
            return jsonify({'status': 'error', 'message': '分数必须在0-100之间'}), 400
        
        # 获取考试信息
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM exams WHERE id = ?", (exam_id,))
        exam = cursor.fetchone()
        
        if not exam:
            return jsonify({'status': 'error', 'message': '考试不存在'}), 404
        
        # 检查权限
        if not current_user.is_admin and current_user.class_id != exam['class_id']:
            return jsonify({'status': 'error', 'message': '您无权修改此考试成绩'}), 403
        
        # 检查学生是否属于该班级
        cursor.execute("SELECT id FROM students WHERE id = ? AND class_id = ?", 
                       (student_id, exam['class_id']))
        student = cursor.fetchone()
        
        if not student:
            return jsonify({'status': 'error', 'message': '学生不属于此考试班级'}), 400
        
        # 获取学生姓名
        cursor.execute("SELECT name FROM students WHERE id = ?", (student_id,))
        student_name = cursor.fetchone()['name']
        
        # 更新成绩
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 检查成绩记录是否存在
        cursor.execute('''
        SELECT * FROM exam_scores 
        WHERE exam_id = ? AND student_id = ? AND subject = ?
        ''', (exam_id, student_id, subject))
        
        score_record = cursor.fetchone()
        
        if score_record:
            # 更新现有记录
            cursor.execute('''
            UPDATE exam_scores 
            SET score = ?, leave_status = ?, updated_at = ? 
            WHERE exam_id = ? AND student_id = ? AND subject = ?
            ''', (score, leave_status, now, exam_id, student_id, subject))
        else:
            # 插入新记录
            cursor.execute('''
            INSERT INTO exam_scores 
            (exam_id, student_id, student_name, class_id, subject, score, leave_status, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (exam_id, student_id, student_name, exam['class_id'], subject, score, leave_status, now, now))
        
        conn.commit()
        
        return jsonify({
            'status': 'ok',
            'message': '成绩更新成功'
        })
    except Exception as e:
        logger.error(f"更新学生成绩时出错: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'更新学生成绩失败: {str(e)}'
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()

# 删除学生成绩
@grade_analysis_bp.route('/api/exams/<int:exam_id>/scores/delete', methods=['POST'])
@login_required
def delete_student_scores(exam_id):
    try:
        data = request.json
        
        # 验证必填字段
        if 'student_id' not in data:
            return jsonify({'status': 'error', 'message': '缺少学生ID字段'}), 400
        
        # 获取考试信息
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM exams WHERE id = ?", (exam_id,))
        exam = cursor.fetchone()
        
        if not exam:
            return jsonify({'status': 'error', 'message': '考试不存在'}), 404
        
        # 检查权限
        if not current_user.is_admin and current_user.class_id != exam['class_id']:
            return jsonify({'status': 'error', 'message': '您无权删除此考试成绩'}), 403
        
        # 删除学生的所有成绩记录
        cursor.execute('''
        DELETE FROM exam_scores 
        WHERE exam_id = ? AND student_id = ?
        ''', (exam_id, data['student_id']))
        
        # 获取影响的行数
        row_count = cursor.rowcount
        
        conn.commit()
        
        if row_count == 0:
            return jsonify({
                'status': 'warning',
                'message': '未找到需要删除的成绩记录'
            })
        
        return jsonify({
            'status': 'ok',
            'message': f'成功删除 {row_count} 条成绩记录'
        })
    except Exception as e:
        logger.error(f"删除学生成绩时出错: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'删除学生成绩失败: {str(e)}'
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()

# 获取多个考试详情，用于对比分析
@grade_analysis_bp.route('/api/exams/compare', methods=['POST'])
@login_required
def get_exams_for_comparison():
    try:
        data = request.json
        
        # 检查考试ID列表
        if not data or 'exam_ids' not in data or not data['exam_ids']:
            return jsonify({'status': 'error', 'message': '请选择要对比的考试'}), 400
        
        exam_ids = data['exam_ids']
        
        # 准备结果容器
        results = {
            'exams': [],
            'subjects': [],
            'students': {},
            'common_subjects': []
        }
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取班级ID (如果是管理员，以第一个考试的班级为准)
        class_id = None
        if not current_user.is_admin:
            class_id = current_user.class_id
        
        # 获取考试详情
        all_subjects = set()
        valid_exam_ids = []
        
        for exam_id in exam_ids:
            # 获取考试信息
            cursor.execute("SELECT * FROM exams WHERE id = ?", (exam_id,))
            exam = cursor.fetchone()
            
            if not exam:
                continue
                
            exam_dict = dict(exam)
            exam_dict['subjects'] = json.loads(exam_dict['subjects'])
            
            # 如果是第一个考试并且是管理员，设置班级ID
            if not class_id and current_user.is_admin:
                class_id = exam_dict['class_id']
            
            # 检查权限
            if exam_dict['class_id'] != class_id:
                continue
                
            # 获取成绩统计
            cursor.execute('''
            SELECT s.*, students.name as student_name, students.class as class_name
            FROM exam_scores s
            LEFT JOIN students ON s.student_id = students.id
            WHERE s.exam_id = ? AND s.class_id = ?
            ORDER BY s.student_id, s.subject
            ''', (exam_id, class_id))
            
            all_scores = [dict(row) for row in cursor.fetchall()]
            
            # 按学科组织成绩数据
            scores_by_subject = {}
            for score in all_scores:
                subject = score['subject']
                if subject not in scores_by_subject:
                    scores_by_subject[subject] = []
                scores_by_subject[subject].append(score)
                
                # 添加到全局学科集合
                all_subjects.add(subject)
            
            # 计算统计信息
            stats = {}
            for subject, subject_scores in scores_by_subject.items():
                # 提取分数列表
                subject_student_ids = set()
                valid_scores = []
                
                # 存储学生ID和分数的映射
                student_scores = {}
                
                for score in subject_scores:
                    student_id = score['student_id']
                    if student_id in subject_student_ids:
                        continue  # 跳过已经计算过的学生
                    
                    subject_student_ids.add(student_id)
                    score_value = score['score']
                    
                    # 添加到学生成绩映射
                    student_scores[student_id] = {
                        'student_id': student_id,
                        'student_name': score['student_name'],
                        'score': score_value
                    }
                    
                    # 验证分数是否有效 - 修复类型比较问题
                    try:
                        # 检查是否为请假状态 - 请假状态的分数为0且leave_status为1
                        if score.get('leave_status', 0) == 1 or score_value == 0:
                            # 请假学生不计入有效分数统计
                            continue
                        
                        # 确保分数是数字且在有效范围内
                        if isinstance(score_value, (int, float)) and 0 < score_value <= 100:
                            valid_scores.append(score_value)
                    except TypeError:
                        # 如果比较出现类型错误，记录日志并继续
                        logger.warning(f"分数类型错误: {score_value}, 类型: {type(score_value)}")
                        continue
                
                # 添加学生成绩到结果中
                for student_id, score_info in student_scores.items():
                    if student_id not in results['students']:
                        results['students'][student_id] = {
                            'student_id': student_id,
                            'student_name': score_info['student_name'],
                            'scores': {}
                        }
                    
                    if 'scores' not in results['students'][student_id]:
                        results['students'][student_id]['scores'] = {}
                        
                    if exam_id not in results['students'][student_id]['scores']:
                        results['students'][student_id]['scores'][exam_id] = {}
                    
                    results['students'][student_id]['scores'][exam_id][subject] = score_info['score']
                
                # 获取该科目的学生总数 (去重后)
                total_students = len(valid_scores)
                
                if total_students == 0:
                    stats[subject] = {
                        'average': 0,
                        'max': 0,
                        'min': 0,
                        'excellent_rate': 0,
                        'pass_rate': 0,
                        'excellent_threshold': 90,
                        'score_distribution': {'0-59': 0, '60-69': 0, '70-79': 0, '80-89': 0, '90-100': 0}
                    }
                    continue
                
                # 计算基本统计信息
                avg_score = round(sum(valid_scores) / total_students, 1)
                max_score = max(valid_scores)
                min_score = min(valid_scores)
                
                # 计算及格率和优秀率
                pass_count = sum(1 for s in valid_scores if s >= 60)
                excellent_threshold = 90  # 默认优秀标准
                excellent_count = sum(1 for s in valid_scores if s >= excellent_threshold)
                
                pass_rate = round((pass_count / total_students) * 100, 1)
                excellent_rate = round((excellent_count / total_students) * 100, 1)
                
                # 计算各分数段人数
                below_60 = sum(1 for s in valid_scores if s < 60)
                from_60_to_69 = sum(1 for s in valid_scores if 60 <= s < 70)
                from_70_to_79 = sum(1 for s in valid_scores if 70 <= s < 80)
                from_80_to_89 = sum(1 for s in valid_scores if 80 <= s < 90)
                from_90_to_100 = sum(1 for s in valid_scores if s >= 90)
                
                # 保存统计数据
                stats[subject] = {
                    'average': avg_score,
                    'max': max_score,
                    'min': min_score,
                    'pass_rate': pass_rate,
                    'excellent_rate': excellent_rate,
                    'excellent_threshold': excellent_threshold,
                    'score_distribution': {
                        '0-59': below_60,
                        '60-69': from_60_to_69,
                        '70-79': from_70_to_79,
                        '80-89': from_80_to_89,
                        '90-100': from_90_to_100
                    }
                }
            
            # 添加考试和统计信息到结果
            valid_exam_ids.append(exam_id)
            results['exams'].append({
                'id': exam_id,
                'name': exam_dict['exam_name'],
                'date': exam_dict['exam_date'],
                'subjects': exam_dict['subjects'],
                'stats': stats
            })
        
        # 转换学生数据为列表
        results['students'] = list(results['students'].values())
        
        # 获取共有学科
        common_subjects = set()
        if results['exams']:
            common_subjects = set(results['exams'][0]['stats'].keys())
            for exam in results['exams'][1:]:
                common_subjects = common_subjects.intersection(set(exam['stats'].keys()))
        
        results['common_subjects'] = list(common_subjects)
        results['all_subjects'] = list(all_subjects)
        
        return jsonify({
            'status': 'ok',
            'results': results
        })
    except Exception as e:
        logger.error(f"获取对比分析数据时出错: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'获取对比分析数据失败: {str(e)}'
        }), 500
    finally:
        if 'conn' in locals():
            conn.close() 