# -*- coding: utf-8 -*-
from flask import Blueprint, request, jsonify, send_from_directory, current_app, make_response, send_file
from flask_login import current_user
import sqlite3
import os
import datetime
import traceback
import logging
import tempfile
import zipfile
import sys
import platform
import threading
import time
import re
from werkzeug.utils import secure_filename
import urllib.parse

# 尝试导入pythoncom，如果不可用则跳过
try:
    import pythoncom
    PYTHONCOM_AVAILABLE = True
except ImportError:
    PYTHONCOM_AVAILABLE = False
    print("警告: pythoncom模块不可用，Word文档转换功能将不可用")

# 导入评语相关的工具类
from utils.comment_processor import batch_update_comments, generate_comments_pdf, generate_preview_html
from utils.comment_generator import CommentGenerator
try:
    from utils.pdf_exporter_fixed import export_comments_to_pdf
except ImportError:
    # 创建一个简化版的PDF导出函数，返回格式与正常函数一致
    def export_comments_to_pdf(*args, **kwargs):
        return {
            'status': 'error',
            'message': 'PDF导出功能不可用，请安装reportlab模块'
        }
    print("! PDF导出功能不可用，请安装reportlab模块")

# 导入Excel相关库
import pandas as pd
import numpy as np
import json
from io import BytesIO

# 创建评语管理蓝图
comments_bp = Blueprint('comments', __name__)
logger = logging.getLogger(__name__)

# 添加导出进度全局变量 - 改为基于用户的进度跟踪
user_export_progress = {}
progress_lock = threading.Lock()

# 更新导出进度信息
def websocket_progress(message, percent=None, request_id=None):
    """
    更新导出进度信息
    
    参数:
    - message: 进度消息内容
    - percent: 百分比值(0-100)
    - request_id: 请求ID
    """
    global user_export_progress
    
    # 获取当前用户ID，如果没有request_id则使用用户ID
    user_id = None
    if hasattr(current_user, 'id') and current_user.is_authenticated:
        user_id = str(current_user.id)
    
    # 使用request_id作为主键，如果没有则使用user_id
    progress_key = request_id if request_id else user_id
    if not progress_key:
        progress_key = 'anonymous'
    
    # 线程安全地更新进度信息
    with progress_lock:
        current_time = time.time()
        if progress_key not in user_export_progress:
            user_export_progress[progress_key] = {}
        
        user_export_progress[progress_key].update({
            'message': message,
            'percent': percent if percent is not None else user_export_progress[progress_key].get('percent', 0),
            'timestamp': current_time,
            'request_id': request_id,
            'user_id': user_id
        })
    
    # 打印日志
    logger.info(f"导出进度更新 [用户:{user_id}, 请求:{request_id}]: {message} ({percent if percent is not None else ''}%)")
    
    return user_export_progress.get(progress_key, {})

# 获取数据库连接函数
def get_db_connection():
    DATABASE = 'students.db'
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')  # 启用外键约束
    return conn

# 获取单个学生评语
@comments_bp.route('/api/comments/<student_id>', methods=['GET'], strict_slashes=False)
def get_student_comment(student_id):
    try:
        # 获取班级ID，优先使用请求参数中的class_id，如果没有则使用当前用户的class_id
        class_id = request.args.get('class_id')
        if not class_id:
            class_id = current_user.class_id
        
        # 移除类型转换，但增加参数类型日志
        logger.info(f"获取评语: 学生ID={student_id} (类型: {type(student_id)}), 班级ID={class_id} (类型: {type(class_id)})")
        
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute('SELECT id, name, comments FROM students WHERE id = ? AND class_id = ?', (student_id, class_id))
            student = cursor.fetchone()
            
            conn.close()
        except Exception as e:
            logger.error(f"获取评语数据库查询错误: {str(e)}")
            return jsonify({
                'status': 'error',
                'message': f'数据库查询错误: {str(e)}'
            }), 500
        
        if not student:
            logger.error(f"未找到学生评语: 学生ID={student_id}, 班级ID={class_id}")
            return jsonify({
                'status': 'error',
                'message': f'未找到ID为 {student_id} 的学生，或学生不在班级 {class_id} 中'
            }), 404
        
        logger.info(f"成功获取评语: 学生ID={student_id}, 班级ID={class_id}, 学生名称={student['name']}")
        return jsonify({
            'status': 'ok',
            'comment': {
                'studentId': student['id'],
                'studentName': student['name'],
                'content': student['comments'] or ''
            }
        })
    except Exception as e:
        logger.error(f"获取评语时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'status': 'error',
            'message': f'获取评语时出错: {str(e)}'
        }), 500

# 保存学生评语
@comments_bp.route('/api/comments', methods=['POST'], strict_slashes=False)
def save_student_comment():
    data = request.json
    
    # 详细记录请求信息以便调试
    logger.info(f"收到保存评语请求: {data}")
    
    if not data or 'studentId' not in data or 'content' not in data:
        logger.error(f"请求数据验证失败: {data}")
        return jsonify({
            'status': 'error',
            'message': '请提供学生ID和评语内容'
        }), 400
        
    student_id = data['studentId']
    content = data['content']
    append_mode = data.get('appendMode', False)
    
    # 获取班级ID，优先使用请求中的classId，如果没有则使用当前用户的class_id
    class_id = data.get('classId')
    if not class_id:
        class_id = current_user.class_id
    
    # 移除类型转换，但添加类型记录以便调试
    logger.info(f"保存学生评语, 学生ID: {student_id} (类型: {type(student_id)}), 班级ID: {class_id} (类型: {type(class_id)}), 追加模式: {append_mode}")
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 修改权限检查逻辑，允许任何用户使用AI生成评语
        # 检查是否是AI评语请求（通过判断请求来源或特定标志）
        is_ai_comment = data.get('isAIComment', False)
        
        # 检查当前用户是否有权限修改该班级的评语
        # 如果是AI评语，或用户是管理员，或这是用户自己班级的学生，则允许操作
        if not is_ai_comment and not current_user.is_admin and class_id != current_user.class_id:
            conn.close()
            logger.error(f"权限检查失败: 用户班级ID={current_user.class_id}, 请求班级ID={class_id}")
            return jsonify({
                'status': 'error',
                'message': '您没有权限修改其他班级学生的评语'
            }), 403
        
        # 先检查学生是否存在于指定班级
        try:
            cursor.execute('SELECT id, comments FROM students WHERE id = ? AND class_id = ?', (student_id, class_id))
            student = cursor.fetchone()
        except Exception as e:
            conn.close()
            logger.error(f"SQL查询错误: {str(e)}, SQL: 'SELECT id, comments FROM students WHERE id = ? AND class_id = ?', 参数: ({student_id}, {class_id})")
            return jsonify({
                'status': 'error',
                'message': f'数据库查询错误: {str(e)}'
            }), 500
        
        if not student:
            conn.close()
            logger.error(f"未找到学生: 学生ID={student_id}, 班级ID={class_id}")
            return jsonify({
                'status': 'error',
                'message': '未找到该班级中的学生'
            }), 404
        
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        try:
            # 如果是追加模式且已有评语，则在原有评语基础上添加新内容
            if append_mode and student['comments']:
                # 在原评语后添加新评语（添加时间戳和分隔符）
                updated_content = f"{student['comments']}\n\n--- {now} ---\n{content}"
            else:
                # 直接使用新评语
                updated_content = content
                
            # 更新学生评语，确保只更新指定班级的学生记录
            cursor.execute('UPDATE students SET comments = ?, updated_at = ? WHERE id = ? AND class_id = ?', 
                          (updated_content, now, student_id, class_id))
            conn.commit()
            
            logger.info(f"评语保存成功: 学生ID={student_id}, 班级ID={class_id}")
            return jsonify({
                'status': 'ok',
                'message': '评语保存成功',
                'updatedContent': updated_content,
                'updateDate': now
            })
        except Exception as e:
            conn.rollback()
            logger.error(f"更新评语SQL错误: {str(e)}, SQL: 'UPDATE students SET comments = ?, updated_at = ? WHERE id = ? AND class_id = ?', 参数: ({len(updated_content)}字节, {now}, {student_id}, {class_id})")
            return jsonify({
                'status': 'error',
                'message': f'保存评语时出错: {str(e)}'
            }), 500
    except Exception as e:
        logger.error(f"保存评语时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'status': 'error',
            'message': f'服务器内部错误: {str(e)}'
        }), 500
    finally:
        if 'conn' in locals() and conn:
            conn.close()

# 获取评语模板API
@comments_bp.route('/api/comment-templates', methods=['GET'])
def get_comment_templates():
    """
    获取预定义的评语模板列表
    """
    templates = [
        # 学习类评语模板
        {"id": 1, "title": "品德优良", "content": "品德优良，尊敬师长，团结同学。", "type": "study"},
        {"id": 2, "title": "学习优秀", "content": "学习刻苦认真，上课认真听讲，作业按时完成。", "type": "study"},
        {"id": 3, "title": "学习积极", "content": "学习态度积极，能够主动思考，乐于探索新知识。", "type": "study"},
        {"id": 4, "title": "学习进步", "content": "近期学习有明显进步，在班级表现积极。", "type": "study"},
        {"id": 5, "title": "成绩优异", "content": "各科成绩优异，是班级的学习标兵。", "type": "study"},
        
        # 体育类评语模板
        {"id": 6, "title": "身体健康", "content": "身体健康，积极参加体育活动。", "type": "physical"},
        {"id": 7, "title": "运动技能", "content": "运动能力强，在体育活动中表现出色。", "type": "physical"},
        {"id": 8, "title": "体育精神", "content": "在体育活动中展现了团队合作精神和拼搏精神。", "type": "physical"},
        
        # 行为类评语模板
        {"id": 9, "title": "全面发展", "content": "德智体美劳全面发展，综合素质优秀。", "type": "behavior"},
        {"id": 10, "title": "行为规范", "content": "行为规范，能够严格遵守校规校纪。", "type": "behavior"},
        {"id": 11, "title": "积极参与", "content": "积极参与班级和学校活动，热心为集体服务。", "type": "behavior"},
        {"id": 12, "title": "有进步空间", "content": "在学习上有进步空间，希望能更加努力。", "type": "behavior"}
    ]
    
    return jsonify({
        "status": "ok",
        "templates": templates
    })

# 批量更新评语
@comments_bp.route('/api/batch-update-comments', methods=['POST'])
def batch_update_comments_api():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'status': 'error', 'message': '未接收到数据'})
        
        content = data.get('content', '').strip()
        append_mode = data.get('appendMode', True)
        student_ids = data.get('studentIds', [])  # 获取选中的学生ID列表
        
        if not content:
            return jsonify({'status': 'error', 'message': '评语内容不能为空'})
        
        if not student_ids:
            return jsonify({'status': 'error', 'message': '未选择学生'})
        
        # 确保class_id是整数
        try:
            class_id = int(current_user.class_id)
        except (TypeError, ValueError):
            logger.error(f"批量更新评语时班级ID类型转换失败: {current_user.class_id}")
            return jsonify({'status': 'error', 'message': '班级ID格式不正确'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 记录更新的学生数量
        updated_count = 0
        
        # 获取当前时间，但不再添加到评语中
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 对选中的每个学生更新评语
        for student_id in student_ids:
            # 查询学生当前评语，确保只查询当前班级的学生
            cursor.execute('SELECT comments FROM students WHERE id = ? AND class_id = ?', (student_id, class_id))
            student = cursor.fetchone()
            
            if not student:
                continue  # 跳过不存在的学生ID或不属于当前班级的学生
            
            current_comment = student[0] or ''
            
            # 根据模式设置新评语
            if append_mode and current_comment:
                # 简化追加模式，直接添加到末尾
                new_comment = f"{current_comment.strip()}\n\n{content}"
            else:
                # 如果是替换模式或无评语，则直接使用新内容
                new_comment = content
            
            # 更新学生评语和更新时间，确保只更新当前班级的学生
            cursor.execute('UPDATE students SET comments = ?, updated_at = ? WHERE id = ? AND class_id = ?', 
                          (new_comment, now, student_id, class_id))
            
            updated_count += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'status': 'ok', 
            'message': f'已成功更新 {updated_count} 名学生的评语'
        })
        
    except Exception as e:
        logger.error(f"批量更新评语时出错: {str(e)}")
        return jsonify({'status': 'error', 'message': f'服务器错误: {str(e)}'})

# 清除当前班级所有评语
@comments_bp.route('/api/clear-all-comments', methods=['POST'])
def clear_all_comments():
    try:
        data = request.get_json()
        if not data:
            logger.error("清除评语 - 未接收到数据")
            return jsonify({'status': 'error', 'message': '未接收到数据'})
        
        # 获取班级ID
        class_id = data.get('class_id')
        
        # 记录班级ID的类型和值，帮助调试
        logger.info(f"清除评语 - 接收到的班级ID: {class_id}, 类型: {type(class_id)}")
        
        # 确保class_id是整数
        try:
            class_id = int(class_id)
            logger.info(f"清除评语 - 转换后的班级ID: {class_id}")
        except (TypeError, ValueError):
            logger.error(f"清除评语时班级ID类型转换失败: {class_id}")
            return jsonify({'status': 'error', 'message': f'班级ID格式不正确: {class_id}'})
        
        # 获取数据库连接
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # 获取当前时间
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # 先获取班级中有评语的学生数量
            cursor.execute('SELECT COUNT(*) FROM students WHERE class_id = ? AND (comments IS NOT NULL AND comments <> "")', (class_id,))
            affected_count = cursor.fetchone()[0]
            
            logger.info(f"清除评语 - 班级ID {class_id} 中有评语的学生数量: {affected_count}")
            
            # 调试：获取班级中的学生总数
            cursor.execute('SELECT COUNT(*) FROM students WHERE class_id = ?', (class_id,))
            total_students = cursor.fetchone()[0]
            logger.info(f"清除评语 - 班级ID {class_id} 中的学生总数: {total_students}")
            
            # 调试：获取所有班级ID和对应的学生数量
            cursor.execute('SELECT class_id, COUNT(*) as student_count FROM students GROUP BY class_id')
            all_classes = cursor.fetchall()
            logger.info(f"清除评语 - 数据库中的所有班级ID和学生数量: {all_classes}")
            
            # 如果没有找到任何学生，尝试查询这个班级是否存在
            if total_students == 0:
                cursor.execute('SELECT DISTINCT class_id FROM students')
                all_class_ids = [row[0] for row in cursor.fetchall()]
                logger.warning(f"清除评语 - 班级ID {class_id} 没有学生。数据库中存在的班级ID: {all_class_ids}")
                return jsonify({
                    'status': 'error', 
                    'message': f'班级ID {class_id} 不存在或没有学生',
                    'available_class_ids': all_class_ids
                })
            
            # 清除指定班级所有学生的评语 - 使用空字符串而不是NULL，确保兼容性
            update_sql = 'UPDATE students SET comments = "", updated_at = ? WHERE class_id = ?'
            logger.info(f"清除评语 - 执行SQL: {update_sql}, 参数: ({now}, {class_id})")
            
            cursor.execute(update_sql, (now, class_id))
            
            # 检查影响的行数
            rows_affected = cursor.rowcount
            logger.info(f"清除评语 - 实际影响的行数: {rows_affected}")
            
            # 验证是否真的清除了评语
            cursor.execute('SELECT COUNT(*) FROM students WHERE class_id = ? AND (comments IS NOT NULL AND comments <> "")', (class_id,))
            remaining_comments = cursor.fetchone()[0]
            logger.info(f"清除评语 - 操作后仍有评语的学生数量: {remaining_comments}")
            
            conn.commit()
        except Exception as e:
            conn.rollback()
            logger.error(f"清除评语 - 数据库操作失败: {str(e)}")
            raise
        finally:
            conn.close()
        
        # 如果没有影响任何行，但有学生有评语，可能是SQL执行问题
        if rows_affected == 0 and affected_count > 0:
            logger.warning(f"清除评语 - 有{affected_count}名学生有评语，但SQL影响了0行")
            return jsonify({
                'status': 'error', 
                'message': f'清除评语失败，请检查数据库权限',
                'affected_count': 0
            })
        
        return jsonify({
            'status': 'ok', 
            'message': f'已成功清除 {affected_count} 名学生的评语',
            'affected_count': affected_count
        })
        
    except Exception as e:
        logger.error(f"清除评语时出错: {str(e)}")
        logger.error(traceback.format_exc())  # 添加堆栈跟踪以便更好地调试
        return jsonify({'status': 'error', 'message': f'服务器错误: {str(e)}'})

# AI生成评语API
@comments_bp.route('/api/generate-comment', methods=['POST'])
def generate_comment():
    """生成AI评语"""
    try:
        # 获取请求数据
        data = request.get_json()
        logger.info(f"收到评语生成请求: {data}")
        
        # 获取全局变量
        deepseek_api = current_app.config.get('deepseek_api')
        
        # 验证请求数据
        comment_generator = CommentGenerator(deepseek_api.api_key if deepseek_api else None)
        is_valid, error_msg = comment_generator.validate_request(data)
        if not is_valid:
            logger.error(f"请求数据验证失败: {error_msg}")
            return jsonify({
                "status": "error",
                "message": error_msg
            })
        
        # 获取学生信息
        student_id = data.get('student_id') or data.get('studentId')
        if not student_id:
            logger.error("缺少学生ID")
            return jsonify({
                "status": "error",
                "message": "缺少学生ID"
            })
            
        # 获取班级ID，优先使用请求中的class_id，如果没有则使用当前用户的class_id
        class_id = data.get('class_id')
        if not class_id:
            class_id = current_user.class_id
            
        # 移除类型转换，但增加参数类型日志
        logger.info(f"生成评语参数: 学生ID={student_id} (类型: {type(student_id)}), 班级ID={class_id} (类型: {type(class_id)})")
        
        # 从数据库获取学生信息
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT name, gender FROM students WHERE id = ? AND class_id = ?', (student_id, class_id))
            student = cursor.fetchone()
            conn.close()
        except Exception as e:
            logger.error(f"数据库查询错误: {str(e)}")
            return jsonify({
                "status": "error",
                "message": f"数据库查询错误: {str(e)}"
            }), 500
        
        if not student:
            logger.error(f"未找到学生: 学生ID={student_id}, 班级ID={class_id}")
            return jsonify({
                "status": "error",
                "message": f"未找到ID为 {student_id} 的学生或该学生不在班级 {class_id} 中"
            })
            
        # 构建学生信息字典
        student_info = {
            "name": student[0],
            "gender": student[1],
            "personality": data.get('personality', ''),
            "study_performance": data.get('study_performance', ''),
            "hobbies": data.get('hobbies', ''),
            "improvement": data.get('improvement', '')
        }
        
        # 获取评语参数
        style = data.get('style', '鼓励性的')
        tone = data.get('tone', '正式的')
        # 强制设置最大字数为260，忽略前端传入的值
        max_length = 50000  # 无论前端传入什么值，都固定为260
        min_length = 200  # 最小字数固定为200
        additional_instructions = data.get('additional_instructions', '')
        
        # 如果有额外指令，添加到学生信息中
        if additional_instructions:
            student_info['additional_instructions'] = additional_instructions
        
        # 记录正在为哪个学生生成评语
        logger.info(f"正在为学生 {student_info['name']}(ID: {student_id}, 班级ID: {class_id}) 生成评语")
        
        # 生成评语
        try:
            result = comment_generator.generate_comment(
                student_info=student_info,
                style=style,
                tone=tone,
                max_length=max_length,
                min_length=min_length  # 添加最小字数参数
            )
        except Exception as e:
            logger.error(f"评语生成引擎错误: {str(e)}")
            return jsonify({
                "status": "error",
                "message": f"评语生成失败: {str(e)}"
            }), 500
        
        logger.info(f"评语生成结果(学生ID: {student_id}, 班级ID: {class_id}): {result.get('status')}")
        
        # 返回结果
        if result["status"] == "ok":
            return jsonify({
                "status": "ok",
                "comment": result["comment"],
                "student_id": student_id,
                "class_id": class_id,
                "reasoning_content": result.get("reasoning_content", ""),
                "content_field": result.get("content_field", "")
            })
        else:
            return jsonify({
                "status": "error",
                "message": result["message"],
                "student_id": student_id,
                "class_id": class_id
            })
                
    except Exception as e:
        logger.error(f"生成评语时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            "status": "error",
            "message": f"生成评语时出错: {str(e)}"
        }), 500

# 导出评语为PDF
@comments_bp.route('/api/export-comments-pdf', methods=['GET'])
def api_export_comments_pdf():
    # 获取班级参数（可选）
    class_name = request.args.get('class')
    logger.info(f"收到导出评语PDF请求，班级: {class_name}")
    
    try:
        # 检查用户权限
        if not hasattr(current_user, 'is_authenticated') or not current_user.is_authenticated:
            return jsonify({'status': 'error', 'message': '您需要登录后才能导出评语', 'code': 'login_required'}), 401
            
        # 如果是班主任，检查是否有权限导出指定班级
        if hasattr(current_user, 'is_admin') and not current_user.is_admin and hasattr(current_user, 'class_id') and current_user.class_id:
            # 如果未指定班级，则使用班主任自己的班级
            if not class_name:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute('SELECT class FROM students WHERE class_id = ? LIMIT 1', (current_user.class_id,))
                result = cursor.fetchone()
                conn.close()
                
                if result:
                    class_name = result['class']
                    logger.info(f"未指定班级，使用班主任班级: {class_name}")
            # 指定了班级，检查是否有权限
            elif class_name:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute('SELECT class FROM students WHERE class_id = ? LIMIT 1', (current_user.class_id,))
                result = cursor.fetchone()
                conn.close()
                
                if result and result['class'] != class_name:
                    return jsonify({
                        'status': 'error',
                        'message': '权限不足，只能导出本班级的评语'
                    }), 403
        
        # 确保exports目录存在且可写
        EXPORTS_FOLDER = 'exports'
        if not os.path.exists(EXPORTS_FOLDER):
            try:
                os.makedirs(EXPORTS_FOLDER, exist_ok=True)
                logger.info(f"创建导出目录: {EXPORTS_FOLDER}")
            except Exception as e:
                logger.error(f"创建导出目录失败: {str(e)}")
                return jsonify({
                    'status': 'error',
                    'message': f'创建导出目录失败: {str(e)}'
                }), 400
        
        # 测试导出目录写入权限
        try:
            test_file = os.path.join(EXPORTS_FOLDER, "test_write.txt")
            with open(test_file, 'w') as f:
                f.write("Test write permission")
            if os.path.exists(test_file):
                os.remove(test_file)
        except Exception as e:
            logger.error(f"导出目录写入权限测试失败: {str(e)}")
            return jsonify({
                'status': 'error',
                'message': f'导出目录没有写入权限: {str(e)}'
            }), 400
            
        # 检查PDF导出功能是否可用
        try:
            from reportlab.lib import colors
            logger.info("ReportLab库成功导入")
        except ImportError as e:
            logger.error(f"ReportLab库导入失败: {str(e)}")
            return jsonify({
                'status': 'error',
                'message': 'PDF导出功能不可用，服务器缺少ReportLab库'
            }), 400
        
        # 限制班级大小 - 如果未指定班级，检查学生总数
        if not class_name:
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute('SELECT COUNT(*) as count FROM students')
                result = cursor.fetchone()
                total_students = result['count'] if result else 0
                conn.close()
                
                # 如果学生总数过多，要求用户选择特定班级
                if total_students > 200:
                    return jsonify({
                        'status': 'error',
                        'message': f'学生总数过多 ({total_students}人)，请选择特定班级后再导出'
                    }), 400
            except Exception as e:
                logger.error(f"检查学生总数时出错: {str(e)}")
                # 继续执行，不中断流程
        
        # 调用PDF导出函数
        logger.info("调用PDF导出函数")
        result = export_comments_to_pdf(class_name)
        logger.info(f"PDF导出结果: {result}")
        
        # 检查result类型，处理可能的返回值为None或元组的情况
        if result is None:
            logger.error("导出函数返回None")
            return jsonify({
                'status': 'error',
                'message': 'PDF导出功能返回None，请检查服务器日志'
            }), 400
        elif isinstance(result, tuple) and len(result) == 2 and result[0] is None:
            logger.error(f"导出函数返回错误元组: {result}")
            return jsonify({
                'status': 'error',
                'message': result[1]
            }), 400
        
        # 正常处理字典类型的结果
        if result.get('status') == 'ok':
            logger.info(f"PDF导出成功: {result.get('file_path')}")
            # 确保下载URL可用
            if 'download_url' in result:
                download_url = result['download_url']
                # 如果URL不是以http开头，添加前缀
                if not download_url.startswith('http'):
                    # 获取服务器的URL前缀
                    if request.host_url:
                        base_url = request.host_url.rstrip('/')
                        download_url = download_url.lstrip('/')
                        result['download_url'] = f"{base_url}/{download_url}"
            return jsonify(result)
        else:
            logger.error(f"PDF导出失败: {result.get('message')}")
            return jsonify(result), 400
    except Exception as e:
        logger.error(f"导出评语PDF时发生错误: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'status': 'error',
            'message': f'导出评语PDF时发生错误: {str(e)}'
        }), 400

# 生成打印预览HTML
@comments_bp.route('/api/preview-comments', methods=['GET'])
def api_preview_comments():
    # 获取班级参数（可选）
    class_name = request.args.get('class')
    
    try:
        # 检查用户权限
        if not hasattr(current_user, 'is_authenticated') or not current_user.is_authenticated:
            return jsonify({'status': 'error', 'message': '未登录', 'code': 'login_required'}), 401
            
        # 如果是班主任，检查是否有权限预览指定班级
        if hasattr(current_user, 'is_admin') and not current_user.is_admin and hasattr(current_user, 'class_id') and current_user.class_id:
            # 如果未指定班级，则使用班主任自己的班级
            if not class_name:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute('SELECT class FROM students WHERE class_id = ? LIMIT 1', (current_user.class_id,))
                result = cursor.fetchone()
                conn.close()
                
                if result:
                    class_name = result['class']
                    logger.info(f"未指定班级，使用班主任班级: {class_name}")
            # 指定了班级，检查是否有权限
            elif class_name:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute('SELECT class FROM students WHERE class_id = ? LIMIT 1', (current_user.class_id,))
                result = cursor.fetchone()
                conn.close()
                
                if result and result['class'] != class_name:
                    return jsonify({
                        'status': 'error',
                        'message': '权限不足，只能预览本班级的评语'
                    }), 403
    
        # 调用预览函数，传递当前用户信息
        result = generate_preview_html(class_name, current_user)
        
        if result['status'] == 'ok':
            # 返回HTML内容
            response = make_response(result['html'])
            response.headers['Content-Type'] = 'text/html'
            return response
        else:
            return jsonify(result), 500
            
    except Exception as e:
        logger.error(f"生成打印预览时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'status': 'error',
            'message': f'生成打印预览时出错: {str(e)}'
        }), 500

# 添加取消导出功能
import threading

# 用于存储活动导出请求的字典和锁
active_export_requests = {}
export_requests_lock = threading.Lock()

# 检查导出请求是否被取消
def is_export_cancelled(request_id):
    """检查指定的导出请求是否已被取消"""
    if not request_id:
        return False
        
    with export_requests_lock:
        request_info = active_export_requests.get(request_id)
        if request_info and request_info.get('cancelled', False):
            logger.info(f"导出请求 {request_id} 已被取消")
            return True
    return False

# 清理过期的导出请求
def cleanup_export_requests():
    """清理超过30分钟的导出请求"""
    now = datetime.datetime.now()
    with export_requests_lock:
        expired_requests = []
        for req_id, info in active_export_requests.items():
            if (now - info['started_at']).total_seconds() > 1800:  # 30分钟
                expired_requests.append(req_id)
                
        for req_id in expired_requests:
            logger.info(f"清理过期的导出请求: {req_id}")
            del active_export_requests[req_id]

# 取消导出API
@comments_bp.route('/api/cancel-export', methods=['POST'])
def cancel_export():
    try:
        data = request.get_json()
        request_id = data.get('requestId')
        
        if not request_id:
            return jsonify({'status': 'error', 'message': '未提供请求ID'})
            
        logger.info(f"收到取消导出请求: {request_id}")
        
        # 清理过期请求
        cleanup_export_requests()
        
        # 标记请求为已取消
        found = False
        with export_requests_lock:
            if request_id in active_export_requests:
                active_export_requests[request_id]['cancelled'] = True
                active_export_requests[request_id]['status'] = 'cancelled'
                active_export_requests[request_id]['cancelled_at'] = datetime.datetime.now()
                found = True
                logger.info(f"已标记导出请求 {request_id} 为已取消")
                
                # 立即更新进度为取消状态
                websocket_progress("导出已取消", 0, request_id)
            
        if found:
            return jsonify({
                'status': 'success',
                'message': '导出操作已取消'
            })
        else:
            logger.warning(f"未找到要取消的导出请求: {request_id}")
            return jsonify({
                'status': 'warning',
                'message': '未找到指定的导出请求，可能已完成或不存在'
            })
    except Exception as e:
        logger.error(f"取消导出请求处理出错: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'status': 'error', 'message': f'取消导出失败: {str(e)}'})

# 在文件转换过程中定期检查取消状态
def convert_docx_to_pdf(input_file, output_file, request_id=None):
    """转换单个Word文档为PDF，确保COM环境正确初始化，并支持取消操作"""
    
    # 检测操作系统
    system = platform.system()
    logger.info(f"当前操作系统: {system}")
    
    # Linux系统使用LibreOffice或unoconv转换
    if system == "Linux":
        try:
            logger.info(f"Linux系统，尝试转换文件：{input_file} -> {output_file}")
            
            # 确保输入文件路径是绝对路径
            input_file_abs = os.path.abspath(input_file)
            output_file_abs = os.path.abspath(output_file)
            output_dir = os.path.dirname(output_file_abs)
            
            # 确保输出目录存在
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
            
            # 先尝试使用unoconv转换
            import subprocess
            check_cmd = ["which", "unoconv"]
            result = subprocess.run(check_cmd, capture_output=True, text=True)
            
            if result.returncode == 0:
                logger.info("检测到unoconv，使用unoconv进行转换")
                # 使用unoconv转换
                cmd = [
                    "unoconv", 
                    "-f", "pdf",
                    "-o", output_file_abs,
                    input_file_abs
                ]
                
                # 执行命令
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
                
                # 检查命令输出
                if result.returncode != 0:
                    logger.error(f"unoconv转换失败，退出码: {result.returncode}")
                    logger.error(f"错误输出: {result.stderr}")
                    logger.warning("unoconv失败，尝试使用LibreOffice")
                else:
                    # 验证文件是否生成
                    if os.path.exists(output_file_abs) and os.path.getsize(output_file_abs) > 0:
                        logger.info(f"unoconv成功转换文件: {output_file_abs}")
                        return True
                    else:
                        logger.error(f"未找到unoconv转换后的PDF文件: {output_file_abs}")
            else:
                logger.info("未检测到unoconv，尝试使用LibreOffice")
                
            # 如果unoconv不可用或失败，尝试使用LibreOffice
            check_cmd = ["which", "libreoffice"]
            result = subprocess.run(check_cmd, capture_output=True, text=True)
            
            if result.returncode != 0:
                logger.warning("在Linux上未找到LibreOffice，尝试使用soffice命令")
                check_cmd = ["which", "soffice"]
                result = subprocess.run(check_cmd, capture_output=True, text=True)
                if result.returncode != 0:
                    logger.error("未找到LibreOffice或soffice命令，无法转换PDF")
                    return False
                libreoffice_cmd = "soffice"
            else:
                libreoffice_cmd = "libreoffice"
            
            # 使用LibreOffice转换
            logger.info(f"使用 {libreoffice_cmd} 转换文档")
            cmd = [
                libreoffice_cmd, 
                "--headless", 
                "--convert-to", 
                "pdf", 
                "--outdir", 
                output_dir,
                input_file_abs
            ]
            
            # 执行命令
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
            
            # 检查命令输出
            if result.returncode != 0:
                logger.error(f"LibreOffice转换失败，退出码: {result.returncode}")
                logger.error(f"错误输出: {result.stderr}")
                return False
                
            # 检查目标文件是否存在
            converted_file = os.path.join(output_dir, os.path.basename(input_file_abs).replace('.docx', '.pdf'))
            
            # 如果输出文件名与默认名不同，需要重命名
            if converted_file != output_file_abs and os.path.exists(converted_file):
                os.rename(converted_file, output_file_abs)
                logger.info(f"重命名转换文件: {converted_file} -> {output_file_abs}")
            
            # 验证文件是否生成
            if os.path.exists(output_file_abs) and os.path.getsize(output_file_abs) > 0:
                logger.info(f"LibreOffice成功转换文件: {output_file_abs}")
                return True
            else:
                logger.error(f"未找到转换后的PDF文件: {output_file_abs}")
                return False
                
        except Exception as e:
            logger.error(f"Linux下文档转换失败: {str(e)}")
            logger.error(traceback.format_exc())
            return False
    
    # Windows系统使用原有代码
    elif system == "Windows":
        # 如果pythoncom不可用，直接返回错误
        if not PYTHONCOM_AVAILABLE:
            logger.error("无法转换Word文档：pythoncom模块不可用，请安装pywin32包")
            return False
            
        try:
            # 检查是否已取消
            if request_id and is_export_cancelled(request_id):
                logger.info(f"文件 {os.path.basename(input_file)} 的转换已被取消")
                return False
                
            # 初始化COM环境
            logger.info(f"为文件 {os.path.basename(input_file)} 初始化COM环境")
            pythoncom.CoInitialize()
            
            # 直接使用win32com进行更可靠的转换，而不是依赖docx2pdf
            try:
                from win32com.client import Dispatch, constants
                
                # 确保输入文件路径是绝对路径
                input_file_abs = os.path.abspath(input_file)
                output_file_abs = os.path.abspath(output_file)
                
                logger.info(f"使用win32com直接转换: {input_file_abs} -> {output_file_abs}")
                
                # 创建Word应用实例
                word = Dispatch('Word.Application')
                word.Visible = False  # 不显示Word界面
                
                # 检查Word是否成功启动
                if word is None:
                    logger.error("无法创建Word应用程序实例")
                    return False
                    
                logger.info("Word应用程序实例创建成功")
                
                # 打开文档
                try:
                    doc = word.Documents.Open(input_file_abs)
                    logger.info(f"成功打开文档: {input_file_abs}")
                except Exception as e:
                    logger.error(f"打开文档失败: {str(e)}")
                    word.Quit()
                    return False
                
                # 另存为PDF
                try:
                    # 检查目录是否存在
                    output_dir = os.path.dirname(output_file_abs)
                    if not os.path.exists(output_dir):
                        os.makedirs(output_dir, exist_ok=True)
                    
                    # 使用不同的常量格式另存为PDF
                    try:
                        # PDF格式常量(17)
                        doc.SaveAs(output_file_abs, FileFormat=17)
                        logger.info(f"使用FileFormat=17成功保存PDF: {output_file_abs}")
                    except Exception as save_error:
                        logger.warning(f"使用FileFormat=17保存失败: {str(save_error)}，尝试使用wdFormatPDF")
                        try:
                            # 尝试使用常量名
                            doc.SaveAs(output_file_abs, FileFormat=constants.wdFormatPDF)
                            logger.info(f"使用wdFormatPDF成功保存PDF: {output_file_abs}")
                        except Exception as e:
                            logger.error(f"所有保存方法都失败: {str(e)}")
                            raise
                except Exception as e:
                    logger.error(f"保存PDF时出错: {str(e)}")
                    doc.Close(False)  # 不保存更改
                    word.Quit()
                    return False
                finally:
                    # 关闭文档
                    try:
                        doc.Close()
                        logger.info("文档已关闭")
                    except:
                        logger.warning("关闭文档时出错")
                
                # 退出Word
                try:
                    word.Quit()
                    logger.info("Word应用已退出")
                except:
                    logger.warning("退出Word时出错")
                
            except ImportError:
                # 如果win32com不可用，尝试使用docx2pdf
                logger.warning("win32com不可用，回退到docx2pdf")
                from docx2pdf import convert
                logger.info(f"使用docx2pdf转换Word文件: {input_file} -> {output_file}")
                convert(input_file, output_file)
            
            # 再次检查是否在转换过程中被取消
            if request_id and is_export_cancelled(request_id):
                logger.info(f"文件 {os.path.basename(input_file)} 转换完成，但请求已被取消")
                # 如果已取消，尝试删除生成的文件
                if os.path.exists(output_file):
                    try:
                        os.remove(output_file)
                        logger.info(f"已删除已取消请求生成的文件: {output_file}")
                    except Exception as e:
                        logger.warning(f"无法删除已取消请求生成的文件: {output_file}, 错误: {str(e)}")
                return False
            
            # 验证文件是否成功创建
            if os.path.exists(output_file):
                file_size = os.path.getsize(output_file)
                if file_size > 0:
                    logger.info(f"成功转换文件为PDF: {output_file}, 文件大小: {file_size} 字节")
                    return True
                else:
                    logger.error(f"转换后的PDF文件大小为0: {output_file}")
                    return False
            else:
                logger.error(f"转换后的PDF文件不存在: {output_file}")
                return False
        except Exception as e:
            logger.error(f"转换文件 {os.path.basename(input_file)} 失败: {str(e)}")
            logger.error(f"错误详情: {traceback.format_exc()}")
            return False
        finally:
            # 释放COM资源
            try:
                if PYTHONCOM_AVAILABLE:
                    logger.info(f"释放文件 {os.path.basename(input_file)} 的COM资源")
                    pythoncom.CoUninitialize()
            except:
                logger.warning("CoUninitialize失败，可能COM资源未正确初始化")
    
    # 其他操作系统（如macOS）
    else:
        logger.error(f"不支持的操作系统: {system}, 无法转换PDF")
        return False

# 导出报告API
@comments_bp.route('/api/export-reports', methods=['POST'])
def api_export_reports():
    try:
        # 检查用户是否已登录
        if not hasattr(current_user, 'is_authenticated') or not current_user.is_authenticated:
            logger.warning("未登录用户或会话过期，尝试导出报告")
            return jsonify({'status': 'error', 'message': '您需要登录后才能导出报告', 'code': 'login_required'}), 401
            
        # 获取请求数据
        data = request.get_json()
        
        # 记录请求ID，用于支持取消功能
        request_id = request.headers.get('X-Export-Request-ID')
        if request_id:
            logger.info(f"收到导出报告请求 [请求ID: {request_id}]: {data}")
            # 将请求ID存储到活动请求字典中
            with export_requests_lock:
                active_export_requests[request_id] = {
                    'status': 'processing',
                    'started_at': datetime.datetime.now(),
                    'cancelled': False
                }
        else:
            logger.info(f"收到导出报告请求 (无请求ID): {data}")
        
        # 验证请求数据
        student_ids = data.get('studentIds', [])
        template_id = data.get('templateId', '泉州东海湾实验学校综合素质发展报告单')
        use_default_template = data.get('useDefaultTemplate', False)
        settings = data.get('settings', {})
        
        # 获取导出类型（默认为word）
        export_type = settings.get('exportType', 'word')
        logger.info(f"导出类型: {export_type}")
        
        if not student_ids:
            return jsonify({'status': 'error', 'message': '未选择任何学生'})
            
        if not template_id:
            template_id = '泉州东海湾实验学校综合素质发展报告单'
            logger.info(f"未指定模板ID，将使用默认模板: {template_id}")
            
        logger.info(f"导出学生: {student_ids}, 模板ID: {template_id}, 使用内置默认模板: {use_default_template}, 设置: {settings}")
            
        # 使用默认设置补充缺失的设置
        default_settings = {
            'schoolYear': '2023-2024',
            'semester': '1',  # 1表示第一学期，2表示第二学期
            'fileNameFormat': 'id_name'  # id_name, name_id, id, name
        }
        
        for key, value in default_settings.items():
            if key not in settings or not settings[key]:
                settings[key] = value
                
        # 初始化报告导出器
        try:
            from utils.report_exporter import ReportExporter
            exporter = ReportExporter()
            logger.info("初始化ReportExporter成功")
        except Exception as e:
            logger.error(f"初始化报告导出器失败: {str(e)}")
            return jsonify({'status': 'error', 'message': f'导出报告失败: {str(e)}'})
        
        # 检查模板是否存在
        template_path = exporter.get_template_path(template_id)
        logger.info(f"模板路径: {template_path}, 模板是否存在: {os.path.exists(template_path)}")
        if not os.path.exists(template_path):
            return jsonify({'status': 'error', 'message': f'模板不存在: {template_id}'})
                
        # 查询学生数据
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # 准备SQL查询，对班主任进行班级过滤
        placeholders = ','.join(['?' for _ in student_ids])
        
        # 构建SQL查询，根据用户身份添加班级过滤条件
        if current_user.is_admin:
            # 管理员可以查询所有班级学生
            student_sql = f'''
                SELECT * FROM students 
                WHERE id IN ({placeholders})
            '''
            cursor.execute(student_sql, student_ids)
        else:
            # 班主任只能查询自己班级的学生
            student_sql = f'''
                SELECT * FROM students 
                WHERE id IN ({placeholders}) AND class_id = ?
            '''
            params = student_ids + [current_user.class_id]
            cursor.execute(student_sql, params)
        
        students = [dict(row) for row in cursor.fetchall()]
        
        # 记录实际查询到的学生数量，并检查是否与请求的学生数量一致
        logger.info(f"找到 {len(students)} 名学生的数据 (请求了 {len(student_ids)} 名学生)")
        if len(students) < len(student_ids) and not current_user.is_admin:
            logger.warning(f"有 {len(student_ids) - len(students)} 名学生不属于当前班主任的班级")
        
        if not students:
            return jsonify({'status': 'error', 'message': '未找到任何学生数据，或没有符合查询条件的学生'})
            
        # 更新student_ids，只包含当前查询到的学生
        student_ids = [s['id'] for s in students]
        
        # 查询学生评语
        comments_dict = {}
        for student_id in student_ids:
            # 对非管理员用户，确保只查询本班级学生
            if not current_user.is_admin:
                cursor.execute('SELECT id, name, comments FROM students WHERE id = ? AND class_id = ?', 
                             (student_id, current_user.class_id))
            else:
                cursor.execute('SELECT id, name, comments FROM students WHERE id = ?', (student_id,))
                
            student = cursor.fetchone()
            if student and student['comments']:
                comments_dict[student_id] = {
                    'studentId': student_id,
                    'studentName': student['name'],
                    'content': student['comments']
                }
            else:
                comments_dict[student_id] = {
                    'studentId': student_id,
                    'studentName': student['name'] if student else 'Unknown',
                    'content': ''  # 空评语
                }
        
        # 查询学生成绩
        grades_dict = {}
        for student_id in student_ids:
            try:
                # 对非管理员用户，确保只查询本班级学生
                if not current_user.is_admin:
                    cursor.execute('''
                        SELECT daof, yuwen, shuxue, yingyu, laodong, tiyu, yinyue, 
                               meishu, kexue, zonghe, xinxi, shufa
                        FROM students WHERE id = ? AND class_id = ?
                    ''', (student_id, current_user.class_id))
                else:
                    cursor.execute('''
                        SELECT daof, yuwen, shuxue, yingyu, laodong, tiyu, yinyue, 
                               meishu, kexue, zonghe, xinxi, shufa
                        FROM students WHERE id = ?
                    ''', (student_id,))
                
                grade = cursor.fetchone()
                
                if grade:
                    grade_dict = dict(grade)
                    grades_dict[student_id] = {
                        'grades': grade_dict
                    }
                else:
                    grades_dict[student_id] = {'grades': {}}
            except sqlite3.OperationalError as e:
                # 处理表不存在或列不存在的情况
                logger.warning(f"获取成绩时出错: {str(e)}, 将使用空成绩数据")
                grades_dict[student_id] = {'grades': {}}
        
        # 关闭数据库连接
        conn.close()
        
        # 执行数据完整性验证
        try:
            from utils.data_validator import DataValidator
            logger.info("导出前进行数据完整性验证...")
            
            is_valid, error_message, problem_students = DataValidator.validate_export_data(
                students=students,
                comments=comments_dict,
                grades=grades_dict
            )
            
            if not is_valid:
                logger.error(f"数据验证失败: {error_message}")
                
                # 处理问题学生列表，对于班主任只显示本班级的学生
                filtered_problem_students = []
                if problem_students:
                    # 如果是班主任且有班级ID，则过滤问题学生列表
                    if hasattr(current_user, 'is_admin') and not current_user.is_admin and hasattr(current_user, 'class_id') and current_user.class_id:
                        for student in problem_students:
                            # 检查学生是否属于班主任的班级
                            for s in students:
                                if s['id'] == student['id'] and s.get('class_id') == current_user.class_id:
                                    filtered_problem_students.append(student)
                                    break
                        logger.info(f"根据班主任班级筛选后的问题学生数量: {len(filtered_problem_students)}/{len(problem_students)}")
                    else:
                        # 如果是管理员，保留所有问题学生
                        filtered_problem_students = problem_students
                
                # 构建更详细的错误信息
                details = ""
                if filtered_problem_students:
                    details = "\n\n问题学生列表:\n"
                    for i, student in enumerate(filtered_problem_students[:10], 1):  # 限制显示前10个
                        problems_text = ", ".join(student['problems'])
                        details += f"{i}. {student['name']} (学号: {student['id']}): {problems_text}\n"
                    
                    if len(filtered_problem_students) > 10:
                        details += f"... 以及其他 {len(filtered_problem_students) - 10} 名学生"
                
                # 更新错误消息，显示筛选后的学生数量
                if hasattr(current_user, 'is_admin') and not current_user.is_admin and hasattr(current_user, 'class_id') and current_user.class_id:
                    if filtered_problem_students:
                        error_message = f"发现 {len(filtered_problem_students)} 名本班学生的数据不完整，无法导出报告。"
                    else:
                        # 如果本班学生数据已完整但有其他验证问题(如全优或全满分)，仍阻止导出
                        if "均为'优'" in error_message or "均为满分" in error_message:
                            logger.info(f"验证失败: {error_message}")
                        # 如果确实只是其他班级的问题，允许导出
                        else:
                            logger.info("本班学生数据已完整，忽略其他班级的数据问题，继续导出报告")
                            # 继续执行导出流程
                            pass
                
                # 只有在本班没有问题学生且不是全优或全满分问题时才继续执行
                if (hasattr(current_user, 'is_admin') and not current_user.is_admin and hasattr(current_user, 'class_id') and current_user.class_id and 
                    not filtered_problem_students and 
                    "均为'优'" not in error_message and 
                    "均为满分" not in error_message):
                    logger.info("跳过验证错误，允许导出本班学生数据")
                else:
                    return jsonify({
                        'status': 'error', 
                        'message': error_message,
                        'details': details,
                        'validation_failed': True,
                        'problem_students': [s['id'] for s in filtered_problem_students]
                    })
            
            logger.info("数据验证通过，继续导出报告...")
        except Exception as e:
            logger.error(f"执行数据验证时出错: {str(e)}")
            logger.error(traceback.format_exc())
            # 出错时不阻止导出，继续执行
            
        # 生成报告
        try:
            from utils.report_exporter import ReportExporter
            
            # 检查必要依赖
            try:
                import docxtpl
                import docx
                logger.info("依赖检查: docxtpl和docx库已安装")
                
                # 如果是PDF导出，检查PDF相关依赖
                if export_type == 'pdf':
                    try:
                        from docx2pdf import convert
                        logger.info("依赖检查: docx2pdf库已安装")
                    except ImportError as e:
                        logger.error(f"缺少PDF导出依赖: {str(e)}")
                        return jsonify({
                            'status': 'error', 
                            'message': f'PDF报告导出失败: 缺少必要依赖，请安装 docx2pdf 库'
                        })
            except ImportError as e:
                logger.error(f"缺少必要依赖: {str(e)}")
                return jsonify({
                    'status': 'error', 
                    'message': f'报告导出失败: 缺少必要依赖，请安装 python-docx 和 docxtpl 库'
                })
            
            logger.info(f"开始生成报告，学生数量: {len(students)}")
            logger.info(f"学生数据示例: {students[0] if students else '无'}")
            logger.info(f"评语数据: {comments_dict}")
            logger.info(f"成绩数据: {grades_dict}")
            
            exporter = ReportExporter()
            success, result = exporter.export_reports(
                students=students,
                comments=comments_dict,
                grades=grades_dict,
                template_id=template_id,
                settings=settings
            )
            
            if not success:
                logger.error(f"导出报告失败: {result}")
                return jsonify({'status': 'error', 'message': f'导出报告失败: {result}'})
            
            # 检查系统是否安装了Microsoft Word
            def check_word_installed():
                """检查系统是否安装了Microsoft Word或LibreOffice"""
                logger.info("开始检查文档转换软件是否安装")
                try:
                    # 根据操作系统类型采用不同的检测方法
                    system = platform.system()
                    logger.info(f"当前操作系统: {system}")
                    
                    if system == "Windows":
                        # Windows系统：尝试通过注册表检查
                        try:
                            import winreg
                            # 尝试打开Word注册表键
                            word_key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Office\Word")
                            winreg.CloseKey(word_key)
                            logger.info("通过注册表确认Microsoft Word已安装")
                            return True
                        except ImportError:
                            logger.warning("无法导入winreg模块，将尝试通过COM接口检测Word")
                        except Exception as e:
                            logger.warning(f"无法通过注册表检测Word: {str(e)}")
                        
                        # 尝试通过COM接口创建Word实例
                        if PYTHONCOM_AVAILABLE:
                            try:
                                from win32com.client import Dispatch
                                pythoncom.CoInitialize()
                                word = Dispatch("Word.Application")
                                word.Quit()
                                pythoncom.CoUninitialize()
                                logger.info("通过COM接口确认Microsoft Word已安装")
                                return True
                            except Exception as e:
                                logger.error(f"通过COM接口检测Word失败: {str(e)}")
                                return False
                        else:
                            logger.warning("pythoncom模块不可用，无法通过COM接口检测Word")
                            return False
                    elif system == "Darwin":  # macOS
                        # 检查macOS上的Word应用位置
                        common_paths = [
                            "/Applications/Microsoft Word.app",
                            "/Applications/Microsoft Office/Microsoft Word.app"
                        ]
                        for path in common_paths:
                            if os.path.exists(path):
                                logger.info(f"在路径 {path} 找到Microsoft Word")
                                return True
                        logger.warning("在macOS上未找到Microsoft Word")
                        return False
                        
                    elif system == "Linux":
                        # Linux上检查LibreOffice或unoconv
                        try:
                            import subprocess
                            logger.info("检查Linux系统上的文档转换工具")
                            
                            # 先尝试检查unoconv
                            which_result = subprocess.run(["which", "unoconv"], capture_output=True, text=True)
                            if which_result.returncode == 0:
                                unoconv_path = which_result.stdout.strip()
                                logger.info(f"在Linux上找到unoconv: {unoconv_path}")
                                return True
                            
                            # 再尝试使用which命令查找libreoffice
                            which_result = subprocess.run(["which", "libreoffice"], capture_output=True, text=True)
                            if which_result.returncode == 0:
                                libreoffice_path = which_result.stdout.strip()
                                logger.info(f"在Linux上找到LibreOffice: {libreoffice_path}")
                                return True
                            
                            # 如果找不到libreoffice，尝试查找soffice
                            which_result = subprocess.run(["which", "soffice"], capture_output=True, text=True)
                            if which_result.returncode == 0:
                                soffice_path = which_result.stdout.strip()
                                logger.info(f"在Linux上找到LibreOffice (soffice): {soffice_path}")
                                return True
                            
                            # 如果以上命令都找不到，尝试检查常见安装位置
                            common_paths = [
                                "/usr/bin/libreoffice",
                                "/usr/bin/soffice",
                                "/usr/bin/unoconv",
                                "/usr/lib/libreoffice/program/soffice"
                            ]
                            for path in common_paths:
                                if os.path.exists(path):
                                    logger.info(f"在Linux上找到文档转换工具: {path}")
                                    return True
                            
                            logger.warning("在Linux上未找到LibreOffice或unoconv")
                            return False
                        except Exception as e:
                            logger.error(f"检查文档转换工具时出错: {str(e)}")
                            return False
                    else:
                        logger.warning(f"未知操作系统类型: {system}，无法确定文档转换软件状态")
                        return False
                except Exception as e:
                    logger.error(f"检查文档转换软件安装状态时出错: {str(e)}")
                    logger.error(traceback.format_exc())
                    return False

            # 如果是PDF导出，需要将Word文档转换为PDF
            if export_type == 'pdf':
                logging.info("开始PDF导出流程")
                try:
                    # 检查是否安装了Word
                    word_installed = check_word_installed()
                    if not word_installed:
                        logger.warning("系统中未检测到Microsoft Word，PDF转换可能失败")
                        # 继续尝试，因为可能有其他方式转换
                    else:
                        logger.info("检测到Microsoft Word，将继续PDF转换")
                    
                    # 保存原始Word文档结果作为回退
                    original_word_result = result
                    
                    # 创建临时目录解压ZIP文件
                    with tempfile.TemporaryDirectory() as temp_dir:
                        # 检查请求是否已取消
                        if request_id and is_export_cancelled(request_id):
                            logger.info(f"PDF转换前检测到请求已取消: {request_id}")
                            return jsonify({
                                'status': 'warning',
                                'message': '导出操作已被用户取消'
                            })
                            
                        # 保存ZIP文件
                        zip_path = os.path.join(temp_dir, "word_reports.zip")
                        logger.info(f"保存ZIP文件到临时目录: {zip_path}")
                        with open(zip_path, 'wb') as f:
                            f.write(result)
                        
                        # 解压ZIP文件
                        extract_dir = os.path.join(temp_dir, "extracted")
                        os.makedirs(extract_dir, exist_ok=True)
                        logger.info(f"解压ZIP文件到: {extract_dir}")
                        
                        # 解压ZIP文件
                        with zipfile.ZipFile(zip_path, 'r') as zipf:
                            file_list = zipf.namelist()
                            logger.info(f"ZIP文件中包含的文件: {file_list}")
                            zipf.extractall(extract_dir)
                        
                        # 检查是否有docx文件
                        extracted_files = os.listdir(extract_dir)
                        logger.info(f"解压后的文件列表: {extracted_files}")
                        docx_files = [f for f in extracted_files if f.endswith('.docx')]
                        
                        if not docx_files:
                            logger.error("未找到任何.docx文件需要转换")
                            raise Exception("解压后未找到任何Word文档文件")
                            
                        # 再次检查请求是否已取消
                        if request_id and is_export_cancelled(request_id):
                            logger.info(f"解压后检测到请求已取消: {request_id}")
                            return jsonify({
                                'status': 'warning',
                                'message': '导出操作已被用户取消'
                            })
                            
                        # 创建PDF输出目录
                        pdf_dir = os.path.join(temp_dir, "pdf")
                        os.makedirs(pdf_dir, exist_ok=True)
                        logger.info(f"创建PDF输出目录: {pdf_dir}")
                        
                        # 建立文件名与学生信息的映射，以便显示正确的学生名称
                        student_name_map = {}
                        for student in students:
                            student_id = student['id']
                            student_name = student['name']
                            # 根据fileNameFormat设置确定文件名格式
                            file_name_format = settings.get('fileNameFormat', 'id_name')
                            if file_name_format == 'id_name':
                                file_prefix = f"{student_id}_{student_name}"
                            elif file_name_format == 'name_id':
                                file_prefix = f"{student_name}_{student_id}"
                            else:
                                file_prefix = student_id
                            
                            # 将文件名前缀与学生名称关联
                            for docx_file in docx_files:
                                if docx_file.startswith(file_prefix) or file_prefix in docx_file:
                                    student_name_map[docx_file] = student_name
                                    break
                        
                        logger.info(f"创建学生文件名映射: {student_name_map}")
                        
                        # 设置转换计数器
                        total_files = len(docx_files)
                        successful_conversions = 0
                        
                        # 转换DOCX文件为PDF
                        pdf_files = []
                        for i, docx_file in enumerate(docx_files):
                            # 检查请求是否已取消
                            if request_id and is_export_cancelled(request_id):
                                logger.info(f"转换过程中检测到请求已取消: {request_id}")
                                return jsonify({
                                    'status': 'warning',
                                    'message': '导出操作已被用户取消'
                                })
                                
                            # 准备转换路径
                            docx_path = os.path.join(extract_dir, docx_file)
                            pdf_file = docx_file.replace('.docx', '.pdf')
                            pdf_path = os.path.join(pdf_dir, pdf_file)
                            
                            # 获取学生信息，尝试从映射中获取学生名称
                            student_name = student_name_map.get(docx_file, "")
                            if not student_name:
                                # 如果映射中没有，则从文件名提取
                                student_name = docx_file.replace('.docx', '').split('_')[-1] if '_' in docx_file else docx_file.replace('.docx', '')
                            
                            # 更新进度信息，显示当前学生名称和进度
                            progress_message = update_progress_in_pdf_conversion(docx_file, student_name, i+1, total_files, request_id)
                            logger.info(progress_message)
                            
                            # 执行转换，添加重试逻辑
                            logger.info(f"开始转换 [{i+1}/{total_files}]: {docx_path} -> {pdf_path}")
                            
                            # 尝试最多3次转换
                            success = False
                            for attempt in range(3):
                                try:
                                    if attempt > 0:
                                        logger.info(f"重试第{attempt+1}次转换: {docx_file}")
                                    
                                    # 执行转换
                                    success = convert_docx_to_pdf(docx_path, pdf_path, request_id)
                                    
                                    if success:
                                        logger.info(f"转换成功 (尝试 {attempt+1}): {docx_file}")
                                        pdf_files.append(pdf_file)
                                        successful_conversions += 1
                                        break  # 成功则退出重试循环
                                except Exception as e:
                                    logger.error(f"尝试转换第{attempt+1}次失败: {str(e)}")
                                    success = False
                                
                                # 如果所有尝试都失败，记录错误
                                if not success and attempt == 2:  # 最后一次尝试
                                    logger.error(f"转换失败 {docx_file} 在3次尝试后")
                        
                        # 记录转换完成情况
                        logger.info(f"完成转换: 成功 {successful_conversions}/{total_files} 文件")
                        
                        # 如果没有成功转换任何文件，返回错误
                        if successful_conversions == 0:
                            logger.error("没有成功转换任何文件为PDF")
                            # 回退到原始Word文档结果
                            return send_file(
                                original_word_result,
                                mimetype='application/zip',
                                as_attachment=True,
                                download_name='student_reports_word.zip'
                            )
                        
                        # 将PDF文件合并为一个文件
                        try:
                            # 尝试导入PyPDF2
                            try:
                                import PyPDF2
                                logger.info("成功导入PyPDF2")
                            except ImportError:
                                logger.error("无法导入PyPDF2库，尝试安装")
                                try:
                                    import subprocess
                                    subprocess.check_call([sys.executable, "-m", "pip", "install", "PyPDF2"])
                                    import PyPDF2
                                    logger.info("成功安装并导入PyPDF2")
                                except Exception as e:
                                    logger.error(f"安装PyPDF2失败: {str(e)}")
                                    # 如果无法安装则跳过合并，只提供单独的PDF文件
                                    raise
                            
                            # 检查是否有PDF文件需要合并
                            if len(pdf_files) > 1:
                                # 更新进度信息，显示正在合并
                                merge_message = f"正在合并 {len(pdf_files)} 个PDF文件..."
                                websocket_progress(merge_message, 90, request_id)
                                logger.info(merge_message)
                                
                                # 根据班主任的class_id获取完整的班级名称
                                class_name = ""
                                try:
                                    # 获取当前登录用户的class_id
                                    class_id = current_user.class_id if hasattr(current_user, 'class_id') else None
                                    
                                    if class_id:
                                        # 连接数据库
                                        conn = get_db_connection()
                                        cursor = conn.cursor()
                                        
                                        # 查询classes表获取班级名称
                                        cursor.execute('SELECT name FROM classes WHERE id = ?', (class_id,))
                                        class_record = cursor.fetchone()
                                        conn.close()
                                        
                                        if class_record and 'name' in class_record:
                                            class_name = class_record['name']
                                            logger.info(f"从数据库获取班级名称: '{class_name}'")
                                        else:
                                            logger.warning(f"未找到class_id为{class_id}的班级记录")
                                            
                                            # 尝试从学生信息中获取班级名称作为备选
                                            if students and len(students) > 0:
                                                class_name = students[0].get('class', '')
                                                logger.info(f"从学生记录获取班级名称: '{class_name}'")
                                    else:
                                        logger.warning("当前用户没有关联的班级ID")
                                        
                                        # 尝试从学生信息中获取班级名称作为备选
                                        if students and len(students) > 0:
                                            class_name = students[0].get('class', '')
                                            logger.info(f"从学生记录获取班级名称: '{class_name}'")
                                except Exception as e:
                                    logger.error(f"获取班级名称时出错: {str(e)}")
                                    # 如果出错，尝试从学生信息获取班级
                                    if students and len(students) > 0:
                                        class_name = students[0].get('class', '')
                                        logger.info(f"从学生记录获取班级名称: '{class_name}'")
                                
                                # 去除班级名称中可能存在的非法文件名字符
                                if class_name:
                                    import re
                                    class_name = re.sub(r'[\\/*?:"<>|]', '', class_name)  # 移除Windows不允许的文件名字符
                                
                                # 生成合并PDF的文件名
                                merged_pdf_filename = f"{class_name}合并.pdf" if class_name else "合并报告.pdf"
                                logger.info(f"使用班级名称 '{class_name}' 生成合并PDF文件名: {merged_pdf_filename}")
                                
                                merged_pdf_path = os.path.join(pdf_dir, merged_pdf_filename)
                                
                                # 创建PDF合并器
                                merger = PyPDF2.PdfMerger()
                                
                                # 添加所有PDF文件，确保按学号顺序合并
                                # 创建一个映射，关联文件名与对应的学生信息
                                student_pdf_mapping = []
                                for pdf_file in pdf_files:
                                    # 跳过合并后的PDF文件
                                    if pdf_file == merged_pdf_filename:
                                        continue
                                        
                                    # 从文件名中提取学生ID
                                    # 假设文件名格式为: ID_姓名.pdf 或 ID.pdf
                                    student_id = pdf_file.split('_')[0].split('.')[0]
                                    try:
                                        # 尝试将学号转为整数进行比较
                                        numeric_id = int(student_id)
                                        student_pdf_mapping.append((numeric_id, pdf_file))
                                    except ValueError:
                                        # 如果学号不是数字，则保持原样
                                        logger.warning(f"无法将学号解析为数字: {student_id}，文件名: {pdf_file}")
                                        student_pdf_mapping.append((student_id, pdf_file))
                                
                                # 按学号排序
                                student_pdf_mapping.sort(key=lambda x: x[0])
                                logger.info(f"按学号顺序排序后的PDF文件列表: {[mapping[1] for mapping in student_pdf_mapping]}")
                                
                                # 按学号顺序合并PDF文件
                                for idx, (_, pdf_file) in enumerate(student_pdf_mapping):
                                    pdf_path = os.path.join(pdf_dir, pdf_file)
                                    if os.path.exists(pdf_path):
                                        try:
                                            merger.append(pdf_path)
                                            # 更新合并进度
                                            merge_progress = f"正在合并PDF文件 ({idx+1}/{len(student_pdf_mapping)})"
                                            websocket_progress(merge_progress, 90 + int((idx+1)/len(student_pdf_mapping)*5), request_id)
                                            logger.info(f"成功添加PDF文件到合并器: {pdf_file}")
                                        except Exception as e:
                                            logger.error(f"添加PDF文件到合并器失败: {pdf_file}, 错误: {str(e)}")
                                
                                # 写入合并后的PDF文件
                                try:
                                    # 更新进度为正在写入合并文件
                                    websocket_progress("正在生成合并后的PDF文件...", 95, request_id)
                                    merger.write(merged_pdf_path)
                                    merger.close()
                                    # 更新进度为合并完成
                                    websocket_progress("PDF文件合并完成", 98, request_id)
                                    logger.info(f"成功创建合并的PDF文件: {merged_pdf_path}")
                                    
                                    # 将合并的PDF文件添加到pdf_files列表中，确保它会被包含在zip包中
                                    pdf_files.append(merged_pdf_filename)
                                except Exception as e:
                                    logger.error(f"写入合并的PDF文件失败: {str(e)}")
                            else:
                                logger.info("只有一个或没有PDF文件，跳过合并步骤")
                        except Exception as e:
                            logger.error(f"合并PDF文件时出错: {str(e)}")
                            logger.error(traceback.format_exc())
                            # 如果合并失败，继续处理，仍然提供单独的PDF文件
                        
                        # 创建PDF文件的压缩包
                        pdf_zip_path = os.path.join(temp_dir, "pdf_reports.zip")
                        logger.info(f"创建PDF ZIP文件: {pdf_zip_path}")
                        
                        with zipfile.ZipFile(pdf_zip_path, 'w') as zipf:
                            for pdf_file in pdf_files:
                                file_path = os.path.join(pdf_dir, pdf_file)
                                logger.info(f"添加PDF到ZIP: {file_path} -> {pdf_file}")
                                zipf.write(file_path, pdf_file)
                        
                        # 读取最终的ZIP文件
                        with open(pdf_zip_path, 'rb') as f:
                            result = f.read()
                            
                        logger.info(f"PDF转换成功，ZIP文件大小: {len(result)} 字节")
                        export_type = 'pdf'
                        
                except Exception as pdf_error:
                    # 检查是否是用户取消导致的错误
                    if isinstance(pdf_error, Exception) and "用户取消" in str(pdf_error):
                        logger.info("用户取消了PDF转换")
                        return jsonify({
                            'status': 'warning',
                            'message': '导出操作已被用户取消'
                        })
                        
                    # 记录转换失败详情
                    logger.error(f"PDF转换失败: {str(pdf_error)}")
                    logger.error(traceback.format_exc())
                    logger.info("将使用原始Word文档作为后备方案")
                    
                    # 回退到Word格式
                    export_type = 'word'
                    result = original_word_result
                    
                    # 如果这是直接下载模式，添加特殊头部
                    if os.environ.get('EXPORT_DIRECT_DOWNLOAD', '1') == '1':
                        response = make_response(result)
                        response.headers['Content-Type'] = 'application/zip'
                        response.headers['Content-Disposition'] = f'attachment; filename=student_reports_word_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
                        response.headers['X-PDF-Conversion-Failed'] = 'true'
                        response.headers['X-PDF-Conversion-Error'] = str(pdf_error)[:200]  # 限制错误消息长度
                        logger.info("返回带有PDF转换失败标记的Word文档")
                        return response
            
            # 直接返回文件
            if os.environ.get('EXPORT_DIRECT_DOWNLOAD', '1') == '1':
                logger.info("使用直接下载模式")
                response = make_response(result)
                response.headers['Content-Type'] = 'application/zip'
                file_ext = 'pdf' if export_type == 'pdf' else 'docx'
                response.headers['Content-Disposition'] = f'attachment; filename=student_reports_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
                return response
            
            # 保存导出文件
            file_ext = 'pdf' if export_type == 'pdf' else 'docx'
            export_filename = f"student_reports_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
            export_path = os.path.join('exports', export_filename)
            
            # 确保exports目录存在
            os.makedirs('exports', exist_ok=True)
            
            # 写入文件
            with open(export_path, 'wb') as f:
                f.write(result)
            
            logger.info(f"报告生成成功，保存为: {export_path}")
            
            return jsonify({
                'status': 'ok',
                'filename': export_filename,
                'message': f'成功导出 {len(students)} 个学生的{export_type}报告'
            })
            
        except Exception as e:
            logger.error(f"导出报告时出错: {str(e)}")
            logger.error(traceback.format_exc())
            return jsonify({'status': 'error', 'message': f'导出报告失败: {str(e)}'})
    except Exception as e:
        logger.error(f"导出报告请求处理出错: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'status': 'error', 'message': f'导出报告失败: {str(e)}'})

# 初始化评语模块
def init_comments(app):
    # 只注册下载评语导出文件的路由，不创建表
    @app.route('/download/exports/<path:filename>', methods=['GET'])
    def download_export(filename):
        EXPORTS_FOLDER = 'exports'
        app.logger.info(f"请求下载文件: {filename}, 从目录: {EXPORTS_FOLDER}")
        
        # 检查文件是否存在
        file_path = os.path.join(EXPORTS_FOLDER, filename)
        if not os.path.exists(file_path):
            app.logger.error(f"请求的文件不存在: {file_path}")
            return jsonify({
                'status': 'error',
                'message': '请求的文件不存在'
            }), 404
            
        # 发送文件
        try:
            return send_from_directory(EXPORTS_FOLDER, filename, as_attachment=True)
        except Exception as e:
            app.logger.error(f"发送文件时出错: {str(e)}")
            return jsonify({
                'status': 'error',
                'message': f'发送文件时出错: {str(e)}'
            }), 500

@comments_bp.route('/api/students', methods=['GET'])
def get_students():
    """
    获取学生列表，如果是班主任则只返回其班级的学生
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取当前用户信息
        if not hasattr(current_user, 'is_authenticated') or not current_user.is_authenticated:
            return jsonify({'status': 'error', 'message': '未登录', 'code': 'login_required'}), 401
            
        # 如果是管理员，返回所有学生
        if hasattr(current_user, 'is_admin') and current_user.is_admin:
            cursor.execute('SELECT * FROM students ORDER BY class, CAST(id AS INTEGER)')
        else:
            # 如果是班主任，只返回其班级的学生
            if hasattr(current_user, 'class_id') and current_user.class_id:
                cursor.execute('SELECT * FROM students WHERE class_id = ? ORDER BY CAST(id AS INTEGER)', (current_user.class_id,))
            else:
                return jsonify({'status': 'error', 'message': '您没有关联的班级'}), 403
            
        students = cursor.fetchall()
        conn.close()
        
        # 转换为字典列表
        students_list = []
        for student in students:
            student_dict = dict(student)
            students_list.append(student_dict)
            
        return jsonify({
            'status': 'ok',
            'students': students_list
        })
    except Exception as e:
        logger.error(f"获取学生列表时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': f'获取学生列表时出错: {str(e)}'}), 500

# 清理过期的导出进度
def cleanup_export_progress():
    """清理超过30分钟的导出进度信息"""
    global user_export_progress
    
    current_time = time.time()
    with progress_lock:
        expired_keys = []
        for key, progress in user_export_progress.items():
            if current_time - progress.get('timestamp', 0) > 1800:  # 30分钟
                expired_keys.append(key)
        
        for key in expired_keys:
            logger.info(f"清理过期的导出进度: {key}")
            del user_export_progress[key]

# 添加导出进度查询接口
@comments_bp.route('/api/export-progress', methods=['GET'])
def get_export_progress():
    """获取当前导出进度信息"""
    global user_export_progress
    
    # 清理过期的进度信息
    cleanup_export_progress()
    
    # 获取当前用户ID
    user_id = None
    if hasattr(current_user, 'id') and current_user.is_authenticated:
        user_id = str(current_user.id)
    
    # 获取请求中的request_id参数
    request_id = request.args.get('request_id')
    
    # 确定要查询的进度键
    progress_key = request_id if request_id else user_id
    if not progress_key:
        progress_key = 'anonymous'
    
    # 线程安全地获取进度信息
    with progress_lock:
        user_progress = user_export_progress.get(progress_key, {})
    
    # 检查是否有有效的进度信息
    current_time = time.time()
    if not user_progress or current_time - user_progress.get('timestamp', 0) > 300:  # 5分钟内的进度才是有效的
        # 超过5分钟没有更新或没有进度信息，返回空进度
        return jsonify({
            'status': 'idle',
            'message': '',
            'percent': 0
        })
    
    # 检查是否有导出完成的标记
    progress_status = 'idle'
    if user_progress.get('message'):
        if user_progress.get('percent') == 100 or '导出完成' in user_progress.get('message', ''):
            progress_status = 'completed'
        else:
            progress_status = 'processing'
    
    # 返回当前用户的进度
    return jsonify({
        'status': progress_status,
        'message': user_progress.get('message', ''),
        'percent': user_progress.get('percent', 0),
        'request_id': user_progress.get('request_id'),
        'user_id': user_progress.get('user_id')
    })

# 更新导出报告的转换和合并进度显示
def update_progress_in_pdf_conversion(docx_file, student_name, current_index, total_files, request_id=None):
    """更新PDF转换进度"""
    # 构建更清晰的进度信息
    progress_message = f"正在导出学生报告: {student_name} ({current_index}/{total_files})"
    # 计算百分比 (但前端不会显示)
    percent = int((current_index / total_files) * 80) if total_files > 0 else 0
    # 调用websocket_progress更新进度
    websocket_progress(progress_message, percent, request_id)
    return progress_message

# 評語导入预览API
@comments_bp.route('/api/preview-import-comments', methods=['POST'])
def preview_import_comments():
    """预览Excel文件中的评语数据，不执行真正的导入"""
    
    try:
        # 检查用户是否已登录
        if not hasattr(current_user, 'is_authenticated') or not current_user.is_authenticated:
            return jsonify({'status': 'error', 'message': '您需要登录后才能导入评语', 'code': 'login_required'}), 401
            
        # 检查是否有文件上传
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': '没有选择文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': '没有选择文件'}), 400
        
        # 检查文件类型
        if not file.filename.lower().endswith('.xlsx'):
            return jsonify({'status': 'error', 'message': '请上传Excel文件（.xlsx格式），不支持旧版.xls格式'}), 400
        
        # 确保uploads目录存在
        os.makedirs('uploads', exist_ok=True)
        
        # 【彻底修复】保存文件逻辑，确保保留.xlsx扩展名
        filename = secure_filename(file.filename)
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        # 确保文件名包含扩展名
        base_name, ext = os.path.splitext(filename)
        if not ext or ext.lower() != '.xlsx':
            ext = '.xlsx'  # 强制添加正确的扩展名
        saved_filename = f"{timestamp}_{base_name}{ext}"
        upload_path = os.path.join('uploads', saved_filename)
        
        logger.info(f"将保存Excel文件为: {upload_path}，原文件名: {filename}")
        
        # 保存文件
        try:
            file.save(upload_path)
            logger.info(f"成功保存上传的Excel文件: {upload_path}")
        except Exception as save_error:
            logger.error(f"保存上传文件失败: {str(save_error)}")
            return jsonify({'status': 'error', 'message': f'保存文件失败: {str(save_error)}'}), 500
        
        # 使用CommentsExcelProcessor处理Excel文件
        try:
            from utils.excel_processor import CommentsExcelProcessor
            
            # 获取当前班级ID
            if not hasattr(current_user, 'class_id') or not current_user.class_id:
                return jsonify({'status': 'error', 'message': '您没有关联的班级，无法导入评语'}), 403
                
            class_id = current_user.class_id
            
            # 创建处理器实例
            processor = CommentsExcelProcessor()
            
            # 处理Excel文件
            result = processor.process_file(upload_path, class_id)
            if 'error' in result:
                # 删除已上传的文件
                try:
                    if os.path.exists(upload_path):
                        os.remove(upload_path)
                except:
                    pass
                
                return jsonify({'status': 'error', 'message': result['error']}), 400
            
            # 获取当前班级学生
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                
                # 查询当前班级所有学生
                cursor.execute('SELECT id, name FROM students WHERE class_id = ?', (class_id,))
                students = cursor.fetchall()
                
                # 转换为字典，方便查找
                students_dict = {student['name']: {'id': student['id']} for student in students}
                conn.close()
                
                logger.info(f"从数据库获取学生信息成功，班级ID: {class_id}, 学生数量: {len(students_dict)}")
            except Exception as db_error:
                logger.error(f"数据库查询失败: {str(db_error)}")
                return jsonify({'status': 'error', 'message': f'数据库查询失败: {str(db_error)}'}), 500
            
            # 将评语与学生匹配
            match_result = processor.match_students_with_comments(result['comments'], students_dict)
            
            # 返回预览结果
            return jsonify({
                'status': 'ok',
                'file_path': upload_path,
                'previews': match_result['previews'],
                'total_count': match_result['total_count'],
                'match_count': match_result['match_count'],
                'valid_count': match_result['valid_count'],
                'all_valid': match_result['all_valid']
            })
            
        except ImportError:
            logger.error("CommentsExcelProcessor模块不可用")
            
            # 回退到原始代码处理Excel文件
            return process_excel_file_legacy(upload_path, class_id=current_user.class_id)
            
    except Exception as e:
        logger.error(f"预览评语导入时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'status': 'error', 'message': f'服务器错误: {str(e)}'}), 500

# 原始Excel处理函数，用作回退
def process_excel_file_legacy(file_path, class_id=None):
    """原始Excel文件处理函数，当新处理器不可用时回退使用"""
    
    try:
        import openpyxl
        
        # 验证是否确实是一个Excel文件
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            logger.info(f"成功读取Excel文件，工作表名称: {ws.title}")
        except Exception as wb_error:
            # 记录详细的解析错误
            logger.error(f"无法解析Excel文件: {str(wb_error)}")
            
            # 尝试删除已保存的文件
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except:
                pass
            
            return jsonify({
                'status': 'error', 
                'message': f'无法解析Excel文件，请确保是标准的.xlsx格式。错误信息: {str(wb_error)}'
            }), 400
    except ImportError:
        logger.error("openpyxl模块未安装")
        return jsonify({'status': 'error', 'message': 'openpyxl模块未安装，请联系管理员'}), 500
    
    # 检查必要的列是否存在
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    if '姓名' not in headers or '评语' not in headers:
        # 删除已上传的文件
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except:
            pass
            
        return jsonify({
            'status': 'error', 
            'message': f'Excel文件格式错误：缺少必要的列"姓名"和"评语"。请下载并使用标准模板。当前表头: {", ".join(headers)}'
        }), 400
    
    logger.info(f"Excel文件校验成功，表头: {headers}")
    
    # 获取姓名和评语列的索引
    name_index = headers.index('姓名')
    comment_index = headers.index('评语')
    
    # 获取当前班级学生
    try:
        # 获取班级ID
        if class_id is None:
            if not hasattr(current_user, 'class_id') or not current_user.class_id:
                return jsonify({'status': 'error', 'message': '您没有关联的班级，无法导入评语'}), 403
            class_id = current_user.class_id
        
        # 获取数据库连接
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 查询当前班级所有学生
        cursor.execute('SELECT id, name FROM students WHERE class_id = ?', (class_id,))
        students = cursor.fetchall()
        
        # 转换为字典，方便查找
        students_dict = {student['name']: student['id'] for student in students}
        conn.close()
        
        logger.info(f"从数据库获取学生信息成功，班级ID: {class_id}, 学生数量: {len(students_dict)}")
    except Exception as db_error:
        logger.error(f"数据库查询失败: {str(db_error)}")
        return jsonify({'status': 'error', 'message': f'数据库查询失败: {str(db_error)}'}), 500
    
    # 解析Excel数据
    previews = []
    row_index = 0
    match_count = 0
    total_count = 0
    
    # 获取所有数据行
    data_rows = list(ws.iter_rows(min_row=2))  # 跳过表头
    
    # 遍历数据行
    for row in data_rows:
        row_index += 1
        if row_index > 30:  # 限制预览行数为30行
            break
            
        # 获取姓名和评语，确保name_index和comment_index在有效范围内
        name = row[name_index].value if name_index < len(row) and row[name_index].value else ""
        comment = row[comment_index].value if comment_index < len(row) and row[comment_index].value else ""
        
        # 确保名称为字符串类型
        if name is not None and not isinstance(name, str):
            name = str(name)
            
        # 确保评语为字符串类型
        if comment is not None and not isinstance(comment, str):
            comment = str(comment)
        
        if not name or not comment:
            continue
        
        total_count += 1
        # 检查是否匹配到学生
        matched = name in students_dict
        if matched:
            match_count += 1
            
        # 添加到预览数据
        previews.append({
            'name': name,
            'comment': comment if len(comment) <= 5000 else comment[:5000] + '...(已截断)',
            'matched': matched
        })
    
    # 统计Excel中的总行数（不包括表头）
    valid_rows_count = 0
    for row in data_rows:
        name = row[name_index].value if name_index < len(row) else None
        comment = row[comment_index].value if comment_index < len(row) else None
        if name and comment:
            valid_rows_count += 1
    
    logger.info(f"Excel文件解析完成，总行数: {valid_rows_count}，匹配学生数: {match_count}")
            
    return jsonify({
        'status': 'ok',
        'file_path': file_path,
        'previews': previews,
        'total_count': valid_rows_count,
        'match_count': match_count
    })

# 确认导入评语API
@comments_bp.route('/api/confirm-import-comments', methods=['POST'])
def confirm_import_comments():
    """确认导入评语数据"""
    
    try:
        # 检查用户是否已登录
        if not hasattr(current_user, 'is_authenticated') or not current_user.is_authenticated:
            return jsonify({'status': 'error', 'message': '您需要登录后才能导入评语', 'code': 'login_required'}), 401
            
        data = request.get_json()
        
        # 验证请求数据
        if not data or 'file_path' not in data:
            logger.error("确认导入请求缺少file_path参数")
            return jsonify({'status': 'error', 'message': '缺少必要的参数'}), 400
        
        file_path = data['file_path']
        append_mode = data.get('append_mode', False)  # 默认使用替换模式
        
        # 详细记录请求信息，帮助调试
        logger.info(f"确认导入请求: 文件路径={file_path}, 追加模式={append_mode}")
        
        # 检查文件是否存在
        if not os.path.exists(file_path):
            logger.error(f"导入文件不存在: {file_path}")
            return jsonify({'status': 'error', 'message': '找不到上传的文件，可能已被删除或过期'}), 400
        
        # 检查文件扩展名，如果缺少，则添加临时副本
        base_name, ext = os.path.splitext(file_path)
        temp_file_path = file_path
        
        if not ext or ext.lower() != '.xlsx':
            # 创建带扩展名的临时文件
            temp_file_path = f"{file_path}.xlsx"
            try:
                import shutil
                shutil.copy2(file_path, temp_file_path)
                logger.info(f"为确保文件格式识别，创建了带扩展名的临时文件副本: {temp_file_path}")
            except Exception as copy_error:
                logger.warning(f"创建临时文件副本失败: {str(copy_error)}，将继续尝试使用原文件")
                temp_file_path = file_path  # 回退到原始文件

        # 尝试验证Excel文件
        try:
            import openpyxl
            
            logger.info(f"尝试打开文件验证Excel格式: {temp_file_path}")
            wb = openpyxl.load_workbook(temp_file_path)
            ws = wb.active
            logger.info(f"成功验证Excel文件: {temp_file_path}, 工作表: {ws.title}")
        except Exception as excel_error:
            logger.error(f"文件不是有效的Excel文件: {temp_file_path}, 错误: {str(excel_error)}")
            
            # 清理临时文件
            if temp_file_path != file_path and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                    logger.info(f"已删除临时文件: {temp_file_path}")
                except:
                    pass
                    
            return jsonify({'status': 'error', 'message': f'无法解析Excel文件: {str(excel_error)}'}), 400

        # 获取当前班级ID
        class_id = current_user.class_id
        
        # 使用CommentsExcelProcessor处理Excel文件
        try:
            from utils.excel_processor import CommentsExcelProcessor
            
            # 创建处理器实例
            processor = CommentsExcelProcessor()
            
            # 处理Excel文件 - 使用带扩展名的临时文件
            result = processor.process_file(temp_file_path, class_id)
            
            if 'error' in result:
                logger.error(f"处理Excel文件失败: {result['error']}")
                
                # 清理临时文件
                if temp_file_path != file_path and os.path.exists(temp_file_path):
                    try:
                        os.remove(temp_file_path)
                    except:
                        pass
                        
                return jsonify({'status': 'error', 'message': result['error']}), 400
                
            # 获取当前班级学生
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # 查询当前班级所有学生
            cursor.execute('SELECT id, name, comments FROM students WHERE class_id = ?', (class_id,))
            students = cursor.fetchall()
            
            # 转换为字典，方便查找
            students_dict = {student['name']: {
                'id': student['id'],
                'comments': student['comments']
            } for student in students}
            
            logger.info(f"从数据库获取学生: {len(students_dict)}名")
            
            # 匹配评语和学生
            logger.info(f"匹配评语和学生...")
            match_result = processor.match_students_with_comments(result['comments'], students_dict)
            previews = match_result['previews']
            logger.info(f"评语匹配结果: 总数={match_result['total_count']}, 匹配={match_result['match_count']}, 有效评语={match_result['valid_count']}")
            
            # 执行导入
            success_count = 0
            error_count = 0
            error_details = []
            skipped_count = 0
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # 开始事务
            conn.execute('BEGIN TRANSACTION')
            
            for preview in previews:
                # 只导入匹配的且有效的评语（不超过5000字）
                if preview['matched'] and preview['valid']:
                    student_name = preview['name']
                    student_id = students_dict[student_name]['id']
                    comment = preview['comment']
                    existing_comment = students_dict[student_name]['comments']
                    
                    # 根据模式更新评语
                    if append_mode and existing_comment:
                        # 追加模式
                        updated_comment = f"{existing_comment}\n\n--- {now} ---\n{comment}"
                    else:
                        # 替换模式
                        updated_comment = comment
                    
                    try:
                        # 更新学生评语
                        cursor.execute(
                            'UPDATE students SET comments = ?, updated_at = ? WHERE id = ? AND class_id = ?',
                            (updated_comment, now, student_id, class_id)
                        )
                        success_count += 1
                    except Exception as update_error:
                        error_count += 1
                        error_detail = f"更新学生 {student_name} 的评语失败: {str(update_error)}"
                        error_details.append(error_detail)
                        logger.error(f"更新学生评语时出错: {error_detail}")
                else:
                    skipped_count += 1
                    # 记录跳过原因 
                    if not preview['matched']:
                        logger.info(f"跳过未匹配学生的评语: {preview['name']}")
                    elif not preview['valid']:
                        logger.info(f"跳过无效评语: {preview['name']}, 长度: {preview['length']}字")
            
            # 提交事务
            conn.commit()
            conn.close()
            
            # 清理临时和原始文件
            try:
                # 先删除临时文件（如果创建了的话）
                if temp_file_path != file_path and os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                    logger.info(f"已删除临时文件: {temp_file_path}")
                    
                # 再删除原始文件
                if os.path.exists(file_path):
                    os.remove(file_path)
                    logger.info(f"已删除原始上传文件: {file_path}")
            except Exception as file_error:
                logger.warning(f"删除文件失败: {str(file_error)}")
            
            # 根据成功和失败的数量确定状态
            status = 'ok' if error_count == 0 else 'partial' if success_count > 0 else 'error'
            
            logger.info(f"评语导入完成: 成功 {success_count} 条, 失败 {error_count} 条, 跳过 {skipped_count} 条")
            
            # 限制错误详情的数量
            if len(error_details) > 10:
                error_details = error_details[:10] + [f"...还有 {len(error_details) - 10} 条错误未显示"]
            
            return jsonify({
                'status': status,
                'success_count': success_count,
                'error_count': error_count,
                'skipped_count': skipped_count,
                'error_details': error_details
            })
            
        except ImportError as import_error:
            logger.warning(f"CommentsExcelProcessor模块不可用: {str(import_error)}")
            logger.warning("回退到原始导入逻辑")
            
            # 清理临时文件
            if temp_file_path != file_path and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except:
                    pass
            
            # 如果CommentsExcelProcessor不可用，使用原始导入逻辑
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # 查询当前班级所有学生
            cursor.execute('SELECT id, name, comments FROM students WHERE class_id = ?', (class_id,))
            students = cursor.fetchall()
            
            # 转换为字典，方便查找
            students_dict = {student['name']: {
                'id': student['id'],
                'comments': student['comments']
            } for student in students}
            
            logger.info(f"已获取班级学生信息，共{len(students_dict)}名学生")
            
            # 使用openpyxl读取Excel文件
            try:
                import openpyxl
                
                try:
                    # 尝试读取临时文件，如果失败则尝试原始文件
                    try:
                        wb = openpyxl.load_workbook(temp_file_path)
                        excel_file_used = temp_file_path
                    except:
                        wb = openpyxl.load_workbook(file_path)
                        excel_file_used = file_path
                        
                    ws = wb.active
                    logger.info(f"导入确认：成功读取Excel文件 {excel_file_used}")
                except Exception as wb_error:
                    conn.close()
                    logger.error(f"导入确认：无法解析Excel文件: {str(wb_error)}")
                    return jsonify({'status': 'error', 'message': f'无法解析Excel文件: {str(wb_error)}'}), 500
            except ImportError as ie:
                conn.close()
                logger.error(f"导入确认：缺少openpyxl模块: {str(ie)}")
                return jsonify({'status': 'error', 'message': 'openpyxl模块未安装，请联系管理员'}), 500
            
            # 获取表头
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            if '姓名' not in headers or '评语' not in headers:
                conn.close()
                logger.error(f"文件格式错误，缺少必要的列，表头: {headers}")
                return jsonify({'status': 'error', 'message': '文件格式错误，缺少必要的列'}), 400
            
            # 获取姓名和评语列的索引
            name_index = headers.index('姓名')
            comment_index = headers.index('评语')
            
            # 开始导入
            success_count = 0
            error_count = 0
            error_details = []
            skipped_count = 0
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # 开始事务
            conn.execute('BEGIN TRANSACTION')
            
            try:
                # 遍历Excel行
                data_rows = list(ws.iter_rows(min_row=2))  # 跳过表头
                logger.info(f"开始处理Excel数据，共{len(data_rows)}行")
                
                for row in data_rows:
                    # 确保索引在有效范围内
                    if name_index >= len(row) or comment_index >= len(row):
                        error_count += 1
                        error_details.append("Excel行数据结构错误，列索引超出范围")
                        continue
                    
                    # 获取姓名和评语
                    name = row[name_index].value
                    comment = row[comment_index].value
                    
                    # 确保名称为字符串类型
                    if name is not None and not isinstance(name, str):
                        name = str(name)
                    
                    # 确保评语为字符串类型
                    if comment is not None and not isinstance(comment, str):
                        comment = str(comment)
                    
                    if not name or not comment:
                        skipped_count += 1
                        continue
                    
                    # 截断评语（如果超过5000字符）
                    if len(comment) > 5000:  # 临时调整为5000字
                        comment = comment[:5000]
                    
                    # 检查是否匹配到学生
                    if name in students_dict:
                        student_info = students_dict[name]
                        student_id = student_info['id']
                        existing_comment = student_info['comments']
                        
                        # 根据模式更新评语
                        if append_mode and existing_comment:
                            # 追加模式
                            updated_comment = f"{existing_comment}\n\n--- {now} ---\n{comment}"
                        else:
                            # 替换模式
                            updated_comment = comment
                        
                        try:
                            # 更新学生评语
                            cursor.execute(
                                'UPDATE students SET comments = ?, updated_at = ? WHERE id = ? AND class_id = ?',
                                (updated_comment, now, student_id, class_id)
                            )
                            success_count += 1
                        except Exception as update_error:
                            error_count += 1
                            error_detail = f"更新学生 {name} 的评语失败: {str(update_error)}"
                            error_details.append(error_detail)
                            logger.error(f"更新学生评语时出错: {error_detail}")
                    else:
                        error_count += 1
                        error_details.append(f"找不到匹配的学生: {name}")
                
                # 提交事务
                conn.commit()
                
                # 关闭数据库连接
                conn.close()
                
                # 清理文件
                try:
                    # 先删除临时文件（如果创建了的话）
                    if temp_file_path != file_path and os.path.exists(temp_file_path):
                        os.remove(temp_file_path)
                        logger.info(f"已删除临时文件: {temp_file_path}")
                        
                    # 再删除原始文件
                    if os.path.exists(file_path):
                        os.remove(file_path)
                        logger.info(f"已删除原始上传文件: {file_path}")
                except Exception as file_error:
                    logger.warning(f"删除文件失败: {str(file_error)}")
                
                # 生成响应
                status = 'ok' if error_count == 0 else 'partial' if success_count > 0 else 'error'
                
                logger.info(f"导入完成: 成功 {success_count} 条, 失败 {error_count} 条, 跳过 {skipped_count} 条")
                
                # 限制错误详情的数量，避免响应太大
                if len(error_details) > 10:
                    error_details = error_details[:10] + [f"...还有 {len(error_details) - 10} 条错误未显示"]
                
                return jsonify({
                    'status': status,
                    'success_count': success_count,
                    'error_count': error_count,
                    'skipped_count': skipped_count,
                    'error_details': error_details
                })
                
            except Exception as e:
                # 回滚事务
                conn.rollback()
                conn.close()
                
                # 清理文件
                try:
                    if temp_file_path != file_path and os.path.exists(temp_file_path):
                        os.remove(temp_file_path)
                except:
                    pass
                
                logger.error(f"导入评语时出错: {str(e)}")
                logger.error(traceback.format_exc())
                return jsonify({
                    'status': 'error',
                    'message': f'导入评语时出错: {str(e)}'
                }), 500
        
    except Exception as e:
        logger.error(f"确认导入评语时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'status': 'error', 'message': f'服务器错误: {str(e)}'}), 500

# 导出评语为Excel
@comments_bp.route('/api/export-comments-excel', methods=['GET'])
def export_comments_excel():
    """导出班级学生评语为Excel文件"""
    try:
        logger.info("收到导出评语Excel请求")
        
        # 检查用户是否已登录
        if not hasattr(current_user, 'is_authenticated') or not current_user.is_authenticated:
            logger.error("用户未登录，无法导出评语")
            return jsonify({'status': 'error', 'message': '您需要登录后才能导出评语', 'code': 'login_required'}), 401
        
        # 获取班级ID，优先使用请求参数中的class_id，如果没有则使用当前用户的class_id
        class_id = request.args.get('class_id')
        if not class_id and hasattr(current_user, 'class_id'):
            class_id = current_user.class_id
            logger.info(f"使用当前用户班级ID: {class_id}")
        
        # 如果没有班级ID，返回错误
        if not class_id:
            logger.error("未指定班级ID")
            return jsonify({'status': 'error', 'message': '未指定班级ID'}), 400
            
        logger.info(f"导出评语Excel: 班级ID={class_id}")
        
        # 获取班级学生数据
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 查询班级名称 - 使用正确的列名class_name
        cursor.execute('SELECT class_name FROM classes WHERE id = ?', (class_id,))
        class_result = cursor.fetchone()
        class_name = class_result['class_name'] if class_result else f"班级{class_id}"
        logger.info(f"班级名称: {class_name}")
        
        # 查询学生数据
        cursor.execute('SELECT id, name, comments FROM students WHERE class_id = ? ORDER BY CAST(id AS INTEGER)', (class_id,))
        students = cursor.fetchall()
        conn.close()
        
        if not students:
            logger.error(f"未找到班级{class_id}的学生数据")
            return jsonify({'status': 'error', 'message': '未找到班级学生数据'}), 404
        
        logger.info(f"找到{len(students)}名学生的数据")
        
        try:
            # 尝试导入pandas
            import pandas as pd
            from io import BytesIO
            import urllib.parse
        except ImportError as e:
            logger.error(f"缺少必要的库，无法导出Excel: {str(e)}")
            return jsonify({'status': 'error', 'message': f'服务器缺少必要的库，无法导出Excel: {str(e)}'}), 500
        
        # 创建DataFrame
        data = []
        for student in students:
            data.append({
                '姓名': student['name'],
                '评语': student['comments'] if student['comments'] else ''
            })
        
        df = pd.DataFrame(data)
        
        # 创建Excel文件
        output = BytesIO()
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='评语数据', index=False)
                
                # 获取工作表以设置列宽
                worksheet = writer.sheets['评语数据']
                worksheet.column_dimensions['A'].width = 15  # 姓名列宽
                worksheet.column_dimensions['B'].width = 60  # 评语列宽
        except Exception as excel_error:
            logger.error(f"创建Excel文件失败: {str(excel_error)}")
            return jsonify({'status': 'error', 'message': f'创建Excel文件失败: {str(excel_error)}'}), 500
        
        # 设置响应头
        output.seek(0)
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        
        # 生成文件名并进行URL编码
        raw_filename = f"{class_name}_评语导出_{timestamp}.xlsx"
        encoded_filename = urllib.parse.quote(raw_filename)
        
        # 创建响应
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
        
        logger.info(f"成功导出评语Excel: 班级={class_name}, 学生数量={len(students)}")
        return response
        
    except Exception as e:
        logger.error(f"导出评语Excel时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'status': 'error',
            'message': f'导出评语Excel时出错: {str(e)}'
        }), 500
